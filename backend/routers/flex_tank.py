import os
from datetime import datetime, timezone, timedelta
from typing import Optional, List
import io
import re
import json
import shutil
import uuid
import logging
from pathlib import Path
from urllib.parse import quote as url_quote

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel as PydanticBaseModel

from models import (
    User, UserCreate, UserLogin, UserResponse, Token,
    CompanySettings, CompanySettingsUpdate,
    Driver, DriverCreate, DriverResponse,
    TransportCompany, TransportCompanyCreate, TransportCompanyResponse,
    Client, ClientCreate, ClientResponse,
    Supplier, SupplierCreate, SupplierResponse,
    ContainerMovement, ContainerMovementCreate, ContainerMovementResponse,
    DailyMovementPoint, DriverRankingEntry, DashboardStats,
    ShippingLine, ShippingLineCreate, ShippingLineResponse,
    ServiceType, ServiceTypeCreate, ServiceTypeResponse,
    Invoice, InvoiceCreate, InvoiceUpdate, InvoiceResponse, InvoiceMovementDetail,
    InvoiceHistory, InvoiceHistoryResponse,
    PhotoRegistry, PhotoRegistryCreate, PhotoRegistryUpdate, PhotoRegistryResponse,
    ContainerInspectionPhoto, ContainerInspection, ContainerInspectionCreate,
    ContainerInspectionUpdate, ContainerInspectionResponse,
    CONTAINER_INSPECTION_PHOTO_TYPES, MAX_CONTAINER_INSPECTION_PHOTOS,
    FlexTankMovement, FlexTankMovementCreate, FlexTankMovementUpdate,
    FlexTankMovementResponse, FlexTankStockSummary,
    Vehicle, VehicleCreate, VehicleUpdate, VehicleResponse,
    VehicleChecklistItem, VehicleChecklistProduct, VehicleChecklistFields,
    VehicleChecklist, VehicleChecklistCreate, VehicleChecklistResponse,
    VEHICLE_CHECKLIST_TEMPLATE, VEHICLE_CHECKLIST_SECTION_LABELS,
    VehicleRevision, VehicleRevisionCreate, VehicleRevisionResponse,
    LoadingScheduleItem, LoadingSchedule, LoadingScheduleCreate, LoadingScheduleResponse,
    DailyRateRequestItem, DailyRateRequest, DailyRateRequestCreate, DailyRateRequestResponse,
    IntlInvoiceItem, IntlInvoice, IntlInvoiceCreate, IntlInvoiceResponse,
    DeliveryStatusItem, DeliveryStatus, DeliveryStatusCreate, DeliveryStatusResponse,
    UnitSegregationItem, UnitSegregation, UnitSegregationCreate, UnitSegregationUpdate,
    UnitSegregationResponse,
    RPAServiceItem, RPATerceiro, RPATerceiroCreate, RPATerceiroUpdate, RPATerceiroResponse,
    OSItem, OrdemServico, OrdemServicoCreate, OrdemServicoUpdate, OrdemServicoResponse,
    ExpenseReportReceipt, ExpenseReportDeposit, ExpenseReportPurchase,
    ExpenseReport, ExpenseReportCreate, ExpenseReportResponse,
)
from auth import get_password_hash, verify_password, create_access_token, get_current_user, decode_token
from reports import (
    generate_pdf_report, generate_excel_report, generate_billing_pdf_report, generate_billing_excel,
    now_brt, to_brt, merge_company, DEFAULT_COMPANY
)

from shared import (
    db, manager, get_current_active_user, get_current_admin_user, get_company_settings,
    get_next_transaction_id, parse_datetime_value, round_money, migrate_inspection_photos,
    load_logo_buffer, validate_and_read_upload, ALLOWED_EXTENSIONS, ALLOWED_RECEIPT_EXTENSIONS,
    MAX_FILE_SIZE, check_rate_limit, client_ip, UPLOADS_DIR, ROOT_DIR
)

api_router = APIRouter(prefix="/api")

# ========== FLEX TANK (CONTROLE DE ESTOQUE DE BOLSAS) ==========

@api_router.get("/flex-tank/movements")
async def list_flex_tank_movements(
    page: int = 1,
    per_page: int = 20,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    client_id: Optional[str] = None,
    movement_number: Optional[int] = None,
    movement_type: Optional[str] = None,
    current_user: dict = Depends(get_current_active_user)
):
    """Lista todas as movimentações de Flex Tank com filtros"""
    query = {}
    
    # Filtro por data
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter["$gte"] = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        if end_date:
            date_filter["$lte"] = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        if date_filter:
            query["movement_date"] = date_filter
    
    # Filtro por cliente
    if client_id:
        query["client_id"] = client_id
    
    # Filtro por número de registro
    if movement_number:
        query["movement_number"] = movement_number
    
    # Filtro por tipo
    if movement_type:
        query["movement_type"] = movement_type
    
    skip = (page - 1) * per_page
    
    total = await db.flex_tank_movements.count_documents(query)
    movements = await db.flex_tank_movements.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page).to_list(per_page)
    
    return {
        "items": movements,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": (total + per_page - 1) // per_page
    }

@api_router.get("/flex-tank/stock")
async def get_flex_tank_stock(
    client_id: Optional[str] = None,
    current_user: dict = Depends(get_current_active_user)
):
    """Retorna o resumo do estoque de Flex Tank"""
    query = {}
    if client_id:
        query["client_id"] = client_id
    
    # Buscar todas as movimentações
    movements = await db.flex_tank_movements.find(query, {"_id": 0}).to_list(None)
    
    # Calcular totais
    total_entries = sum(1 for m in movements if m.get("movement_type") == "ENTRADA")
    total_exits = sum(1 for m in movements if m.get("movement_type") == "SAIDA")
    total_bags = total_entries - total_exits
    
    # Agrupar por cliente
    client_stock = {}
    for m in movements:
        cid = m.get("client_id") or "sem_cliente"
        cname = m.get("client_name") or "Sem Cliente"
        if cid not in client_stock:
            client_stock[cid] = {"client_id": cid, "client_name": cname, "entries": 0, "exits": 0, "stock": 0}
        if m.get("movement_type") == "ENTRADA":
            client_stock[cid]["entries"] += 1
        else:
            client_stock[cid]["exits"] += 1
        client_stock[cid]["stock"] = client_stock[cid]["entries"] - client_stock[cid]["exits"]
    
    # Agrupar por tamanho
    size_stock = {}
    for m in movements:
        size = m.get("bag_size") or "Não informado"
        if size not in size_stock:
            size_stock[size] = {"size": size, "entries": 0, "exits": 0, "stock": 0}
        if m.get("movement_type") == "ENTRADA":
            size_stock[size]["entries"] += 1
        else:
            size_stock[size]["exits"] += 1
        size_stock[size]["stock"] = size_stock[size]["entries"] - size_stock[size]["exits"]
    
    return {
        "total_bags": total_bags,
        "total_entries": total_entries,
        "total_exits": total_exits,
        "by_client": list(client_stock.values()),
        "by_size": list(size_stock.values())
    }

@api_router.get("/flex-tank/movements/{movement_id}", response_model=FlexTankMovementResponse)
async def get_flex_tank_movement(movement_id: str, current_user: dict = Depends(get_current_active_user)):
    """Obtém uma movimentação de Flex Tank pelo ID"""
    movement = await db.flex_tank_movements.find_one({"id": movement_id}, {"_id": 0})
    if not movement:
        raise HTTPException(status_code=404, detail="Movimentação não encontrada")
    return FlexTankMovementResponse(**movement)

@api_router.post("/flex-tank/movements", response_model=FlexTankMovementResponse)
async def create_flex_tank_movement(
    data: FlexTankMovementCreate,
    current_user: dict = Depends(get_current_active_user)
):
    """Cria uma nova movimentação de Flex Tank"""
    # Gerar número sequencial
    counter = await db.counters.find_one_and_update(
        {"_id": "flex_tank_movement_number"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    movement_number = counter["seq"]
    
    # Buscar nome do cliente
    client_name = None
    if data.client_id:
        client = await db.clients.find_one({"id": data.client_id}, {"_id": 0, "name": 1})
        client_name = client.get("name") if client else None
    
    # Buscar nome do cliente destino
    destination_client_name = None
    if data.destination_client_id:
        dest_client = await db.clients.find_one({"id": data.destination_client_id}, {"_id": 0, "name": 1})
        destination_client_name = dest_client.get("name") if dest_client else None
    
    movement = FlexTankMovement(
        movement_number=movement_number,
        bag_number=data.bag_number,
        bag_size=data.bag_size,
        movement_date=data.movement_date,
        movement_type=data.movement_type,
        client_id=data.client_id,
        client_name=client_name,
        destination_client_id=data.destination_client_id,
        destination_client_name=destination_client_name,
        container_number=data.container_number,
        observations=data.observations,
        created_by=current_user["sub"],
        created_by_name=current_user["name"]
    )
    
    movement_dict = movement.model_dump()
    movement_dict["created_at"] = movement_dict["created_at"].isoformat()
    movement_dict["movement_date"] = movement_dict["movement_date"].isoformat()
    
    await db.flex_tank_movements.insert_one(movement_dict)
    
    return FlexTankMovementResponse(**movement_dict)

@api_router.put("/flex-tank/movements/{movement_id}", response_model=FlexTankMovementResponse)
async def update_flex_tank_movement(
    movement_id: str,
    data: FlexTankMovementUpdate,
    current_user: dict = Depends(get_current_active_user)
):
    """Atualiza uma movimentação de Flex Tank"""
    movement = await db.flex_tank_movements.find_one({"id": movement_id}, {"_id": 0})
    if not movement:
        raise HTTPException(status_code=404, detail="Movimentação não encontrada")
    
    update_data = {}
    
    if data.bag_number is not None:
        update_data["bag_number"] = data.bag_number
    
    if data.bag_size is not None:
        update_data["bag_size"] = data.bag_size
    
    if data.movement_date is not None:
        update_data["movement_date"] = data.movement_date.isoformat()
    
    if data.movement_type is not None:
        update_data["movement_type"] = data.movement_type
    
    if data.container_number is not None:
        update_data["container_number"] = data.container_number
    
    if data.observations is not None:
        update_data["observations"] = data.observations
    
    if data.client_id is not None:
        update_data["client_id"] = data.client_id
        if data.client_id:
            client = await db.clients.find_one({"id": data.client_id}, {"_id": 0, "name": 1})
            update_data["client_name"] = client.get("name") if client else None
        else:
            update_data["client_name"] = None
    
    if data.destination_client_id is not None:
        update_data["destination_client_id"] = data.destination_client_id
        if data.destination_client_id:
            dest_client = await db.clients.find_one({"id": data.destination_client_id}, {"_id": 0, "name": 1})
            update_data["destination_client_name"] = dest_client.get("name") if dest_client else None
        else:
            update_data["destination_client_name"] = None
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.flex_tank_movements.update_one({"id": movement_id}, {"$set": update_data})
    
    updated = await db.flex_tank_movements.find_one({"id": movement_id}, {"_id": 0})
    return FlexTankMovementResponse(**updated)

@api_router.delete("/flex-tank/movements/{movement_id}")
async def delete_flex_tank_movement(
    movement_id: str,
    current_user: dict = Depends(get_current_active_user)
):
    """Exclui uma movimentação de Flex Tank"""
    movement = await db.flex_tank_movements.find_one({"id": movement_id}, {"_id": 0})
    if not movement:
        raise HTTPException(status_code=404, detail="Movimentação não encontrada")
    
    await db.flex_tank_movements.delete_one({"id": movement_id})
    
    return {"message": "Movimentação excluída com sucesso"}

@api_router.get("/flex-tank/report/excel")
async def download_flex_tank_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    client_id: Optional[str] = None,
    movement_type: Optional[str] = None,
    current_user: dict = Depends(get_current_active_user)
):
    """Gera relatório Excel das movimentações de Flex Tank"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    query = {}
    
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter["$gte"] = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        if end_date:
            date_filter["$lte"] = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        if date_filter:
            query["movement_date"] = date_filter
    
    if client_id:
        query["client_id"] = client_id
    
    if movement_type:
        query["movement_type"] = movement_type
    
    movements = await db.flex_tank_movements.find(query, {"_id": 0}).sort("movement_date", -1).to_list(None)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimentações Flex Tank"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    entrada_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    saida_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Cabeçalhos - adicionado Cliente Destino
    headers = ["Nº Registro", "Nº Bolsa", "Tamanho", "Data", "Tipo", "Cliente", "Cliente Destino", "Container", "Observações"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Congelar primeira linha
    ws.freeze_panes = 'A2'
    
    # Dados
    for row, m in enumerate(movements, 2):
        # Nº Registro
        cell = ws.cell(row=row, column=1, value=m.get("movement_number"))
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        
        # Nº Bolsa
        cell = ws.cell(row=row, column=2, value=m.get("bag_number"))
        cell.border = thin_border
        
        # Tamanho
        cell = ws.cell(row=row, column=3, value=m.get("bag_size"))
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        
        # Data
        movement_date = m.get("movement_date")
        if isinstance(movement_date, str):
            movement_date = datetime.fromisoformat(movement_date.replace('Z', '+00:00'))
        cell = ws.cell(row=row, column=4, value=movement_date.strftime("%d/%m/%Y") if movement_date else "")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        
        # Tipo (com formatação condicional)
        tipo = m.get("movement_type")
        cell = ws.cell(row=row, column=5, value=tipo)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if tipo == "ENTRADA":
            cell.fill = entrada_fill
        elif tipo == "SAIDA":
            cell.fill = saida_fill
        
        # Cliente
        cell = ws.cell(row=row, column=6, value=m.get("client_name") or "-")
        cell.border = thin_border
        
        # Cliente Destino
        cell = ws.cell(row=row, column=7, value=m.get("destination_client_name") or "-")
        cell.border = thin_border
        
        # Container
        cell = ws.cell(row=row, column=8, value=m.get("container_number") or "-")
        cell.border = thin_border
        
        # Observações
        cell = ws.cell(row=row, column=9, value=m.get("observations") or "-")
        cell.border = thin_border
        
        # Linha alternada
        if row % 2 == 0:
            for col in [1, 2, 3, 4, 6, 7, 8, 9]:  # Não aplicar na coluna Tipo (5)
                ws.cell(row=row, column=col).fill = alt_fill
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 35
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 35
    
    # Salvar em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"relatorio_flex_tank_{now_brt().strftime('%d-%m-%Y_%H-%M')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

