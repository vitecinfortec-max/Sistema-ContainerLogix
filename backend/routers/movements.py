import asyncio
import os
from datetime import datetime, timezone, timedelta
from typing import Optional, List, Literal
import io
import re
import json
import shutil
import uuid
import logging
from pathlib import Path
from urllib.parse import quote as url_quote

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel as PydanticBaseModel

from models import (
    User, UserCreate, UserLogin, UserResponse, Token,
    CompanySettings, CompanySettingsUpdate,
    Driver, DriverCreate, DriverResponse,
    TransportCompany, TransportCompanyCreate, TransportCompanyResponse,
    Client, ClientCreate, ClientResponse,
    Supplier, SupplierCreate, SupplierResponse,
    ContainerMovement, ContainerMovementCreate, ContainerMovementResponse,
    DailyMovementPoint, DailyBillingPoint, DriverRankingEntry, DashboardStats,
    ShippingLine, ShippingLineCreate, ShippingLineResponse,
    ServiceType, ServiceTypeCreate, ServiceTypeResponse,
    Invoice, InvoiceCreate, InvoiceUpdate, InvoiceResponse, InvoiceMovementDetail,
    InvoiceHistory, InvoiceHistoryResponse,
    PhotoRegistry, PhotoRegistryCreate, PhotoRegistryUpdate, PhotoRegistryResponse,
    ContainerInspectionPhoto, ContainerInspection, ContainerInspectionCreate,
    ContainerInspectionUpdate, ContainerInspectionResponse,
    CONTAINER_INSPECTION_PHOTO_TYPES, MAX_CONTAINER_INSPECTION_PHOTOS,
    FlexTankMovement, FlexTankMovementCreate, FlexTankMovementUpdate,
    FlexTankMovementResponse, FlexTankStockSummary,
    Vehicle, VehicleCreate, VehicleUpdate, VehicleResponse,
    VehicleChecklistItem, VehicleChecklistProduct, VehicleChecklistFields,
    VehicleChecklist, VehicleChecklistCreate, VehicleChecklistResponse,
    VEHICLE_CHECKLIST_TEMPLATE, VEHICLE_CHECKLIST_SECTION_LABELS,
    VehicleRevision, VehicleRevisionCreate, VehicleRevisionResponse,
    LoadingScheduleItem, LoadingSchedule, LoadingScheduleCreate, LoadingScheduleResponse,
    DailyRateRequestItem, DailyRateRequest, DailyRateRequestCreate, DailyRateRequestResponse,
    IntlInvoiceItem, IntlInvoice, IntlInvoiceCreate, IntlInvoiceResponse,
    DeliveryStatusItem, DeliveryStatus, DeliveryStatusCreate, DeliveryStatusResponse,
    UnitSegregationItem, UnitSegregation, UnitSegregationCreate, UnitSegregationUpdate,
    UnitSegregationResponse,
    RPAServiceItem, RPATerceiro, RPATerceiroCreate, RPATerceiroUpdate, RPATerceiroResponse,
    OSItem, OrdemServico, OrdemServicoCreate, OrdemServicoUpdate, OrdemServicoResponse,
    ExpenseReportReceipt, ExpenseReportDeposit, ExpenseReportPurchase,
    ExpenseReport, ExpenseReportCreate, ExpenseReportResponse,
)
from auth import get_password_hash, verify_password, create_access_token, get_current_user, decode_token
from reports import (
    generate_pdf_report, generate_excel_report, generate_billing_pdf_report, generate_billing_excel,
    generate_movement_voucher_pdf, generate_yard_control_pdf,
    now_brt, to_brt, merge_company, DEFAULT_COMPANY
)

from shared import (
    db, manager, get_current_active_user, get_current_admin_user, get_company_settings,
    get_next_transaction_id, parse_datetime_value, round_money, migrate_inspection_photos,
    load_logo_buffer, validate_and_read_upload, ALLOWED_EXTENSIONS, ALLOWED_RECEIPT_EXTENSIONS,
    MAX_FILE_SIZE, check_rate_limit, client_ip, UPLOADS_DIR, ROOT_DIR
)

api_router = APIRouter(prefix="/api")

@api_router.post("/movements", response_model=ContainerMovementResponse)
async def create_movement(movement_input: ContainerMovementCreate, current_user: dict = Depends(get_current_active_user)):
    # Usar contador atômico para garantir sequência única
    next_transaction_id = await get_next_transaction_id()
    
    movement = ContainerMovement(
        transaction_id=next_transaction_id,
        operation_type=movement_input.operation_type,
        driver_name=movement_input.driver_name,
        driver_cpf=movement_input.driver_cpf,
        truck_plate=movement_input.truck_plate,
        trailer_plate_1=movement_input.trailer_plate_1,
        trailer_plate_2=movement_input.trailer_plate_2,
        transport_company=movement_input.transport_company,
        client_name=movement_input.client_name,
        container_number=movement_input.container_number,
        status=movement_input.status,
        size_type=movement_input.size_type,
        tare=movement_input.tare,
        shipping_line=movement_input.shipping_line,
        seal=movement_input.seal,
        genset=movement_input.genset,
        booking=movement_input.booking,
        origin_terminal=movement_input.origin_terminal,
        service_type=movement_input.service_type,
        invoice_number=movement_input.invoice_number,
        service_value=round_money(movement_input.service_value),
        observations=movement_input.observations,
        container_photos=movement_input.container_photos,
        container_damages=movement_input.container_damages,
        inspection_notes=movement_input.inspection_notes,
        created_by=current_user['sub'],
        user_name=current_user['name']
    )
    
    doc = movement.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    if doc.get('billed_at'):
        doc['billed_at'] = doc['billed_at'].isoformat()
    await db.movements.insert_one(doc)
    
    response = ContainerMovementResponse(
        id=movement.id,
        transaction_id=movement.transaction_id,
        operation_type=movement.operation_type,
        driver_name=movement.driver_name,
        driver_cpf=movement.driver_cpf,
        truck_plate=movement.truck_plate,
        trailer_plate_1=movement.trailer_plate_1,
        trailer_plate_2=movement.trailer_plate_2,
        transport_company=movement.transport_company,
        client_name=movement.client_name,
        container_number=movement.container_number,
        status=movement.status,
        size_type=movement.size_type,
        tare=movement.tare,
        shipping_line=movement.shipping_line,
        seal=movement.seal,
        genset=movement.genset,
        booking=movement.booking,
        origin_terminal=movement.origin_terminal,
        service_type=movement.service_type,
        invoice_number=movement.invoice_number,
        service_value=movement.service_value,
        observations=movement.observations,
        container_photos=movement.container_photos,
        container_damages=movement.container_damages,
        inspection_notes=movement.inspection_notes,
        billed=movement.billed,
        billed_at=movement.billed_at,
        created_at=movement.created_at,
        user_name=movement.user_name
    )
    
    # Notificar todos os clientes conectados via WebSocket
    await manager.broadcast({
        "type": "MOVEMENT_CREATED",
        "data": {
            "id": response.id,
            "transaction_id": response.transaction_id,
            "operation_type": response.operation_type,
            "driver_name": response.driver_name,
            "driver_cpf": response.driver_cpf,
            "truck_plate": response.truck_plate,
            "trailer_plate_1": response.trailer_plate_1,
            "trailer_plate_2": response.trailer_plate_2,
            "transport_company": response.transport_company,
            "container_number": response.container_number,
            "status": response.status,
            "size_type": response.size_type,
            "tare": response.tare,
            "shipping_line": response.shipping_line,
            "seal": response.seal,
            "genset": response.genset,
            "booking": response.booking,
            "created_at": response.created_at.isoformat(),
            "user_name": response.user_name
        }
    })
    
    return response

class MovementPDFRequest(PydanticBaseModel):
    movement_ids: List[str]
    via: Literal["TERMINAL", "MOTORISTA"]

@api_router.post("/movements/pdf")
async def download_movements_pdf(request: MovementPDFRequest, current_user: dict = Depends(get_current_active_user)):
    if not request.movement_ids:
        raise HTTPException(status_code=400, detail="Selecione pelo menos um registro")

    docs = await db.movements.find({"id": {"$in": request.movement_ids}}, {"_id": 0}).to_list(None)
    if not docs:
        raise HTTPException(status_code=404, detail="Nenhum registro encontrado")

    order = {mid: i for i, mid in enumerate(request.movement_ids)}
    docs.sort(key=lambda d: order.get(d['id'], 0))

    company = await get_company_settings()
    pdf_bytes = generate_movement_voucher_pdf(docs, request.via, company)

    filename = f"registro-gate-{docs[0].get('transaction_id')}.pdf" if len(docs) == 1 else f"registros-gate-{len(docs)}-documentos.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@api_router.get("/movements")
async def get_movements(
    operation_type: Optional[str] = None,
    status_filter: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    transaction_id: Optional[str] = None,
    container_number: Optional[str] = None,
    driver_name: Optional[str] = None,
    client_name: Optional[str] = None,
    page: int = 1,
    per_page: int = 0,  # 0 = sem limite (compatível com quem já chamava sem paginar)
    current_user: dict = Depends(get_current_active_user)
):
    # Caso especial: Estoque Atual - containers que entraram mas não saíram
    if operation_type == "ESTOQUE":
        # Buscar todas as movimentações com projeção de campos necessários
        all_movements = await db.movements.find(
            {}, 
            {"_id": 0, "operation_type": 1, "container_number": 1, "status": 1, "created_at": 1,
             "driver_name": 1, "driver_cpf": 1, "truck_plate": 1, "trailer_plate_1": 1,
             "trailer_plate_2": 1, "transport_company": 1, "client_name": 1, "size_type": 1,
             "tare": 1, "shipping_line": 1, "seal": 1, "genset": 1, "booking": 1,
             "service_type": 1, "invoice_number": 1, "service_value": 1, "container_photos": 1,
             "billed": 1, "billed_at": 1, "user_name": 1, "id": 1, "transaction_id": 1}
        ).sort("created_at", -1).to_list(None)
        
        # Agrupar por container_number e verificar o status
        container_status = {}
        for m in all_movements:
            container = m['container_number']
            if container not in container_status:
                container_status[container] = {
                    'last_movement': m,
                    'has_entry': False,
                    'has_exit': False
                }
            
            if m['operation_type'] == 'ENTRADA':
                container_status[container]['has_entry'] = True
            elif m['operation_type'] == 'SAIDA':
                container_status[container]['has_exit'] = True
        
        # Filtrar containers em estoque (entrada sem saída correspondente)
        # Um container está em estoque se a contagem de entradas > contagem de saídas
        container_counts = {}
        for m in sorted(all_movements, key=lambda x: parse_datetime_value(x['created_at'])):
            container = m['container_number']
            if container not in container_counts:
                container_counts[container] = {'entries': 0, 'exits': 0, 'last_entry': None}
            
            if m['operation_type'] == 'ENTRADA':
                container_counts[container]['entries'] += 1
                container_counts[container]['last_entry'] = m
            elif m['operation_type'] == 'SAIDA':
                container_counts[container]['exits'] += 1
        
        # Coletar as últimas entradas dos containers em estoque
        in_stock_movements = []
        for container, counts in container_counts.items():
            if counts['entries'] > counts['exits'] and counts['last_entry']:
                m = counts['last_entry']
                # Aplicar filtro de status se especificado
                if status_filter and m['status'] != status_filter:
                    continue
                in_stock_movements.append(m)
        
        # Ordenar por data decrescente
        in_stock_movements.sort(key=lambda x: parse_datetime_value(x['created_at']), reverse=True)
        
        return [
            ContainerMovementResponse(
                id=m['id'],
                transaction_id=m.get('transaction_id', 0),
                operation_type=m['operation_type'],
                driver_name=m['driver_name'],
                driver_cpf=m['driver_cpf'],
                truck_plate=m['truck_plate'],
                trailer_plate_1=m['trailer_plate_1'],
                trailer_plate_2=m.get('trailer_plate_2'),
                transport_company=m['transport_company'],
                client_name=m.get('client_name'),
                container_number=m['container_number'],
                status=m['status'],
                size_type=m['size_type'],
                tare=m.get('tare'),
                shipping_line=m['shipping_line'],
                seal=m.get('seal'),
                genset=m.get('genset'),
                booking=m.get('booking'),
                service_type=m.get('service_type'),
                invoice_number=m.get('invoice_number'),
                service_value=m.get('service_value'),
                container_photos=m.get('container_photos'),
                container_damages=m.get('container_damages', []),
                billed=m.get('billed', False),
                billed_at=parse_datetime_value(m['billed_at']) if m.get('billed_at') else None,
                created_at=parse_datetime_value(m['created_at']),
                user_name=m['user_name']
            )
            for m in in_stock_movements
        ]
    
    # Fluxo normal para outros tipos de operação
    query = {}
    if operation_type:
        query['operation_type'] = operation_type
    if status_filter:
        query['status'] = status_filter
    if transaction_id:
        try:
            query['transaction_id'] = int(transaction_id)
        except ValueError:
            query['transaction_id'] = -1  # nº inválido -> nenhum resultado, em vez de ignorar o filtro
    if container_number:
        query['container_number'] = {"$regex": re.escape(container_number), "$options": "i"}
    if driver_name:
        query['driver_name'] = {"$regex": re.escape(driver_name), "$options": "i"}
    if client_name:
        query['client_name'] = {"$regex": re.escape(client_name), "$options": "i"}
    # created_at é salvo como string ISO 8601 (sempre UTC) — comparação de string
    # funciona corretamente para ordenar/filtrar por data nesse formato.
    date_range = {}
    if date_from:
        try:
            date_range['$gte'] = datetime.fromisoformat(date_from).replace(tzinfo=timezone.utc).isoformat()
        except ValueError:
            pass
    if date_to:
        try:
            date_range['$lte'] = datetime.fromisoformat(date_to).replace(hour=23, minute=59, second=59, tzinfo=timezone.utc).isoformat()
        except ValueError:
            pass
    if date_range:
        query['created_at'] = date_range

    total = await db.movements.count_documents(query) if per_page else None
    cursor = db.movements.find(query, {"_id": 0}).sort("created_at", -1)
    if per_page:
        cursor = cursor.skip((page - 1) * per_page).limit(per_page)
    movements = await cursor.to_list(None if not per_page else per_page)

    items = [
        ContainerMovementResponse(
            id=m['id'],
            transaction_id=m.get('transaction_id', 0),
            operation_type=m['operation_type'],
            driver_name=m.get('driver_name', ''),
            driver_cpf=m.get('driver_cpf', ''),
            truck_plate=m.get('truck_plate', m.get('vehicle_plate', '')),
            trailer_plate_1=m.get('trailer_plate_1', ''),
            trailer_plate_2=m.get('trailer_plate_2'),
            transport_company=m.get('transport_company', m.get('transport_company_name', '')),
            client_name=m.get('client_name'),
            container_number=m['container_number'],
            status=m['status'],
            size_type=m['size_type'],
            tare=m.get('tare'),
            shipping_line=m['shipping_line'],
            seal=m.get('seal'),
            genset=m.get('genset'),
            booking=m.get('booking'),
            origin_terminal=m.get('origin_terminal'),
            service_type=m.get('service_type'),
            invoice_number=m.get('invoice_number'),
            service_value=m.get('service_value'),
            container_photos=m.get('container_photos'),
            container_damages=m.get('container_damages', []),
            inspection_notes=m.get('inspection_notes'),
            billed=m.get('billed', False),
            billed_at=parse_datetime_value(m['billed_at']) if m.get('billed_at') else None,
            created_at=parse_datetime_value(m['created_at']),
            user_name=m.get('user_name', m.get('created_by_name', ''))
        )
        for m in movements
    ]

    if per_page:
        return {
            "items": items,
            "total": total,
            "page": page,
            "per_page": per_page,
            "pages": (total + per_page - 1) // per_page
        }
    return items

@api_router.get("/movements/unbilled", response_model=List[ContainerMovementResponse])
async def get_unbilled_movements(
    client_name: Optional[str] = None,
    client_cnpj: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_active_user)
):
    """Lista movimentações não faturadas para seleção"""
    query = {"billed": {"$ne": True}}
    
    # Filtrar por cliente (nome ou CNPJ)
    if client_name:
        # Buscar clientes que correspondem ao nome ou CNPJ
        client_name_escaped = re.escape(client_name)
        client_query = await db.clients.find_one({
            "$or": [
                {"name": {"$regex": client_name_escaped, "$options": "i"}},
                {"cnpj": {"$regex": client_name_escaped, "$options": "i"}}
            ]
        }, {"_id": 0})

        if client_query:
            query['client_name'] = client_query['name']
        else:
            # Se não encontrou cliente exato, buscar por nome parcial na movimentação
            query['client_name'] = {"$regex": client_name_escaped, "$options": "i"}
    
    if client_cnpj:
        # Buscar cliente pelo CNPJ
        client_by_cnpj = await db.clients.find_one({"cnpj": client_cnpj}, {"_id": 0})
        if client_by_cnpj:
            query['client_name'] = client_by_cnpj['name']
    
    if search:
        # Buscar por transaction_id, container_number ou ID
        search_escaped = re.escape(search)
        search_conditions = [
            {"container_number": {"$regex": search_escaped, "$options": "i"}},
            {"id": {"$regex": search_escaped, "$options": "i"}},
        ]
        # Tentar buscar por transaction_id se for número
        try:
            transaction_id = int(search)
            search_conditions.append({"transaction_id": transaction_id})
        except ValueError:
            pass
        
        query['$or'] = search_conditions
    
    movements = await db.movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(None)
    
    return [
        ContainerMovementResponse(
            id=m['id'],
            transaction_id=m.get('transaction_id', 0),
            operation_type=m['operation_type'],
            driver_name=m['driver_name'],
            driver_cpf=m['driver_cpf'],
            truck_plate=m['truck_plate'],
            trailer_plate_1=m['trailer_plate_1'],
            trailer_plate_2=m.get('trailer_plate_2'),
            transport_company=m['transport_company'],
            client_name=m.get('client_name'),
            container_number=m['container_number'],
            status=m['status'],
            size_type=m['size_type'],
            tare=m.get('tare'),
            shipping_line=m['shipping_line'],
            seal=m.get('seal'),
            genset=m.get('genset'),
            booking=m.get('booking'),
            origin_terminal=m.get('origin_terminal'),
            service_type=m.get('service_type'),
            invoice_number=m.get('invoice_number'),
            service_value=m.get('service_value'),
            container_photos=m.get('container_photos'),
            container_damages=m.get('container_damages', []),
            inspection_notes=m.get('inspection_notes'),
            billed=m.get('billed', False),
            billed_at=parse_datetime_value(m['billed_at']) if m.get('billed_at') else None,
            created_at=parse_datetime_value(m['created_at']),
            user_name=m['user_name']
        )
        for m in movements
    ]

@api_router.get("/movements/open-entry/{container_number}")
async def get_open_entry_for_container(container_number: str, current_user: dict = Depends(get_current_active_user)):
    """Retorna a ENTRADA mais recente desse container que ainda não teve uma SAÍDA
    correspondente — usado para autopreencher a Saída com os dados da Entrada."""
    entry = await db.movements.find_one(
        {"container_number": container_number.upper(), "operation_type": "ENTRADA"},
        {"_id": 0},
        sort=[("created_at", -1)]
    )
    if not entry:
        return {"entry": None}

    existing_exit = await db.movements.find_one({
        "container_number": entry["container_number"],
        "operation_type": "SAIDA",
        "created_at": {"$gt": entry["created_at"]}
    }, {"_id": 0})

    if existing_exit:
        return {"entry": None}

    return {"entry": entry}

@api_router.get("/movements/{movement_id}", response_model=ContainerMovementResponse)
async def get_movement(movement_id: str, current_user: dict = Depends(get_current_active_user)):
    movement = await db.movements.find_one({"id": movement_id}, {"_id": 0})
    if not movement:
        raise HTTPException(status_code=404, detail="Movimentação não encontrada")
    
    return ContainerMovementResponse(
        id=movement['id'],
        transaction_id=movement.get('transaction_id', 0),
        operation_type=movement['operation_type'],
        driver_name=movement['driver_name'],
        driver_cpf=movement['driver_cpf'],
        truck_plate=movement['truck_plate'],
        trailer_plate_1=movement['trailer_plate_1'],
        trailer_plate_2=movement.get('trailer_plate_2'),
        transport_company=movement['transport_company'],
        client_name=movement.get('client_name'),
        container_number=movement['container_number'],
        status=movement['status'],
        size_type=movement['size_type'],
        tare=movement.get('tare'),
        shipping_line=movement['shipping_line'],
        seal=movement.get('seal'),
        genset=movement.get('genset'),
        booking=movement.get('booking'),
        origin_terminal=movement.get('origin_terminal'),
        service_type=movement.get('service_type'),
        invoice_number=movement.get('invoice_number'),
        service_value=movement.get('service_value'),
        observations=movement.get('observations'),
        container_photos=movement.get('container_photos'),
        container_damages=movement.get('container_damages', []),
        inspection_notes=movement.get('inspection_notes'),
        billed=movement.get('billed', False),
        billed_at=parse_datetime_value(movement['billed_at']) if movement.get('billed_at') else None,
        created_at=parse_datetime_value(movement['created_at']),
        user_name=movement['user_name']
    )

@api_router.put("/movements/{movement_id}", response_model=ContainerMovementResponse)
async def update_movement(movement_id: str, movement_input: ContainerMovementCreate, current_user: dict = Depends(get_current_active_user)):
    existing = await db.movements.find_one({"id": movement_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Movimentação não encontrada")

    if existing.get('billed'):
        requester = await db.users.find_one({"id": current_user['sub']}, {"_id": 0, "role": 1})
        if not requester or requester.get('role') != 'admin':
            raise HTTPException(status_code=403, detail="Movimentação já faturada não pode ser editada. Solicite a um administrador.")

    update_data = movement_input.model_dump()
    update_data['id'] = movement_id
    update_data['transaction_id'] = existing.get('transaction_id', 0)
    update_data['created_at'] = existing['created_at']
    update_data['created_by'] = existing['created_by']
    update_data['user_name'] = existing['user_name']
    update_data['billed'] = existing.get('billed', False)
    update_data['billed_at'] = existing.get('billed_at')
    
    # Arredondar valor monetário para evitar problemas de precisão
    if update_data.get('service_value') is not None:
        update_data['service_value'] = round_money(update_data['service_value'])
    
    await db.movements.replace_one({"id": movement_id}, update_data)
    
    return ContainerMovementResponse(
        id=movement_id,
        transaction_id=update_data['transaction_id'],
        operation_type=update_data['operation_type'],
        driver_name=update_data['driver_name'],
        driver_cpf=update_data['driver_cpf'],
        truck_plate=update_data['truck_plate'],
        trailer_plate_1=update_data['trailer_plate_1'],
        trailer_plate_2=update_data.get('trailer_plate_2'),
        transport_company=update_data['transport_company'],
        client_name=update_data.get('client_name'),
        container_number=update_data['container_number'],
        status=update_data['status'],
        size_type=update_data['size_type'],
        tare=update_data.get('tare'),
        shipping_line=update_data['shipping_line'],
        seal=update_data.get('seal'),
        genset=update_data.get('genset'),
        booking=update_data.get('booking'),
        origin_terminal=update_data.get('origin_terminal'),
        service_type=update_data.get('service_type'),
        invoice_number=update_data.get('invoice_number'),
        service_value=update_data.get('service_value'),
        observations=update_data.get('observations'),
        container_photos=update_data.get('container_photos'),
        container_damages=update_data.get('container_damages', []),
        inspection_notes=update_data.get('inspection_notes'),
        billed=update_data.get('billed', False),
        billed_at=parse_datetime_value(update_data['billed_at']) if update_data.get('billed_at') else None,
        created_at=parse_datetime_value(update_data['created_at']),
        user_name=update_data['user_name']
    )

@api_router.delete("/movements/{movement_id}")
async def delete_movement(movement_id: str, current_user: dict = Depends(get_current_active_user)):
    existing = await db.movements.find_one({"id": movement_id}, {"_id": 0, "billed": 1})
    if not existing:
        raise HTTPException(status_code=404, detail="Movimentação não encontrada")

    if existing.get('billed'):
        requester = await db.users.find_one({"id": current_user['sub']}, {"_id": 0, "role": 1})
        if not requester or requester.get('role') != 'admin':
            raise HTTPException(status_code=403, detail="Movimentação já faturada não pode ser excluída. Solicite a um administrador.")

    result = await db.movements.delete_one({"id": movement_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Movimentação não encontrada")

    # Notificar todos os clientes conectados via WebSocket
    await manager.broadcast({
        "type": "MOVEMENT_DELETED",
        "data": {"id": movement_id}
    })
    
    return {"message": "Movimentação deletada com sucesso"}

@api_router.get("/user/shortcuts")
async def get_user_shortcuts(current_user: dict = Depends(get_current_active_user)):
    user = await db.users.find_one({"email": current_user["email"]}, {"_id": 0, "shortcuts": 1})
    if user and "shortcuts" in user:
        return {"shortcuts": user["shortcuts"]}
    return {"shortcuts": None}


@api_router.put("/user/shortcuts")
async def update_user_shortcuts(data: dict, current_user: dict = Depends(get_current_active_user)):
    shortcuts = data.get("shortcuts", [])
    await db.users.update_one(
        {"email": current_user["email"]},
        {"$set": {"shortcuts": shortcuts}}
    )
    return {"message": "Atalhos atualizados", "shortcuts": shortcuts}


def _created_at_gte(dt: datetime) -> dict:
    """Filtro de data sobre 'created_at' (armazenado como string ISO), convertendo
    para Date dentro do próprio MongoDB - evita comparação de string frágil e evita
    trazer a coleção inteira para o Python só para filtrar por data."""
    return {"$expr": {"$gte": [{"$toDate": "$created_at"}, dt]}}


async def _compute_stock_by_status() -> dict:
    """Para cada container, soma entradas/saídas e pega o status da última entrada
    (histórico completo, feito no MongoDB via agregação em vez de carregar todas as
    movimentações em memória no Python)."""
    pipeline = [
        {"$sort": {"created_at": 1}},
        {"$group": {
            "_id": "$container_number",
            "entries": {"$sum": {"$cond": [{"$eq": ["$operation_type", "ENTRADA"]}, 1, 0]}},
            "exits": {"$sum": {"$cond": [{"$eq": ["$operation_type", "SAIDA"]}, 1, 0]}},
            "last_status": {"$last": "$status"},
        }},
        {"$match": {"$expr": {"$gt": ["$entries", "$exits"]}}},
        {"$group": {"_id": "$last_status", "count": {"$sum": 1}}},
    ]
    result = {r["_id"]: r["count"] async for r in db.movements.aggregate(pipeline)}
    return {"VAZIO": result.get("VAZIO", 0), "CHEIO": result.get("CHEIO", 0)}


async def _compute_daily_chart(today: datetime) -> list[DailyMovementPoint]:
    """Entradas/saídas por dia dos últimos 14 dias - agregação restrita à janela de
    datas (usa o índice de created_at) em vez de escanear a coleção inteira."""
    day0 = today - timedelta(days=13)
    pipeline = [
        {"$match": _created_at_gte(day0)},
        {"$addFields": {"_day": {"$dateToString": {
            "format": "%Y-%m-%d", "date": {"$toDate": "$created_at"}, "timezone": "UTC"
        }}}},
        {"$group": {"_id": {"day": "$_day", "op": "$operation_type"}, "count": {"$sum": 1}}},
    ]
    by_day: dict = {}
    async for r in db.movements.aggregate(pipeline):
        by_day.setdefault(r["_id"]["day"], {})[r["_id"]["op"]] = r["count"]

    daily_chart = []
    for i in range(13, -1, -1):
        day_start = today - timedelta(days=i)
        day_key = day_start.strftime('%Y-%m-%d')
        counts = by_day.get(day_key, {})
        daily_chart.append(DailyMovementPoint(
            date=day_key,
            entries=counts.get('ENTRADA', 0),
            exits=counts.get('SAIDA', 0)
        ))
    return daily_chart


async def _compute_daily_billing_chart(today: datetime) -> list[DailyBillingPoint]:
    """Valor faturado/não faturado por dia dos últimos 14 dias - mesma estratégia de
    agregação restrita à janela de datas usada em _compute_daily_chart."""
    day0 = today - timedelta(days=13)
    pipeline = [
        {"$match": {**_created_at_gte(day0), "service_value": {"$ne": None}}},
        {"$addFields": {"_day": {"$dateToString": {
            "format": "%Y-%m-%d", "date": {"$toDate": "$created_at"}, "timezone": "UTC"
        }}}},
        {"$group": {
            "_id": {"day": "$_day", "billed": {"$eq": ["$billed", True]}},
            "total": {"$sum": "$service_value"}
        }},
    ]
    by_day: dict = {}
    async for r in db.movements.aggregate(pipeline):
        key = 'billed' if r["_id"]["billed"] else 'unbilled'
        by_day.setdefault(r["_id"]["day"], {})[key] = r["total"]

    daily_chart = []
    for i in range(13, -1, -1):
        day_start = today - timedelta(days=i)
        day_key = day_start.strftime('%Y-%m-%d')
        totals = by_day.get(day_key, {})
        daily_chart.append(DailyBillingPoint(
            date=day_key,
            billed=totals.get('billed', 0),
            unbilled=totals.get('unbilled', 0)
        ))
    return daily_chart


async def _compute_driver_ranking(first_day_of_month: datetime) -> list[DriverRankingEntry]:
    """Ranking de motoristas no mês vigente - agregação restrita ao mês (usa o índice
    de created_at) em vez de escanear a coleção inteira."""
    pipeline = [
        {"$match": _created_at_gte(first_day_of_month)},
        {"$group": {
            "_id": {"driver": {"$ifNull": ["$driver_name", "Não informado"]}, "op": "$operation_type"},
            "count": {"$sum": 1}
        }},
    ]
    driver_stats: dict = {}
    async for r in db.movements.aggregate(pipeline):
        name = r["_id"]["driver"]
        data = driver_stats.setdefault(name, {"entries": 0, "exits": 0})
        if r["_id"]["op"] == "ENTRADA":
            data["entries"] = r["count"]
        elif r["_id"]["op"] == "SAIDA":
            data["exits"] = r["count"]

    return sorted(
        [
            DriverRankingEntry(
                driver_name=name,
                entries=data['entries'],
                exits=data['exits'],
                total=data['entries'] + data['exits']
            )
            for name, data in driver_stats.items()
        ],
        key=lambda x: x.total,
        reverse=True
    )[:10]


@api_router.get("/dashboard", response_model=DashboardStats)
async def get_dashboard_stats(current_user: dict = Depends(get_current_active_user)):
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    # Início do mês atual
    first_day_of_month = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # Otimização: cada estatística vira uma query indexada/agregação enxuta no
    # MongoDB (executadas em paralelo) em vez de trazer a coleção inteira de
    # movimentações para o Python a cada carregamento do dashboard.
    (
        total_movements,
        entries_today, exits_today,
        full_containers, empty_containers,
        total_entries, total_exits,
        entries_month, exits_month,
        stock_by_status,
        daily_chart,
        driver_ranking,
        recent,
    ) = await asyncio.gather(
        db.movements.count_documents({}),
        db.movements.count_documents({"operation_type": "ENTRADA", **_created_at_gte(today)}),
        db.movements.count_documents({"operation_type": "SAIDA", **_created_at_gte(today)}),
        db.movements.count_documents({"status": "CHEIO"}),
        db.movements.count_documents({"status": "VAZIO"}),
        db.movements.count_documents({"operation_type": "ENTRADA"}),
        db.movements.count_documents({"operation_type": "SAIDA"}),
        db.movements.count_documents({"operation_type": "ENTRADA", **_created_at_gte(first_day_of_month)}),
        db.movements.count_documents({"operation_type": "SAIDA", **_created_at_gte(first_day_of_month)}),
        _compute_stock_by_status(),
        _compute_daily_chart(today),
        _compute_driver_ranking(first_day_of_month),
        db.movements.find({}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5),
    )

    current_stock = total_entries - total_exits
    stock_empty = stock_by_status["VAZIO"]
    stock_full = stock_by_status["CHEIO"]

    recent_movements = [
        ContainerMovementResponse(
            id=m['id'],
            transaction_id=m.get('transaction_id', 0),
            operation_type=m['operation_type'],
            driver_name=m['driver_name'],
            driver_cpf=m['driver_cpf'],
            truck_plate=m['truck_plate'],
            trailer_plate_1=m['trailer_plate_1'],
            trailer_plate_2=m.get('trailer_plate_2'),
            transport_company=m['transport_company'],
            container_number=m['container_number'],
            status=m['status'],
            size_type=m['size_type'],
            tare=m.get('tare'),
            shipping_line=m['shipping_line'],
            seal=m.get('seal'),
            genset=m.get('genset'),
            booking=m.get('booking'),
            created_at=parse_datetime_value(m['created_at']),
            user_name=m['user_name']
        )
        for m in recent
    ]
    
    return DashboardStats(
        total_movements=total_movements,
        entries_today=entries_today,
        exits_today=exits_today,
        full_containers=full_containers,
        empty_containers=empty_containers,
        total_entries=total_entries,
        total_exits=total_exits,
        current_stock=current_stock,
        stock_empty=stock_empty,
        stock_full=stock_full,
        entries_month=entries_month,
        exits_month=exits_month,
        recent_movements=recent_movements,
        daily_chart=daily_chart,
        driver_ranking=driver_ranking
    )

# ==================== CONTROLE DE ESTOQUE / CONTAINERS NO PÁTIO ====================

@api_router.get("/yard-control")
async def get_yard_control(
    status_filter: Optional[str] = None,
    client_name: Optional[str] = None,
    shipping_line: Optional[str] = None,
    min_days: Optional[int] = None,
    movement_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_active_user)
):
    """Retorna containers em estoque no pátio com contagem de dias"""
    # Otimização: projeta só os campos usados abaixo (id, datas, status etc.) em vez
    # do documento inteiro de cada movimentação (que inclui fotos, observações etc.
    # não usados aqui) - reduz bastante o volume de dados trafegado, sem mudar a
    # lógica de pareamento entrada/saída que continua precisando do histórico inteiro.
    all_movements = await db.movements.find({}, {
        "_id": 0, "id": 1, "transaction_id": 1, "container_number": 1, "status": 1,
        "size_type": 1, "shipping_line": 1, "client_name": 1, "booking": 1,
        "created_at": 1, "tare": 1, "seal": 1, "service_type": 1, "operation_type": 1,
    }).sort("created_at", -1).to_list(None)
    
    # Agrupar por container - rastrear pares de entrada/saída
    container_data = {}
    for m in sorted(all_movements, key=lambda x: parse_datetime_value(x['created_at'])):
        container = m['container_number']
        if container not in container_data:
            container_data[container] = {
                'entries': [],
                'exits': [],
                'movements': []
            }
        
        if m['operation_type'] == 'ENTRADA':
            container_data[container]['entries'].append(m)
            container_data[container]['movements'].append(m)
        elif m['operation_type'] == 'SAIDA':
            container_data[container]['exits'].append(m)
            container_data[container]['movements'].append(m)
    
    # Calcular dias no pátio
    now = datetime.now(timezone.utc)
    yard_containers = []
    
    # Parse date filters
    date_from_dt = None
    date_to_dt = None
    if date_from:
        try:
            date_from_dt = datetime.fromisoformat(date_from).replace(tzinfo=timezone.utc)
        except:
            pass
    if date_to:
        try:
            date_to_dt = datetime.fromisoformat(date_to).replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
        except:
            pass
    
    for container, data in container_data.items():
        num_entries = len(data['entries'])
        num_exits = len(data['exits'])
        
        # Container ainda em estoque (mais entradas que saídas)
        if num_entries > num_exits:
            # Filtro por tipo - se for SAIDA, pular containers em estoque
            if movement_type == 'SAIDA':
                continue
                
            m = data['entries'][-1]  # Última entrada
            entry_date = parse_datetime_value(m['created_at'])
            
            # Aplicar filtros básicos primeiro
            if status_filter and m['status'] != status_filter:
                continue
            # Filtro de cliente: verificar se todas as palavras da busca estão no nome
            if client_name:
                container_client = (m.get('client_name') or '').lower()
                search_words = client_name.lower().split()
                if not all(word in container_client for word in search_words):
                    continue
            if shipping_line and m.get('shipping_line') != shipping_line:
                continue
            
            # Para ESTOQUE ou TODOS com filtro de data:
            # - Mostra containers que estão em estoque E entraram dentro ou antes do período
            if (movement_type == 'ESTOQUE' or not movement_type) and (date_from_dt or date_to_dt):
                # Se container entrou DEPOIS do período final, não mostrar
                if date_to_dt and entry_date > date_to_dt:
                    continue
                
                # Data de referência para início do cálculo: maior entre data de entrada e date_from
                calc_start = entry_date
                if date_from_dt and entry_date < date_from_dt:
                    calc_start = date_from_dt
                
                # Data de referência para fim do cálculo: menor entre agora e date_to
                calc_end = now
                if date_to_dt and now > date_to_dt:
                    calc_end = date_to_dt
                
                # Calcular dias no pátio dentro do período selecionado
                days_in_yard = max(0, (calc_end - calc_start).days)
            elif movement_type == 'ENTRADA' and (date_from_dt or date_to_dt):
                # Para filtro ENTRADA, filtrar pela data de entrada
                if date_from_dt and entry_date < date_from_dt:
                    continue
                if date_to_dt and entry_date > date_to_dt:
                    continue
                days_in_yard = (now - entry_date).days
            else:
                # Cálculo padrão: dias desde a entrada até agora
                days_in_yard = (now - entry_date).days
            
            if min_days is not None and days_in_yard < min_days:
                continue
            
            yard_containers.append({
                'id': m['id'],
                'transaction_id': m.get('transaction_id', 0),
                'container_number': container,
                'status': m['status'],
                'size_type': m['size_type'],
                'shipping_line': m['shipping_line'],
                'client_name': m.get('client_name'),
                'booking': m.get('booking'),
                'entry_date': entry_date.isoformat(),
                'exit_date': None,
                'days_in_yard': days_in_yard,
                'tare': m.get('tare'),
                'seal': m.get('seal'),
                'service_type': m.get('service_type'),
                'operation_type': 'ENTRADA',
                'in_stock': True
            })
        
        # Containers que já saíram - mostrar quando filtro for SAIDA ou TODOS (não ESTOQUE)
        if num_exits > 0 and movement_type != 'ESTOQUE':
            # Para cada par entrada/saída, calcular tempo no pátio
            for i, exit_m in enumerate(data['exits']):
                # Encontrar a entrada correspondente (entrada anterior à saída)
                exit_date = parse_datetime_value(exit_m['created_at'])
                entry_m = None
                entry_date = None
                
                # Buscar a entrada mais recente antes desta saída
                for entry in reversed(data['entries'][:i+1] if i < len(data['entries']) else data['entries']):
                    e_date = parse_datetime_value(entry['created_at'])
                    if e_date < exit_date:
                        entry_m = entry
                        entry_date = e_date
                        break
                
                if not entry_date:
                    # Se não encontrou entrada correspondente, usar a primeira entrada
                    if data['entries']:
                        entry_m = data['entries'][0]
                        entry_date = parse_datetime_value(entry_m['created_at'])
                    else:
                        continue
                
                # Cálculo dos dias no pátio:
                # - Sem filtro de data: tempo total entre entrada e saída
                # - Com filtro de data: limitar ao período selecionado (sobreposição)
                if date_from_dt or date_to_dt:
                    calc_start = entry_date
                    if date_from_dt and entry_date < date_from_dt:
                        calc_start = date_from_dt
                    calc_end = exit_date
                    if date_to_dt and exit_date > date_to_dt:
                        calc_end = date_to_dt
                    days_in_yard = max(0, (calc_end - calc_start).days)
                else:
                    days_in_yard = (exit_date - entry_date).days
                
                # Aplicar filtros na saída
                if status_filter and exit_m['status'] != status_filter:
                    continue
                # Filtro de cliente: verificar se todas as palavras da busca estão no nome
                if client_name:
                    exit_client = (exit_m.get('client_name') or '').lower()
                    search_words = client_name.lower().split()
                    if not all(word in exit_client for word in search_words):
                        continue
                if shipping_line and exit_m.get('shipping_line') != shipping_line:
                    continue
                if min_days is not None and days_in_yard < min_days:
                    continue
                # Filtro por período: incluir a saída se houve sobreposição com o período
                # (entrada antes do date_to E saída depois do date_from)
                if date_from_dt and exit_date < date_from_dt:
                    continue
                if date_to_dt and entry_date > date_to_dt:
                    continue
                
                # Evitar duplicatas - verificar se já foi adicionado
                existing = next((c for c in yard_containers if c['id'] == exit_m['id']), None)
                if existing:
                    continue
                
                yard_containers.append({
                    'id': exit_m['id'],
                    'transaction_id': exit_m.get('transaction_id', 0),
                    'container_number': container,
                    'status': exit_m['status'],
                    'size_type': exit_m['size_type'],
                    'shipping_line': exit_m['shipping_line'],
                    'client_name': exit_m.get('client_name'),
                    'booking': exit_m.get('booking'),
                    'entry_date': entry_date.isoformat() if entry_date else None,
                    'exit_date': exit_date.isoformat(),
                    'days_in_yard': days_in_yard,
                    'tare': exit_m.get('tare'),
                    'seal': exit_m.get('seal'),
                    'service_type': exit_m.get('service_type'),
                    'operation_type': 'SAIDA',
                    'in_stock': False
                })
    
    # Ordenar por dias no pátio (maior primeiro)
    yard_containers.sort(key=lambda x: x['days_in_yard'], reverse=True)

    # Estatísticas (cards do topo, breakdown por cliente/armador) - sempre calculadas
    # só a partir do estoque atual (in_stock), independente do filtro "Tipo" da
    # tabela abaixo. Sem isso, escolher "Todos" ou "Saída" em Tipo inflava "Total
    # no Pátio" somando containers que já saíram junto com os que ainda estão lá.
    stock_only = [c for c in yard_containers if c['in_stock']]
    total_containers = len(stock_only)
    total_empty = sum(1 for c in stock_only if c['status'] == 'VAZIO')
    total_full = sum(1 for c in stock_only if c['status'] == 'CHEIO')
    avg_days = sum(c['days_in_yard'] for c in stock_only) / total_containers if total_containers > 0 else 0
    max_days = max((c['days_in_yard'] for c in stock_only), default=0)

    # Containers com mais de 30 dias (alerta)
    over_30_days = sum(1 for c in stock_only if c['days_in_yard'] > 30)
    over_60_days = sum(1 for c in stock_only if c['days_in_yard'] > 60)
    over_90_days = sum(1 for c in stock_only if c['days_in_yard'] > 90)

    # Estoque por Cliente
    stock_by_client = {}
    for c in stock_only:
        client = c.get('client_name') or 'Sem Cliente'
        if client not in stock_by_client:
            stock_by_client[client] = {'total': 0, 'empty': 0, 'full': 0}
        stock_by_client[client]['total'] += 1
        if c['status'] == 'VAZIO':
            stock_by_client[client]['empty'] += 1
        else:
            stock_by_client[client]['full'] += 1

    by_client = [{'client': k, **v} for k, v in sorted(stock_by_client.items(), key=lambda x: x[1]['total'], reverse=True)]

    # Estoque por Armador
    stock_by_shipping = {}
    for c in stock_only:
        shipping = c.get('shipping_line') or 'Sem Armador'
        if shipping not in stock_by_shipping:
            stock_by_shipping[shipping] = {'total': 0, 'empty': 0, 'full': 0}
        stock_by_shipping[shipping]['total'] += 1
        if c['status'] == 'VAZIO':
            stock_by_shipping[shipping]['empty'] += 1
        else:
            stock_by_shipping[shipping]['full'] += 1

    by_shipping = [{'shipping_line': k, **v} for k, v in sorted(stock_by_shipping.items(), key=lambda x: x[1]['total'], reverse=True)]
    
    return {
        'containers': yard_containers,
        'stats': {
            'total': total_containers,
            'empty': total_empty,
            'full': total_full,
            'avg_days': round(avg_days, 1),
            'max_days': max_days,
            'over_30_days': over_30_days,
            'over_60_days': over_60_days,
            'over_90_days': over_90_days
        },
        'by_client': by_client,
        'by_shipping': by_shipping
    }

@api_router.get("/search")
async def global_search(q: str, current_user: dict = Depends(get_current_active_user)):
    """Busca global usada pela paleta de comandos (Ctrl+K): contêineres, motoristas,
    clientes e transportadoras que batem com o termo digitado."""
    term = q.strip()
    if len(term) < 2:
        return {"results": []}

    regex = {"$regex": re.escape(term), "$options": "i"}
    results = []

    movements = await db.movements.find(
        {"container_number": regex},
        {"_id": 0, "id": 1, "container_number": 1, "operation_type": 1, "status": 1, "driver_name": 1}
    ).sort("created_at", -1).limit(20).to_list(20)
    seen_containers = set()
    for m in movements:
        if m['container_number'] in seen_containers:
            continue
        seen_containers.add(m['container_number'])
        op_label = 'Entrada' if m['operation_type'] == 'ENTRADA' else 'Saída'
        results.append({
            "type": "container",
            "label": m['container_number'],
            "subtitle": f"{op_label} · {m['status']}" + (f" · {m['driver_name']}" if m.get('driver_name') else ""),
            "path": f"/movements/{m['id']}",
        })
        if len(seen_containers) >= 6:
            break

    drivers = await db.drivers.find(
        {"$or": [{"name": regex}, {"cpf": regex}]}, {"_id": 0, "name": 1, "cpf": 1}
    ).limit(6).to_list(6)
    for d in drivers:
        results.append({
            "type": "driver",
            "label": d['name'],
            "subtitle": f"Motorista · CPF {d['cpf']}" if d.get('cpf') else "Motorista",
            "path": f"/drivers?q={url_quote(d['name'])}",
        })

    clients = await db.clients.find(
        {"$or": [{"name": regex}, {"cnpj": regex}]}, {"_id": 0, "name": 1, "cnpj": 1}
    ).limit(6).to_list(6)
    for c in clients:
        results.append({
            "type": "client",
            "label": c['name'],
            "subtitle": f"Cliente · {c['cnpj']}" if c.get('cnpj') else "Cliente",
            "path": f"/clients?q={url_quote(c['name'])}",
        })

    companies = await db.transport_companies.find(
        {"$or": [{"name": regex}, {"cnpj": regex}]}, {"_id": 0, "name": 1, "cnpj": 1}
    ).limit(6).to_list(6)
    for c in companies:
        results.append({
            "type": "company",
            "label": c['name'],
            "subtitle": f"Transportadora · {c['cnpj']}" if c.get('cnpj') else "Transportadora",
            "path": f"/companies?q={url_quote(c['name'])}",
        })

    return {"results": results}

@api_router.get("/alerts/summary")
async def get_alerts_summary(current_user: dict = Depends(get_current_active_user)):
    """Resumo leve de alertas - usado pelo sino do topo e pelo card 'Alertas do
    Sistema' do dashboard (containers parados no pátio há mais de 30/60/90 dias)."""
    yard = await get_yard_control(current_user=current_user)
    return {
        'yard_over_30_days': yard['stats']['over_30_days'],
        'yard_over_60_days': yard['stats']['over_60_days'],
        'yard_over_90_days': yard['stats']['over_90_days'],
    }

from pydantic import BaseModel as PydanticBaseModel

class QuickExitRequest(PydanticBaseModel):
    entry_movement_id: str
    driver_id: Optional[str] = None
    driver_name: Optional[str] = None
    vehicle_plate: Optional[str] = None
    transport_company_id: Optional[str] = None
    transport_company_name: Optional[str] = None
    observations: Optional[str] = None

@api_router.post("/yard-control/quick-exit")
async def register_quick_exit(
    data: QuickExitRequest,
    current_user: dict = Depends(get_current_active_user)
):
    """Registra saída rápida de container do pátio"""
    # Buscar movimentação de entrada
    entry_movement = await db.movements.find_one({"id": data.entry_movement_id}, {"_id": 0})
    if not entry_movement:
        raise HTTPException(status_code=404, detail="Movimentação de entrada não encontrada")
    
    # Verificar se já existe saída
    existing_exit = await db.movements.find_one({
        "container_number": entry_movement["container_number"],
        "operation_type": "SAIDA",
        "created_at": {"$gt": entry_movement["created_at"]}
    }, {"_id": 0})
    
    if existing_exit:
        raise HTTPException(status_code=400, detail="Container já possui saída registrada")
    
    # Gerar número sequencial
    counter = await db.counters.find_one_and_update(
        {"_id": "transaction_id"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    transaction_id = counter["seq"]
    
    # Criar movimentação de saída baseada na entrada
    exit_movement = {
        "id": str(uuid.uuid4()),
        "transaction_id": transaction_id,
        "operation_type": "SAIDA",
        "container_number": entry_movement["container_number"],
        "status": entry_movement["status"],
        "size_type": entry_movement["size_type"],
        "shipping_line": entry_movement["shipping_line"],
        "client_id": entry_movement.get("client_id"),
        "client_name": entry_movement.get("client_name"),
        "booking": entry_movement.get("booking"),
        "bl_number": entry_movement.get("bl_number"),
        "tare": entry_movement.get("tare"),
        "seal": entry_movement.get("seal"),
        "service_type": entry_movement.get("service_type"),
        "driver_id": data.driver_id,
        "driver_name": data.driver_name,
        "vehicle_plate": data.vehicle_plate,
        "transport_company_id": data.transport_company_id,
        "transport_company_name": data.transport_company_name,
        "observations": data.observations or f"Saída rápida - Entrada #{entry_movement.get('transaction_id')}",
        "created_by": current_user["sub"],
        "created_by_name": current_user["name"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": None
    }
    
    await db.movements.insert_one(exit_movement)
    
    # Remover _id se foi adicionado pelo MongoDB
    exit_movement.pop('_id', None)
    
    return {
        "message": "Saída registrada com sucesso",
        "movement": exit_movement,
        "transaction_id": transaction_id
    }

@api_router.get("/yard-control/excel")
async def download_yard_control_excel(
    status_filter: Optional[str] = None,
    client_name: Optional[str] = None,
    shipping_line: Optional[str] = None,
    min_days: Optional[int] = None,
    movement_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_active_user)
):
    """Gera relatório Excel de containers no pátio"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    # Buscar dados com todos os filtros
    yard_data = await get_yard_control(status_filter, client_name, shipping_line, min_days, movement_type, date_from, date_to, current_user)
    containers = yard_data['containers']
    stats = yard_data['stats']
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Controle de Pátio"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )
    warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alert_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    danger_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    entrada_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    saida_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    
    # Título
    ws.merge_cells('A1:J1')
    title_cell = ws.cell(row=1, column=1, value="CONTROLE DE CONTAINERS NO PÁTIO")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Filtros aplicados
    filter_text = "Filtros: "
    if movement_type:
        filter_text += f"Tipo: {movement_type} | "
    if status_filter:
        filter_text += f"Status: {status_filter} | "
    if date_from or date_to:
        filter_text += f"Período: {date_from or 'início'} a {date_to or 'atual'} | "
    ws.merge_cells('A2:J2')
    ws.cell(row=2, column=1, value=filter_text.rstrip(' | ') if filter_text != "Filtros: " else "Filtros: Todos")
    
    # Estatísticas
    ws.cell(row=4, column=1, value=f"Total de Registros: {stats['total']}")
    ws.cell(row=4, column=3, value=f"Vazios: {stats['empty']}")
    ws.cell(row=4, column=5, value=f"Cheios: {stats['full']}")
    ws.cell(row=5, column=1, value=f"Média de Dias: {stats['avg_days']}")
    ws.cell(row=5, column=3, value=f"Máximo de Dias: {stats['max_days']}")
    ws.cell(row=5, column=5, value=f">30 dias: {stats['over_30_days']} | >60 dias: {stats['over_60_days']} | >90 dias: {stats['over_90_days']}")
    
    # Cabeçalhos
    headers = ["Nº Container", "Tipo", "Status", "Tamanho", "Armador", "Cliente", "Data Entrada", "Data Saída", "Dias no Pátio", "Booking"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Congelar
    ws.freeze_panes = 'A8'
    
    # Dados
    for row, c in enumerate(containers, 8):
        ws.cell(row=row, column=1, value=c['container_number']).border = thin_border
        
        tipo_cell = ws.cell(row=row, column=2, value=c.get('operation_type', 'ENTRADA'))
        tipo_cell.border = thin_border
        if c.get('operation_type') == 'ENTRADA' or c.get('in_stock', True):
            tipo_cell.fill = entrada_fill
        else:
            tipo_cell.fill = saida_fill
        
        ws.cell(row=row, column=3, value=c['status']).border = thin_border
        ws.cell(row=row, column=4, value=c['size_type']).border = thin_border
        ws.cell(row=row, column=5, value=c['shipping_line']).border = thin_border
        ws.cell(row=row, column=6, value=c['client_name'] or '-').border = thin_border
        
        # Data Entrada
        if c.get('entry_date'):
            try:
                entry_date = datetime.fromisoformat(c['entry_date'].replace('Z', '+00:00'))
                ws.cell(row=row, column=7, value=entry_date.strftime('%d/%m/%Y')).border = thin_border
            except:
                ws.cell(row=row, column=7, value='-').border = thin_border
        else:
            ws.cell(row=row, column=7, value='-').border = thin_border
        
        # Data Saída
        if c.get('exit_date'):
            try:
                exit_date = datetime.fromisoformat(c['exit_date'].replace('Z', '+00:00'))
                ws.cell(row=row, column=8, value=exit_date.strftime('%d/%m/%Y')).border = thin_border
            except:
                ws.cell(row=row, column=8, value='-').border = thin_border
        else:
            ws.cell(row=row, column=8, value='-').border = thin_border
        
        days_cell = ws.cell(row=row, column=9, value=c['days_in_yard'])
        days_cell.border = thin_border
        days_cell.alignment = Alignment(horizontal='center')
        
        # Colorir baseado nos dias
        if c['days_in_yard'] > 90:
            days_cell.fill = danger_fill
        elif c['days_in_yard'] > 60:
            days_cell.fill = warning_fill
        elif c['days_in_yard'] > 30:
            days_cell.fill = alert_fill
        
        ws.cell(row=row, column=10, value=c['booking'] or '-').border = thin_border
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    
    # Salvar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"controle_patio_{now_brt().strftime('%d-%m-%Y_%H-%M')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/yard-control/pdf")
async def download_yard_control_pdf(
    status_filter: Optional[str] = None,
    client_name: Optional[str] = None,
    shipping_line: Optional[str] = None,
    min_days: Optional[int] = None,
    movement_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_active_user)
):
    """Gera relatório PDF de containers no pátio"""
    yard_data = await get_yard_control(status_filter, client_name, shipping_line, min_days, movement_type, date_from, date_to, current_user)
    company = await get_company_settings()
    pdf_bytes = generate_yard_control_pdf(yard_data['containers'], yard_data['stats'], company)

    filename = f"controle_patio_{now_brt().strftime('%d-%m-%Y_%H-%M')}.pdf"

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

async def _get_current_stock_movements(status_filter: Optional[str] = None, client_name: Optional[str] = None) -> list:
    """Movimentação (última entrada) de cada container que ainda está em estoque
    (mais entradas que saídas) - usado no relatório 'Estoque Atual' (PDF e Excel).

    Otimização: descobrir quais containers estão em estoque exige olhar o histórico
    inteiro (não tem como saber sem contar entradas/saídas de cada um), mas isso é
    feito com uma projeção enxuta (poucos campos) em vez de carregar o documento
    completo de cada movimentação. Só depois o documento completo é buscado, e
    apenas para os containers que realmente estão em estoque hoje - normalmente uma
    fração pequena do histórico total.
    """
    lean_movements = await db.movements.find(
        {}, {"_id": 0, "id": 1, "container_number": 1, "operation_type": 1, "created_at": 1}
    ).sort("created_at", -1).to_list(None)

    container_counts: dict = {}
    for m in sorted(lean_movements, key=lambda x: parse_datetime_value(x['created_at'])):
        container = m['container_number']
        counts = container_counts.setdefault(container, {'entries': 0, 'exits': 0, 'last_entry_id': None})
        if m['operation_type'] == 'ENTRADA':
            counts['entries'] += 1
            counts['last_entry_id'] = m['id']
        elif m['operation_type'] == 'SAIDA':
            counts['exits'] += 1

    in_stock_ids = [
        counts['last_entry_id']
        for counts in container_counts.values()
        if counts['entries'] > counts['exits'] and counts['last_entry_id']
    ]
    if not in_stock_ids:
        return []

    full_docs = {
        doc['id']: doc
        for doc in await db.movements.find({"id": {"$in": in_stock_ids}}, {"_id": 0}).to_list(None)
    }

    movements = []
    for mid in in_stock_ids:
        m = full_docs.get(mid)
        if not m:
            continue
        if status_filter and m['status'] != status_filter:
            continue
        if client_name and m.get('client_name') != client_name:
            continue
        movements.append(m)

    movements.sort(key=lambda x: parse_datetime_value(x['created_at']), reverse=True)
    return movements


@api_router.get("/reports/pdf")
async def download_pdf_report(
    operation_type: Optional[str] = None,
    status_filter: Optional[str] = None,
    client_name: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_active_user)
):
    # Caso especial: Estoque Atual
    if operation_type == "ESTOQUE":
        movements = await _get_current_stock_movements(status_filter, client_name)
    else:
        query = {}
        if operation_type:
            query['operation_type'] = operation_type
        if status_filter:
            query['status'] = status_filter
        if client_name:
            query['client_name'] = client_name
        
        movements = await db.movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(None)
    
    # Filtrar por data se fornecido
    if date_from or date_to:
        filtered_movements = []
        for m in movements:
            try:
                movement_date = parse_datetime_value(m['created_at']).date()
                if date_from and movement_date < datetime.fromisoformat(date_from).date():
                    continue
                if date_to and movement_date > datetime.fromisoformat(date_to).date():
                    continue
                filtered_movements.append(m)
            except:
                filtered_movements.append(m)
        movements = filtered_movements
    
    # Gerar título do relatório baseado no filtro
    report_title = "Relatório de Movimentações"
    if operation_type == "ENTRADA":
        report_title = "Relatório de Entradas"
    elif operation_type == "SAIDA":
        report_title = "Relatório de Saídas"
    elif operation_type == "ESTOQUE":
        report_title = "Relatório de Estoque Atual"
    
    if client_name:
        report_title += f" - Cliente: {client_name}"
    
    company = await get_company_settings()
    pdf_buffer = generate_pdf_report(movements, report_title, company=company)

    return StreamingResponse(
        io.BytesIO(pdf_buffer),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=relatorio_movimentacoes.pdf"}
    )

@api_router.get("/reports/excel")
async def download_excel_report(
    operation_type: Optional[str] = None,
    status_filter: Optional[str] = None,
    client_name: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_active_user)
):
    # Caso especial: Estoque Atual
    if operation_type == "ESTOQUE":
        movements = await _get_current_stock_movements(status_filter, client_name)
    else:
        query = {}
        if operation_type:
            query['operation_type'] = operation_type
        if status_filter:
            query['status'] = status_filter
        if client_name:
            query['client_name'] = client_name

        movements = await db.movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(None)

    # Filtrar por data se fornecido
    if date_from or date_to:
        filtered_movements = []
        for m in movements:
            try:
                movement_date = parse_datetime_value(m['created_at']).date()
                if date_from and movement_date < datetime.fromisoformat(date_from).date():
                    continue
                if date_to and movement_date > datetime.fromisoformat(date_to).date():
                    continue
                filtered_movements.append(m)
            except:
                filtered_movements.append(m)
        movements = filtered_movements
    
    # Gerar título do relatório baseado no filtro
    report_title = "Relatório de Movimentações"
    if operation_type == "ENTRADA":
        report_title = "Relatório de Entradas"
    elif operation_type == "SAIDA":
        report_title = "Relatório de Saídas"
    elif operation_type == "ESTOQUE":
        report_title = "Relatório de Estoque Atual"
    
    if client_name:
        report_title += f" - Cliente: {client_name}"
    
    company = await get_company_settings()
    excel_buffer = generate_excel_report(movements, report_title, company=company)

    return StreamingResponse(
        io.BytesIO(excel_buffer),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=relatorio_movimentacoes.xlsx"}
    )

# ==================== RELATÓRIO DE FATURAMENTO ====================

@api_router.get("/reports/billing/daily-chart", response_model=list[DailyBillingPoint])
async def get_billing_daily_chart(current_user: dict = Depends(get_current_admin_user)):
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    return await _compute_daily_billing_chart(today)


@api_router.get("/reports/billing/pdf")
async def download_billing_pdf_report(
    operation_type: Optional[str] = None,
    status_filter: Optional[str] = None,
    client_name: Optional[str] = None,
    billed_filter: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_admin_user)
):
    query = {}
    if operation_type and operation_type != 'all':
        query['operation_type'] = operation_type
    if status_filter and status_filter != 'all':
        query['status'] = status_filter
    if client_name and client_name != 'all':
        query['client_name'] = client_name
    if billed_filter == 'billed':
        query['billed'] = True
    elif billed_filter == 'unbilled':
        query['billed'] = {"$ne": True}

    movements = await db.movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(None)

    if date_from or date_to:
        filtered = []
        for m in movements:
            try:
                movement_date = parse_datetime_value(m['created_at']).date()
                if date_from and movement_date < datetime.fromisoformat(date_from).date():
                    continue
                if date_to and movement_date > datetime.fromisoformat(date_to).date():
                    continue
                filtered.append(m)
            except:
                filtered.append(m)
        movements = filtered

    report_title = "Relatório de Faturamento"
    if client_name and client_name != 'all':
        report_title += f" - Cliente: {client_name}"

    company = await get_company_settings()
    pdf_buffer = generate_billing_pdf_report(movements, report_title, company=company)

    return StreamingResponse(
        io.BytesIO(pdf_buffer),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=relatorio_faturamento.pdf"}
    )


@api_router.get("/reports/billing/excel")
async def download_billing_excel_report(
    operation_type: Optional[str] = None,
    status_filter: Optional[str] = None,
    client_name: Optional[str] = None,
    billed_filter: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_admin_user)
):
    query = {}
    if operation_type and operation_type != 'all':
        query['operation_type'] = operation_type
    if status_filter and status_filter != 'all':
        query['status'] = status_filter
    if client_name and client_name != 'all':
        query['client_name'] = client_name
    if billed_filter == 'billed':
        query['billed'] = True
    elif billed_filter == 'unbilled':
        query['billed'] = {"$ne": True}

    movements = await db.movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(None)

    if date_from or date_to:
        filtered = []
        for m in movements:
            try:
                movement_date = parse_datetime_value(m['created_at']).date()
                if date_from and movement_date < datetime.fromisoformat(date_from).date():
                    continue
                if date_to and movement_date > datetime.fromisoformat(date_to).date():
                    continue
                filtered.append(m)
            except:
                filtered.append(m)
        movements = filtered

    report_title = "Relatório de Faturamento"
    if client_name and client_name != 'all':
        report_title += f" - Cliente: {client_name}"

    company = await get_company_settings()
    excel_buffer = generate_billing_excel(movements, company=company)

    return StreamingResponse(
        io.BytesIO(excel_buffer),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=relatorio_faturamento.xlsx"}
    )


from pydantic import BaseModel

class BillingRequest(BaseModel):
    movement_ids: List[str]

@api_router.post("/billing/report")
async def generate_billing_report(
    request: BillingRequest,
    current_user: dict = Depends(get_current_admin_user)
):
    from reports import generate_billing_excel
    
    # Buscar movimentações pelos IDs
    movements = await db.movements.find(
        {"id": {"$in": request.movement_ids}},
        {"_id": 0}
    ).to_list(None)
    
    if not movements:
        raise HTTPException(status_code=404, detail="Nenhuma movimentação encontrada")
    
    # Marcar movimentações como faturadas
    billed_at = datetime.now(timezone.utc).isoformat()
    await db.movements.update_many(
        {"id": {"$in": request.movement_ids}},
        {"$set": {"billed": True, "billed_at": billed_at}}
    )
    
    # Ordenar por transaction_id
    movements.sort(key=lambda x: x.get('transaction_id', 0), reverse=True)

    company = await get_company_settings()
    excel_buffer = generate_billing_excel(movements, company=company)

    return StreamingResponse(
        io.BytesIO(excel_buffer),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=faturamento.xlsx"}
    )

