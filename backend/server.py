from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, WebSocket, WebSocketDisconnect, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv

# Carregar variáveis de ambiente ANTES de qualquer outra importação
load_dotenv()

from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from datetime import datetime, timezone, timedelta
from typing import Optional, List
import io
import uuid
import json
import shutil

from models import (
    User, UserCreate, UserLogin, UserResponse, Token,
    Driver, DriverCreate, DriverResponse,
    TransportCompany, TransportCompanyCreate, TransportCompanyResponse,
    ContainerMovement, ContainerMovementCreate, ContainerMovementResponse,
    DashboardStats,
    ShippingLine, ShippingLineCreate, ShippingLineResponse,
    Client, ClientCreate, ClientResponse,
    ServiceType, ServiceTypeCreate, ServiceTypeResponse,
    Invoice, InvoiceCreate, InvoiceUpdate, InvoiceResponse, InvoiceMovementDetail,
    InvoiceHistory, InvoiceHistoryResponse,
    PhotoRegistry, PhotoRegistryCreate, PhotoRegistryUpdate, PhotoRegistryResponse,
    ContainerInspection, ContainerInspectionCreate, ContainerInspectionUpdate, ContainerInspectionResponse,
    FlexTankMovement, FlexTankMovementCreate, FlexTankMovementUpdate, FlexTankMovementResponse, FlexTankStockSummary,
    VehicleRevision, VehicleRevisionCreate, VehicleRevisionResponse,
    IntlInvoice, IntlInvoiceCreate, IntlInvoiceResponse, IntlInvoiceItem
)
from pydantic import BaseModel as PydanticBaseModel
from auth import get_password_hash, verify_password, create_access_token, get_current_user
from reports import generate_pdf_report, generate_excel_report, generate_billing_pdf_report, generate_billing_excel, now_brt, to_brt

# Helper function to parse datetime from MongoDB (handles both string and datetime objects)
def parse_datetime_value(dt_value):
    """Parse datetime from string or return datetime directly - always timezone-aware"""
    if isinstance(dt_value, datetime):
        # Se não tem timezone, assume UTC
        if dt_value.tzinfo is None:
            return dt_value.replace(tzinfo=timezone.utc)
        return dt_value
    if isinstance(dt_value, str):
        parsed = datetime.fromisoformat(dt_value.replace('Z', '+00:00'))
        # Se não tem timezone, assume UTC
        if parsed.tzinfo is None:
            return parsed.replace(tzinfo=timezone.utc)
        return parsed
    return datetime.now(timezone.utc)

# Helper function to round monetary values to 2 decimal places
def round_money(value):
    """Round monetary value to 2 decimal places to avoid floating point precision issues"""
    if value is None:
        return None
    return round(float(value), 2)

ROOT_DIR = Path(__file__).parent
UPLOADS_DIR = ROOT_DIR.parent / 'uploads'
UPLOADS_DIR.mkdir(exist_ok=True)

load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

# Montar diretório de uploads para servir arquivos estáticos via /api/uploads
app.mount("/api/uploads", StaticFiles(directory=str(UPLOADS_DIR)), name="uploads")

# WebSocket connection manager para sincronização em tempo real
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)

    async def broadcast(self, message: dict):
        """Envia mensagem para todos os clientes conectados"""
        disconnected = []
        for connection in self.active_connections:
            try:
                await connection.send_json(message)
            except Exception:
                disconnected.append(connection)
        
        # Remove conexões que falharam
        for conn in disconnected:
            self.disconnect(conn)

manager = ConnectionManager()

# Função para obter próximo transaction_id (sequencial, nunca reutiliza)
async def get_next_transaction_id():
    """Obtém o próximo transaction_id usando um contador atômico"""
    result = await db.counters.find_one_and_update(
        {"_id": "transaction_id"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    return result["seq"]

@api_router.get("/")
async def root():
    return {"message": "ContainerLogix API"}

@api_router.post("/auth/register", response_model=Token)
async def register(user_input: UserCreate):
    existing = await db.users.find_one({"email": user_input.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email já cadastrado")
    
    user = User(
        name=user_input.name,
        email=user_input.email,
        password=get_password_hash(user_input.password),
        role=user_input.role
    )
    
    doc = user.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.users.insert_one(doc)
    
    access_token = create_access_token(data={"sub": user.id, "email": user.email, "name": user.name})
    
    user_response = UserResponse(
        id=user.id,
        name=user.name,
        email=user.email,
        role=user.role,
        created_at=user.created_at
    )
    
    return Token(access_token=access_token, token_type="bearer", user=user_response)

# Password Recovery with Resend
import resend
import asyncio
import secrets
import string

resend.api_key = os.environ.get('RESEND_API_KEY', '')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'onboarding@resend.dev')

def generate_temp_password(length=8):
    """Gera uma senha temporária aleatória"""
    chars = string.ascii_letters + string.digits
    return ''.join(secrets.choice(chars) for _ in range(length))

class ForgotPasswordRequest(PydanticBaseModel):
    email: str

class ChangePasswordRequest(PydanticBaseModel):
    current_password: str
    new_password: str

@api_router.post("/auth/forgot-password")
async def forgot_password(request: ForgotPasswordRequest):
    """Envia senha provisória por email"""
    user_doc = await db.users.find_one({"email": request.email}, {"_id": 0})
    if not user_doc:
        # Não revelamos se o email existe ou não (segurança)
        return {"message": "Se o email estiver cadastrado, você receberá uma senha provisória."}
    
    # Gerar senha provisória
    temp_password = generate_temp_password()
    hashed_password = get_password_hash(temp_password)
    
    # Atualizar no banco com flag must_change_password
    await db.users.update_one(
        {"email": request.email},
        {"$set": {"password": hashed_password, "must_change_password": True}}
    )
    
    # Enviar email com senha provisória
    try:
        html_content = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
            <div style="text-align: center; margin-bottom: 30px;">
                <h1 style="color: #3B9BA8;">J.A Logística</h1>
                <h2 style="color: #333;">ContainerLogix</h2>
            </div>
            
            <div style="background-color: #f5f5f5; padding: 20px; border-radius: 8px; margin-bottom: 20px;">
                <h3 style="color: #333; margin-top: 0;">Recuperação de Senha</h3>
                <p>Olá <strong>{user_doc['name']}</strong>,</p>
                <p>Recebemos uma solicitação de recuperação de senha para sua conta.</p>
                <p>Sua senha provisória é:</p>
                <div style="background-color: #3B9BA8; color: white; padding: 15px; border-radius: 4px; text-align: center; font-size: 24px; font-family: monospace; letter-spacing: 2px;">
                    {temp_password}
                </div>
            </div>
            
            <div style="background-color: #fff3cd; padding: 15px; border-radius: 8px; border-left: 4px solid #ffc107;">
                <p style="margin: 0; color: #856404;">
                    <strong>Importante:</strong> Ao fazer login com esta senha, você será solicitado a criar uma nova senha.
                </p>
            </div>
            
            <p style="color: #666; font-size: 12px; margin-top: 30px; text-align: center;">
                Se você não solicitou esta recuperação, ignore este email.
            </p>
        </div>
        """
        
        params = {
            "from": SENDER_EMAIL,
            "to": [request.email],
            "subject": "ContainerLogix - Recuperação de Senha",
            "html": html_content
        }
        
        # Enviar email de forma assíncrona
        await asyncio.to_thread(resend.Emails.send, params)
        
    except Exception as e:
        logging.error(f"Erro ao enviar email de recuperação: {e}")
        # Mesmo com erro no email, não revelamos detalhes ao usuário
    
    return {"message": "Se o email estiver cadastrado, você receberá uma senha provisória."}

@api_router.post("/auth/change-password")
async def change_password(request: ChangePasswordRequest, current_user: dict = Depends(get_current_user)):
    """Altera a senha do usuário logado"""
    user_doc = await db.users.find_one({"id": current_user['sub']}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    
    # Verificar senha atual
    if not verify_password(request.current_password, user_doc['password']):
        raise HTTPException(status_code=400, detail="Senha atual incorreta")
    
    # Atualizar senha e remover flag must_change_password
    new_hashed_password = get_password_hash(request.new_password)
    await db.users.update_one(
        {"id": current_user['sub']},
        {"$set": {"password": new_hashed_password, "must_change_password": False}}
    )
    
    return {"message": "Senha alterada com sucesso"}

@api_router.post("/auth/login", response_model=Token)
async def login(credentials: UserLogin):
    user_doc = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="Email ou senha incorretos")
    
    if not verify_password(credentials.password, user_doc['password']):
        raise HTTPException(status_code=401, detail="Email ou senha incorretos")
    
    access_token = create_access_token(data={"sub": user_doc['id'], "email": user_doc['email'], "name": user_doc['name']})
    
    user_response = UserResponse(
        id=user_doc['id'],
        name=user_doc['name'],
        email=user_doc['email'],
        role=user_doc['role'],
        must_change_password=user_doc.get('must_change_password', False),
        created_at=datetime.fromisoformat(user_doc['created_at'])
    )
    
    return Token(access_token=access_token, token_type="bearer", user=user_response)

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    user_doc = await db.users.find_one({"id": current_user['sub']}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    
    return UserResponse(
        id=user_doc['id'],
        name=user_doc['name'],
        email=user_doc['email'],
        role=user_doc['role'],
        must_change_password=user_doc.get('must_change_password', False),
        created_at=datetime.fromisoformat(user_doc['created_at'])
    )

@api_router.post("/drivers", response_model=DriverResponse)
async def create_driver(driver_input: DriverCreate, current_user: dict = Depends(get_current_user)):
    driver = Driver(
        name=driver_input.name,
        cpf=driver_input.cpf,
        phone=driver_input.phone,
        created_by=current_user['sub']
    )
    
    doc = driver.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.drivers.insert_one(doc)
    
    return DriverResponse(
        id=driver.id,
        name=driver.name,
        cpf=driver.cpf,
        phone=driver.phone,
        created_at=driver.created_at
    )

@api_router.get("/drivers", response_model=List[DriverResponse])
async def get_drivers(
    page: int = 1,
    per_page: int = 0,  # 0 = sem limite
    current_user: dict = Depends(get_current_user)
):
    # Se per_page = 0, retorna todos os motoristas
    if per_page == 0:
        drivers = await db.drivers.find({}, {"_id": 0}).sort("created_at", -1).to_list(None)
    else:
        skip = (page - 1) * per_page
        drivers = await db.drivers.find({}, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page).to_list(per_page)
    return [
        DriverResponse(
            id=d['id'],
            name=d['name'],
            cpf=d['cpf'],
            phone=d.get('phone'),
            created_at=datetime.fromisoformat(d['created_at'])
        )
        for d in drivers
    ]

@api_router.put("/drivers/{driver_id}", response_model=DriverResponse)
async def update_driver(driver_id: str, driver_input: DriverCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.drivers.find_one({"id": driver_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Motorista não encontrado")
    
    update_data = {
        "id": driver_id,
        "name": driver_input.name,
        "cpf": driver_input.cpf,
        "phone": driver_input.phone,
        "created_at": existing['created_at'],
        "created_by": existing['created_by']
    }
    
    await db.drivers.replace_one({"id": driver_id}, update_data)
    
    return DriverResponse(
        id=driver_id,
        name=update_data['name'],
        cpf=update_data['cpf'],
        phone=update_data['phone'],
        created_at=datetime.fromisoformat(update_data['created_at'])
    )

@api_router.delete("/drivers/{driver_id}")
async def delete_driver(driver_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.drivers.delete_one({"id": driver_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Motorista não encontrado")
    return {"message": "Motorista deletado com sucesso"}

@api_router.post("/transport-companies", response_model=TransportCompanyResponse)
async def create_transport_company(company_input: TransportCompanyCreate, current_user: dict = Depends(get_current_user)):
    company = TransportCompany(
        name=company_input.name,
        cnpj=company_input.cnpj,
        phone=company_input.phone,
        created_by=current_user['sub']
    )
    
    doc = company.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.transport_companies.insert_one(doc)
    
    return TransportCompanyResponse(
        id=company.id,
        name=company.name,
        cnpj=company.cnpj,
        phone=company.phone,
        created_at=company.created_at
    )

@api_router.get("/transport-companies", response_model=List[TransportCompanyResponse])
async def get_transport_companies(
    page: int = 1,
    per_page: int = 0,  # 0 = sem limite
    current_user: dict = Depends(get_current_user)
):
    if per_page == 0:
        companies = await db.transport_companies.find({}, {"_id": 0}).sort("created_at", -1).to_list(None)
    else:
        skip = (page - 1) * per_page
        companies = await db.transport_companies.find({}, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page).to_list(per_page)
    return [
        TransportCompanyResponse(
            id=c['id'],
            name=c['name'],
            cnpj=c.get('cnpj'),
            phone=c.get('phone'),
            created_at=datetime.fromisoformat(c['created_at'])
        )
        for c in companies
    ]

@api_router.post("/shipping-lines", response_model=ShippingLineResponse)
async def create_shipping_line(line_input: ShippingLineCreate, current_user: dict = Depends(get_current_user)):
    line = ShippingLine(
        name=line_input.name,
        code=line_input.code,
        created_by=current_user['sub']
    )
    
    doc = line.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.shipping_lines.insert_one(doc)
    
    return ShippingLineResponse(
        id=line.id,
        name=line.name,
        code=line.code,
        created_at=line.created_at
    )

@api_router.get("/shipping-lines", response_model=List[ShippingLineResponse])
async def get_shipping_lines(
    page: int = 1,
    per_page: int = 0,  # 0 = sem limite
    current_user: dict = Depends(get_current_user)
):
    if per_page == 0:
        lines = await db.shipping_lines.find({}, {"_id": 0}).sort("created_at", -1).to_list(None)
    else:
        skip = (page - 1) * per_page
        lines = await db.shipping_lines.find({}, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page).to_list(per_page)
    return [
        ShippingLineResponse(
            id=line['id'],
            name=line['name'],
            code=line.get('code'),
            created_at=datetime.fromisoformat(line['created_at'])
        )
        for line in lines
    ]

@api_router.put("/shipping-lines/{line_id}", response_model=ShippingLineResponse)
async def update_shipping_line(line_id: str, line_input: ShippingLineCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.shipping_lines.find_one({"id": line_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Armador não encontrado")
    
    update_data = {
        "id": line_id,
        "name": line_input.name,
        "code": line_input.code,
        "created_at": existing['created_at'],
        "created_by": existing['created_by']
    }
    
    await db.shipping_lines.replace_one({"id": line_id}, update_data)
    
    return ShippingLineResponse(
        id=line_id,
        name=update_data['name'],
        code=update_data['code'],
        created_at=datetime.fromisoformat(update_data['created_at'])
    )

@api_router.delete("/shipping-lines/{line_id}")
async def delete_shipping_line(line_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.shipping_lines.delete_one({"id": line_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Armador não encontrado")
    return {"message": "Armador deletado com sucesso"}

@api_router.put("/transport-companies/{company_id}", response_model=TransportCompanyResponse)
async def update_transport_company(company_id: str, company_input: TransportCompanyCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.transport_companies.find_one({"id": company_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Transportadora não encontrada")
    
    update_data = {
        "id": company_id,
        "name": company_input.name,
        "cnpj": company_input.cnpj,
        "phone": company_input.phone,
        "created_at": existing['created_at'],
        "created_by": existing['created_by']
    }
    
    await db.transport_companies.replace_one({"id": company_id}, update_data)
    
    return TransportCompanyResponse(
        id=company_id,
        name=update_data['name'],
        cnpj=update_data['cnpj'],
        phone=update_data['phone'],
        created_at=datetime.fromisoformat(update_data['created_at'])
    )

@api_router.delete("/transport-companies/{company_id}")
async def delete_transport_company(company_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.transport_companies.delete_one({"id": company_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Transportadora não encontrada")
    return {"message": "Transportadora deletada com sucesso"}

# CRUD de Clientes
@api_router.post("/clients", response_model=ClientResponse)
async def create_client(client_input: ClientCreate, current_user: dict = Depends(get_current_user)):
    client = Client(
        name=client_input.name,
        cnpj=client_input.cnpj,
        phone=client_input.phone,
        email=client_input.email,
        address=client_input.address,
        created_by=current_user['sub']
    )
    
    doc = client.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.clients.insert_one(doc)
    
    return ClientResponse(
        id=client.id,
        name=client.name,
        cnpj=client.cnpj,
        phone=client.phone,
        email=client.email,
        address=client.address,
        created_at=client.created_at
    )

@api_router.get("/clients", response_model=List[ClientResponse])
async def get_clients(
    page: int = 1,
    per_page: int = 100,
    current_user: dict = Depends(get_current_user)
):
    skip = (page - 1) * per_page
    clients = await db.clients.find({}, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page).to_list(per_page)
    return [
        ClientResponse(
            id=c['id'],
            name=c['name'],
            cnpj=c.get('cnpj'),
            phone=c.get('phone'),
            email=c.get('email'),
            address=c.get('address'),
            created_at=datetime.fromisoformat(c['created_at'])
        )
        for c in clients
    ]

@api_router.put("/clients/{client_id}", response_model=ClientResponse)
async def update_client(client_id: str, client_input: ClientCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    
    update_data = {
        "id": client_id,
        "name": client_input.name,
        "cnpj": client_input.cnpj,
        "phone": client_input.phone,
        "email": client_input.email,
        "address": client_input.address,
        "created_at": existing['created_at'],
        "created_by": existing['created_by']
    }
    
    await db.clients.replace_one({"id": client_id}, update_data)
    
    return ClientResponse(
        id=client_id,
        name=update_data['name'],
        cnpj=update_data['cnpj'],
        phone=update_data['phone'],
        email=update_data['email'],
        address=update_data['address'],
        created_at=datetime.fromisoformat(update_data['created_at'])
    )

@api_router.delete("/clients/{client_id}")
async def delete_client(client_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.clients.delete_one({"id": client_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    return {"message": "Cliente deletado com sucesso"}

# CRUD de Tipos de Serviço
@api_router.post("/service-types", response_model=ServiceTypeResponse)
async def create_service_type(service_type_input: ServiceTypeCreate, current_user: dict = Depends(get_current_user)):
    service_type = ServiceType(
        name=service_type_input.name,
        description=service_type_input.description,
        created_by=current_user['sub']
    )
    
    doc = service_type.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.service_types.insert_one(doc)
    
    return ServiceTypeResponse(
        id=service_type.id,
        name=service_type.name,
        description=service_type.description,
        created_at=service_type.created_at
    )

@api_router.get("/service-types", response_model=List[ServiceTypeResponse])
async def get_service_types(
    page: int = 1,
    per_page: int = 100,
    current_user: dict = Depends(get_current_user)
):
    skip = (page - 1) * per_page
    service_types = await db.service_types.find({}, {"_id": 0}).sort("name", 1).skip(skip).limit(per_page).to_list(per_page)
    return [
        ServiceTypeResponse(
            id=st['id'],
            name=st['name'],
            description=st.get('description'),
            created_at=datetime.fromisoformat(st['created_at'])
        )
        for st in service_types
    ]

@api_router.put("/service-types/{service_type_id}", response_model=ServiceTypeResponse)
async def update_service_type(service_type_id: str, service_type_input: ServiceTypeCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.service_types.find_one({"id": service_type_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Tipo de Serviço não encontrado")
    
    update_data = {
        "id": service_type_id,
        "name": service_type_input.name,
        "description": service_type_input.description,
        "created_at": existing['created_at'],
        "created_by": existing['created_by']
    }
    
    await db.service_types.replace_one({"id": service_type_id}, update_data)
    
    return ServiceTypeResponse(
        id=service_type_id,
        name=update_data['name'],
        description=update_data['description'],
        created_at=datetime.fromisoformat(update_data['created_at'])
    )

@api_router.delete("/service-types/{service_type_id}")
async def delete_service_type(service_type_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.service_types.delete_one({"id": service_type_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tipo de Serviço não encontrado")
    return {"message": "Tipo de Serviço deletado com sucesso"}

# Upload de fotos de containers
ALLOWED_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.webp'}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

@api_router.post("/upload")
async def upload_file(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Upload de arquivo (imagem) para o servidor"""
    logging.info(f"[Upload] Recebendo arquivo: {file.filename}, content_type: {file.content_type}")
    
    # Verificar extensão
    file_ext = Path(file.filename).suffix.lower() if file.filename else ''
    if file_ext not in ALLOWED_EXTENSIONS:
        logging.error(f"[Upload] Extensão não permitida: {file_ext}")
        raise HTTPException(
            status_code=400, 
            detail=f"Tipo de arquivo não permitido. Use: {', '.join(ALLOWED_EXTENSIONS)}"
        )
    
    # Ler o arquivo
    content = await file.read()
    logging.info(f"[Upload] Arquivo lido: {len(content)} bytes")
    
    # Verificar tamanho
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400, 
            detail=f"Arquivo muito grande. Máximo: {MAX_FILE_SIZE // (1024*1024)}MB"
        )
    
    # Gerar nome único para o arquivo
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = UPLOADS_DIR / unique_filename
    
    # Salvar arquivo
    try:
        with open(file_path, 'wb') as f:
            f.write(content)
        logging.info(f"[Upload] Arquivo salvo: {file_path}")
    except Exception as e:
        logging.error(f"Erro ao salvar arquivo: {e}")
        raise HTTPException(status_code=500, detail="Erro ao salvar arquivo")
    
    # Verificar se o arquivo foi salvo corretamente
    if not file_path.exists():
        logging.error(f"[Upload] Arquivo não encontrado após salvar: {file_path}")
        raise HTTPException(status_code=500, detail="Erro ao verificar arquivo salvo")
    
    url = f"/api/uploads/{unique_filename}"
    logging.info(f"[Upload] Sucesso! URL: {url}")
    
    # Retornar URL do arquivo
    return {
        "filename": unique_filename,
        "url": url,
        "size": len(content),
        "content_type": file.content_type
    }

@api_router.delete("/upload/{filename}")
async def delete_file(filename: str, current_user: dict = Depends(get_current_user)):
    """Deletar arquivo do servidor"""
    file_path = UPLOADS_DIR / filename
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Arquivo não encontrado")
    
    try:
        file_path.unlink()
        return {"message": "Arquivo deletado com sucesso"}
    except Exception as e:
        logging.error(f"Erro ao deletar arquivo: {e}")
        raise HTTPException(status_code=500, detail="Erro ao deletar arquivo")

@api_router.post("/movements", response_model=ContainerMovementResponse)
async def create_movement(movement_input: ContainerMovementCreate, current_user: dict = Depends(get_current_user)):
    # Usar contador atômico para garantir sequência única
    next_transaction_id = await get_next_transaction_id()
    
    movement = ContainerMovement(
        transaction_id=next_transaction_id,
        operation_type=movement_input.operation_type,
        driver_name=movement_input.driver_name,
        driver_cpf=movement_input.driver_cpf,
        truck_plate=movement_input.truck_plate,
        trailer_plate_1=movement_input.trailer_plate_1,
        trailer_plate_2=movement_input.trailer_plate_2,
        transport_company=movement_input.transport_company,
        client_name=movement_input.client_name,
        container_number=movement_input.container_number,
        status=movement_input.status,
        size_type=movement_input.size_type,
        tare=movement_input.tare,
        shipping_line=movement_input.shipping_line,
        seal=movement_input.seal,
        genset=movement_input.genset,
        booking=movement_input.booking,
        service_type=movement_input.service_type,
        invoice_number=movement_input.invoice_number,
        service_value=round_money(movement_input.service_value),
        observations=movement_input.observations,
        container_photos=movement_input.container_photos,
        container_damages=movement_input.container_damages,
        created_by=current_user['sub'],
        user_name=current_user['name']
    )
    
    doc = movement.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    if doc.get('billed_at'):
        doc['billed_at'] = doc['billed_at'].isoformat()
    await db.movements.insert_one(doc)
    
    response = ContainerMovementResponse(
        id=movement.id,
        transaction_id=movement.transaction_id,
        operation_type=movement.operation_type,
        driver_name=movement.driver_name,
        driver_cpf=movement.driver_cpf,
        truck_plate=movement.truck_plate,
        trailer_plate_1=movement.trailer_plate_1,
        trailer_plate_2=movement.trailer_plate_2,
        transport_company=movement.transport_company,
        client_name=movement.client_name,
        container_number=movement.container_number,
        status=movement.status,
        size_type=movement.size_type,
        tare=movement.tare,
        shipping_line=movement.shipping_line,
        seal=movement.seal,
        genset=movement.genset,
        booking=movement.booking,
        service_type=movement.service_type,
        invoice_number=movement.invoice_number,
        service_value=movement.service_value,
        observations=movement.observations,
        container_photos=movement.container_photos,
        container_damages=movement.container_damages,
        billed=movement.billed,
        billed_at=movement.billed_at,
        created_at=movement.created_at,
        user_name=movement.user_name
    )
    
    # Notificar todos os clientes conectados via WebSocket
    await manager.broadcast({
        "type": "MOVEMENT_CREATED",
        "data": {
            "id": response.id,
            "transaction_id": response.transaction_id,
            "operation_type": response.operation_type,
            "driver_name": response.driver_name,
            "driver_cpf": response.driver_cpf,
            "truck_plate": response.truck_plate,
            "trailer_plate_1": response.trailer_plate_1,
            "trailer_plate_2": response.trailer_plate_2,
            "transport_company": response.transport_company,
            "container_number": response.container_number,
            "status": response.status,
            "size_type": response.size_type,
            "tare": response.tare,
            "shipping_line": response.shipping_line,
            "seal": response.seal,
            "genset": response.genset,
            "booking": response.booking,
            "created_at": response.created_at.isoformat(),
            "user_name": response.user_name
        }
    })
    
    return response

@api_router.get("/movements", response_model=List[ContainerMovementResponse])
async def get_movements(
    operation_type: Optional[str] = None,
    status_filter: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    # Caso especial: Estoque Atual - containers que entraram mas não saíram
    if operation_type == "ESTOQUE":
        # Buscar todas as movimentações com projeção de campos necessários
        all_movements = await db.movements.find(
            {}, 
            {"_id": 0, "operation_type": 1, "container_number": 1, "status": 1, "created_at": 1,
             "driver_name": 1, "driver_cpf": 1, "truck_plate": 1, "trailer_plate_1": 1,
             "trailer_plate_2": 1, "transport_company": 1, "client_name": 1, "size_type": 1,
             "tare": 1, "shipping_line": 1, "seal": 1, "genset": 1, "booking": 1,
             "service_type": 1, "invoice_number": 1, "service_value": 1, "container_photos": 1,
             "billed": 1, "billed_at": 1, "user_name": 1, "id": 1, "transaction_id": 1}
        ).sort("created_at", -1).to_list(10000)
        
        # Agrupar por container_number e verificar o status
        container_status = {}
        for m in all_movements:
            container = m['container_number']
            if container not in container_status:
                container_status[container] = {
                    'last_movement': m,
                    'has_entry': False,
                    'has_exit': False
                }
            
            if m['operation_type'] == 'ENTRADA':
                container_status[container]['has_entry'] = True
            elif m['operation_type'] == 'SAIDA':
                container_status[container]['has_exit'] = True
        
        # Filtrar containers em estoque (entrada sem saída correspondente)
        # Um container está em estoque se a contagem de entradas > contagem de saídas
        container_counts = {}
        for m in sorted(all_movements, key=lambda x: parse_datetime_value(x['created_at'])):
            container = m['container_number']
            if container not in container_counts:
                container_counts[container] = {'entries': 0, 'exits': 0, 'last_entry': None}
            
            if m['operation_type'] == 'ENTRADA':
                container_counts[container]['entries'] += 1
                container_counts[container]['last_entry'] = m
            elif m['operation_type'] == 'SAIDA':
                container_counts[container]['exits'] += 1
        
        # Coletar as últimas entradas dos containers em estoque
        in_stock_movements = []
        for container, counts in container_counts.items():
            if counts['entries'] > counts['exits'] and counts['last_entry']:
                m = counts['last_entry']
                # Aplicar filtro de status se especificado
                if status_filter and m['status'] != status_filter:
                    continue
                in_stock_movements.append(m)
        
        # Ordenar por data decrescente
        in_stock_movements.sort(key=lambda x: parse_datetime_value(x['created_at']), reverse=True)
        
        return [
            ContainerMovementResponse(
                id=m['id'],
                transaction_id=m.get('transaction_id', 0),
                operation_type=m['operation_type'],
                driver_name=m['driver_name'],
                driver_cpf=m['driver_cpf'],
                truck_plate=m['truck_plate'],
                trailer_plate_1=m['trailer_plate_1'],
                trailer_plate_2=m.get('trailer_plate_2'),
                transport_company=m['transport_company'],
                client_name=m.get('client_name'),
                container_number=m['container_number'],
                status=m['status'],
                size_type=m['size_type'],
                tare=m.get('tare'),
                shipping_line=m['shipping_line'],
                seal=m.get('seal'),
                genset=m.get('genset'),
                booking=m.get('booking'),
                service_type=m.get('service_type'),
                invoice_number=m.get('invoice_number'),
                service_value=m.get('service_value'),
                container_photos=m.get('container_photos'),
                container_damages=m.get('container_damages', []),
                billed=m.get('billed', False),
                billed_at=parse_datetime_value(m['billed_at']) if m.get('billed_at') else None,
                created_at=parse_datetime_value(m['created_at']),
                user_name=m['user_name']
            )
            for m in in_stock_movements
        ]
    
    # Fluxo normal para outros tipos de operação
    query = {}
    if operation_type:
        query['operation_type'] = operation_type
    if status_filter:
        query['status'] = status_filter
    
    movements = await db.movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(None)
    
    return [
        ContainerMovementResponse(
            id=m['id'],
            transaction_id=m.get('transaction_id', 0),
            operation_type=m['operation_type'],
            driver_name=m.get('driver_name', ''),
            driver_cpf=m.get('driver_cpf', ''),
            truck_plate=m.get('truck_plate', m.get('vehicle_plate', '')),
            trailer_plate_1=m.get('trailer_plate_1', ''),
            trailer_plate_2=m.get('trailer_plate_2'),
            transport_company=m.get('transport_company', m.get('transport_company_name', '')),
            client_name=m.get('client_name'),
            container_number=m['container_number'],
            status=m['status'],
            size_type=m['size_type'],
            tare=m.get('tare'),
            shipping_line=m['shipping_line'],
            seal=m.get('seal'),
            genset=m.get('genset'),
            booking=m.get('booking'),
            service_type=m.get('service_type'),
            invoice_number=m.get('invoice_number'),
            service_value=m.get('service_value'),
            container_photos=m.get('container_photos'),
            container_damages=m.get('container_damages', []),
            billed=m.get('billed', False),
            billed_at=parse_datetime_value(m['billed_at']) if m.get('billed_at') else None,
            created_at=parse_datetime_value(m['created_at']),
            user_name=m.get('user_name', m.get('created_by_name', ''))
        )
        for m in movements
    ]

@api_router.get("/movements/unbilled", response_model=List[ContainerMovementResponse])
async def get_unbilled_movements(
    client_name: Optional[str] = None,
    client_cnpj: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lista movimentações não faturadas para seleção"""
    query = {"billed": {"$ne": True}}
    
    # Filtrar por cliente (nome ou CNPJ)
    if client_name:
        # Buscar clientes que correspondem ao nome ou CNPJ
        client_query = await db.clients.find_one({
            "$or": [
                {"name": {"$regex": client_name, "$options": "i"}},
                {"cnpj": {"$regex": client_name, "$options": "i"}}
            ]
        }, {"_id": 0})
        
        if client_query:
            query['client_name'] = client_query['name']
        else:
            # Se não encontrou cliente exato, buscar por nome parcial na movimentação
            query['client_name'] = {"$regex": client_name, "$options": "i"}
    
    if client_cnpj:
        # Buscar cliente pelo CNPJ
        client_by_cnpj = await db.clients.find_one({"cnpj": client_cnpj}, {"_id": 0})
        if client_by_cnpj:
            query['client_name'] = client_by_cnpj['name']
    
    if search:
        # Buscar por transaction_id, container_number ou ID
        search_conditions = [
            {"container_number": {"$regex": search, "$options": "i"}},
            {"id": {"$regex": search, "$options": "i"}},
        ]
        # Tentar buscar por transaction_id se for número
        try:
            transaction_id = int(search)
            search_conditions.append({"transaction_id": transaction_id})
        except ValueError:
            pass
        
        query['$or'] = search_conditions
    
    movements = await db.movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(None)
    
    return [
        ContainerMovementResponse(
            id=m['id'],
            transaction_id=m.get('transaction_id', 0),
            operation_type=m['operation_type'],
            driver_name=m['driver_name'],
            driver_cpf=m['driver_cpf'],
            truck_plate=m['truck_plate'],
            trailer_plate_1=m['trailer_plate_1'],
            trailer_plate_2=m.get('trailer_plate_2'),
            transport_company=m['transport_company'],
            client_name=m.get('client_name'),
            container_number=m['container_number'],
            status=m['status'],
            size_type=m['size_type'],
            tare=m.get('tare'),
            shipping_line=m['shipping_line'],
            seal=m.get('seal'),
            genset=m.get('genset'),
            booking=m.get('booking'),
            service_type=m.get('service_type'),
            invoice_number=m.get('invoice_number'),
            service_value=m.get('service_value'),
            container_photos=m.get('container_photos'),
            container_damages=m.get('container_damages', []),
            billed=m.get('billed', False),
            billed_at=parse_datetime_value(m['billed_at']) if m.get('billed_at') else None,
            created_at=parse_datetime_value(m['created_at']),
            user_name=m['user_name']
        )
        for m in movements
    ]

@api_router.get("/movements/{movement_id}", response_model=ContainerMovementResponse)
async def get_movement(movement_id: str, current_user: dict = Depends(get_current_user)):
    movement = await db.movements.find_one({"id": movement_id}, {"_id": 0})
    if not movement:
        raise HTTPException(status_code=404, detail="Movimentação não encontrada")
    
    return ContainerMovementResponse(
        id=movement['id'],
        transaction_id=movement.get('transaction_id', 0),
        operation_type=movement['operation_type'],
        driver_name=movement['driver_name'],
        driver_cpf=movement['driver_cpf'],
        truck_plate=movement['truck_plate'],
        trailer_plate_1=movement['trailer_plate_1'],
        trailer_plate_2=movement.get('trailer_plate_2'),
        transport_company=movement['transport_company'],
        client_name=movement.get('client_name'),
        container_number=movement['container_number'],
        status=movement['status'],
        size_type=movement['size_type'],
        tare=movement.get('tare'),
        shipping_line=movement['shipping_line'],
        seal=movement.get('seal'),
        genset=movement.get('genset'),
        booking=movement.get('booking'),
        service_type=movement.get('service_type'),
        invoice_number=movement.get('invoice_number'),
        service_value=movement.get('service_value'),
        observations=movement.get('observations'),
        container_photos=movement.get('container_photos'),
        container_damages=movement.get('container_damages', []),
        billed=movement.get('billed', False),
        billed_at=parse_datetime_value(movement['billed_at']) if movement.get('billed_at') else None,
        created_at=parse_datetime_value(movement['created_at']),
        user_name=movement['user_name']
    )

@api_router.put("/movements/{movement_id}", response_model=ContainerMovementResponse)
async def update_movement(movement_id: str, movement_input: ContainerMovementCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.movements.find_one({"id": movement_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Movimentação não encontrada")
    
    update_data = movement_input.model_dump()
    update_data['id'] = movement_id
    update_data['transaction_id'] = existing.get('transaction_id', 0)
    update_data['created_at'] = existing['created_at']
    update_data['created_by'] = existing['created_by']
    update_data['user_name'] = existing['user_name']
    update_data['billed'] = existing.get('billed', False)
    update_data['billed_at'] = existing.get('billed_at')
    
    # Arredondar valor monetário para evitar problemas de precisão
    if update_data.get('service_value') is not None:
        update_data['service_value'] = round_money(update_data['service_value'])
    
    await db.movements.replace_one({"id": movement_id}, update_data)
    
    return ContainerMovementResponse(
        id=movement_id,
        transaction_id=update_data['transaction_id'],
        operation_type=update_data['operation_type'],
        driver_name=update_data['driver_name'],
        driver_cpf=update_data['driver_cpf'],
        truck_plate=update_data['truck_plate'],
        trailer_plate_1=update_data['trailer_plate_1'],
        trailer_plate_2=update_data.get('trailer_plate_2'),
        transport_company=update_data['transport_company'],
        client_name=update_data.get('client_name'),
        container_number=update_data['container_number'],
        status=update_data['status'],
        size_type=update_data['size_type'],
        tare=update_data.get('tare'),
        shipping_line=update_data['shipping_line'],
        seal=update_data.get('seal'),
        genset=update_data.get('genset'),
        booking=update_data.get('booking'),
        service_type=update_data.get('service_type'),
        invoice_number=update_data.get('invoice_number'),
        service_value=update_data.get('service_value'),
        observations=update_data.get('observations'),
        container_photos=update_data.get('container_photos'),
        container_damages=update_data.get('container_damages', []),
        billed=update_data.get('billed', False),
        billed_at=parse_datetime_value(update_data['billed_at']) if update_data.get('billed_at') else None,
        created_at=parse_datetime_value(update_data['created_at']),
        user_name=update_data['user_name']
    )

@api_router.delete("/movements/{movement_id}")
async def delete_movement(movement_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.movements.delete_one({"id": movement_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Movimentação não encontrada")
    
    # Notificar todos os clientes conectados via WebSocket
    await manager.broadcast({
        "type": "MOVEMENT_DELETED",
        "data": {"id": movement_id}
    })
    
    return {"message": "Movimentação deletada com sucesso"}

@api_router.get("/user/shortcuts")
async def get_user_shortcuts(current_user: dict = Depends(get_current_user)):
    user = await db.users.find_one({"email": current_user["email"]}, {"_id": 0, "shortcuts": 1})
    if user and "shortcuts" in user:
        return {"shortcuts": user["shortcuts"]}
    return {"shortcuts": None}


@api_router.put("/user/shortcuts")
async def update_user_shortcuts(data: dict, current_user: dict = Depends(get_current_user)):
    shortcuts = data.get("shortcuts", [])
    await db.users.update_one(
        {"email": current_user["email"]},
        {"$set": {"shortcuts": shortcuts}}
    )
    return {"message": "Atalhos atualizados", "shortcuts": shortcuts}


@api_router.get("/dashboard", response_model=DashboardStats)
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    # Início do mês atual
    first_day_of_month = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    total_movements = await db.movements.count_documents({})
    
    # Otimização: usar agregação do MongoDB para estatísticas
    # Buscar apenas campos necessários com projeção
    all_movements = await db.movements.find(
        {}, 
        {"_id": 0, "operation_type": 1, "status": 1, "created_at": 1, "container_number": 1}
    ).to_list(10000)
    
    entries_today = sum(1 for m in all_movements if m['operation_type'] == 'ENTRADA' and parse_datetime_value(m['created_at']) >= today)
    exits_today = sum(1 for m in all_movements if m['operation_type'] == 'SAIDA' and parse_datetime_value(m['created_at']) >= today)
    full_containers = sum(1 for m in all_movements if m['status'] == 'CHEIO')
    empty_containers = sum(1 for m in all_movements if m['status'] == 'VAZIO')
    
    total_entries = sum(1 for m in all_movements if m['operation_type'] == 'ENTRADA')
    total_exits = sum(1 for m in all_movements if m['operation_type'] == 'SAIDA')
    current_stock = total_entries - total_exits
    
    # Entradas e saídas do mês vigente
    entries_month = sum(1 for m in all_movements if m['operation_type'] == 'ENTRADA' and parse_datetime_value(m['created_at']) >= first_day_of_month)
    exits_month = sum(1 for m in all_movements if m['operation_type'] == 'SAIDA' and parse_datetime_value(m['created_at']) >= first_day_of_month)
    
    # Calcular estoque atual de vazios e cheios
    # Precisamos agrupar por container e verificar quais estão em estoque
    container_status = {}
    for m in sorted(all_movements, key=lambda x: parse_datetime_value(x['created_at'])):
        container = m['container_number']
        if container not in container_status:
            container_status[container] = {'entries': 0, 'exits': 0, 'last_status': None}
        
        if m['operation_type'] == 'ENTRADA':
            container_status[container]['entries'] += 1
            container_status[container]['last_status'] = m['status']
        elif m['operation_type'] == 'SAIDA':
            container_status[container]['exits'] += 1
    
    # Contar containers em estoque por status
    stock_empty = 0
    stock_full = 0
    for container, data in container_status.items():
        if data['entries'] > data['exits']:  # Container em estoque
            if data['last_status'] == 'VAZIO':
                stock_empty += 1
            elif data['last_status'] == 'CHEIO':
                stock_full += 1
    
    recent = await db.movements.find({}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    recent_movements = [
        ContainerMovementResponse(
            id=m['id'],
            transaction_id=m.get('transaction_id', 0),
            operation_type=m['operation_type'],
            driver_name=m['driver_name'],
            driver_cpf=m['driver_cpf'],
            truck_plate=m['truck_plate'],
            trailer_plate_1=m['trailer_plate_1'],
            trailer_plate_2=m.get('trailer_plate_2'),
            transport_company=m['transport_company'],
            container_number=m['container_number'],
            status=m['status'],
            size_type=m['size_type'],
            tare=m.get('tare'),
            shipping_line=m['shipping_line'],
            seal=m.get('seal'),
            genset=m.get('genset'),
            booking=m.get('booking'),
            created_at=parse_datetime_value(m['created_at']),
            user_name=m['user_name']
        )
        for m in recent
    ]
    
    return DashboardStats(
        total_movements=total_movements,
        entries_today=entries_today,
        exits_today=exits_today,
        full_containers=full_containers,
        empty_containers=empty_containers,
        total_entries=total_entries,
        total_exits=total_exits,
        current_stock=current_stock,
        stock_empty=stock_empty,
        stock_full=stock_full,
        entries_month=entries_month,
        exits_month=exits_month,
        recent_movements=recent_movements
    )

# ==================== CONTROLE DE ESTOQUE / CONTAINERS NO PÁTIO ====================

@api_router.get("/yard-control")
async def get_yard_control(
    status_filter: Optional[str] = None,
    client_name: Optional[str] = None,
    shipping_line: Optional[str] = None,
    min_days: Optional[int] = None,
    movement_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Retorna containers em estoque no pátio com contagem de dias"""
    all_movements = await db.movements.find({}, {"_id": 0}).sort("created_at", -1).to_list(10000)
    
    # Agrupar por container - rastrear pares de entrada/saída
    container_data = {}
    for m in sorted(all_movements, key=lambda x: parse_datetime_value(x['created_at'])):
        container = m['container_number']
        if container not in container_data:
            container_data[container] = {
                'entries': [],
                'exits': [],
                'movements': []
            }
        
        if m['operation_type'] == 'ENTRADA':
            container_data[container]['entries'].append(m)
            container_data[container]['movements'].append(m)
        elif m['operation_type'] == 'SAIDA':
            container_data[container]['exits'].append(m)
            container_data[container]['movements'].append(m)
    
    # Calcular dias no pátio
    now = datetime.now(timezone.utc)
    yard_containers = []
    
    # Parse date filters
    date_from_dt = None
    date_to_dt = None
    if date_from:
        try:
            date_from_dt = datetime.fromisoformat(date_from).replace(tzinfo=timezone.utc)
        except:
            pass
    if date_to:
        try:
            date_to_dt = datetime.fromisoformat(date_to).replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
        except:
            pass
    
    for container, data in container_data.items():
        num_entries = len(data['entries'])
        num_exits = len(data['exits'])
        
        # Container ainda em estoque (mais entradas que saídas)
        if num_entries > num_exits:
            # Filtro por tipo - se for SAIDA, pular containers em estoque
            if movement_type == 'SAIDA':
                continue
                
            m = data['entries'][-1]  # Última entrada
            entry_date = parse_datetime_value(m['created_at'])
            
            # Aplicar filtros básicos primeiro
            if status_filter and m['status'] != status_filter:
                continue
            # Filtro de cliente: verificar se todas as palavras da busca estão no nome
            if client_name:
                container_client = (m.get('client_name') or '').lower()
                search_words = client_name.lower().split()
                if not all(word in container_client for word in search_words):
                    continue
            if shipping_line and m.get('shipping_line') != shipping_line:
                continue
            
            # Para ESTOQUE ou TODOS com filtro de data:
            # - Mostra containers que estão em estoque E entraram dentro ou antes do período
            if (movement_type == 'ESTOQUE' or not movement_type) and (date_from_dt or date_to_dt):
                # Se container entrou DEPOIS do período final, não mostrar
                if date_to_dt and entry_date > date_to_dt:
                    continue
                
                # Data de referência para início do cálculo: maior entre data de entrada e date_from
                calc_start = entry_date
                if date_from_dt and entry_date < date_from_dt:
                    calc_start = date_from_dt
                
                # Data de referência para fim do cálculo: menor entre agora e date_to
                calc_end = now
                if date_to_dt and now > date_to_dt:
                    calc_end = date_to_dt
                
                # Calcular dias no pátio dentro do período selecionado
                days_in_yard = max(0, (calc_end - calc_start).days)
            elif movement_type == 'ENTRADA' and (date_from_dt or date_to_dt):
                # Para filtro ENTRADA, filtrar pela data de entrada
                if date_from_dt and entry_date < date_from_dt:
                    continue
                if date_to_dt and entry_date > date_to_dt:
                    continue
                days_in_yard = (now - entry_date).days
            else:
                # Cálculo padrão: dias desde a entrada até agora
                days_in_yard = (now - entry_date).days
            
            if min_days is not None and days_in_yard < min_days:
                continue
            
            yard_containers.append({
                'id': m['id'],
                'transaction_id': m.get('transaction_id', 0),
                'container_number': container,
                'status': m['status'],
                'size_type': m['size_type'],
                'shipping_line': m['shipping_line'],
                'client_name': m.get('client_name'),
                'booking': m.get('booking'),
                'entry_date': entry_date.isoformat(),
                'exit_date': None,
                'days_in_yard': days_in_yard,
                'tare': m.get('tare'),
                'seal': m.get('seal'),
                'service_type': m.get('service_type'),
                'operation_type': 'ENTRADA',
                'in_stock': True
            })
        
        # Containers que já saíram - mostrar quando filtro for SAIDA ou TODOS (não ESTOQUE)
        if num_exits > 0 and movement_type != 'ESTOQUE':
            # Para cada par entrada/saída, calcular tempo no pátio
            for i, exit_m in enumerate(data['exits']):
                # Encontrar a entrada correspondente (entrada anterior à saída)
                exit_date = parse_datetime_value(exit_m['created_at'])
                entry_m = None
                entry_date = None
                
                # Buscar a entrada mais recente antes desta saída
                for entry in reversed(data['entries'][:i+1] if i < len(data['entries']) else data['entries']):
                    e_date = parse_datetime_value(entry['created_at'])
                    if e_date < exit_date:
                        entry_m = entry
                        entry_date = e_date
                        break
                
                if not entry_date:
                    # Se não encontrou entrada correspondente, usar a primeira entrada
                    if data['entries']:
                        entry_m = data['entries'][0]
                        entry_date = parse_datetime_value(entry_m['created_at'])
                    else:
                        continue
                
                # Cálculo dos dias no pátio:
                # - Sem filtro de data: tempo total entre entrada e saída
                # - Com filtro de data: limitar ao período selecionado (sobreposição)
                if date_from_dt or date_to_dt:
                    calc_start = entry_date
                    if date_from_dt and entry_date < date_from_dt:
                        calc_start = date_from_dt
                    calc_end = exit_date
                    if date_to_dt and exit_date > date_to_dt:
                        calc_end = date_to_dt
                    days_in_yard = max(0, (calc_end - calc_start).days)
                else:
                    days_in_yard = (exit_date - entry_date).days
                
                # Aplicar filtros na saída
                if status_filter and exit_m['status'] != status_filter:
                    continue
                # Filtro de cliente: verificar se todas as palavras da busca estão no nome
                if client_name:
                    exit_client = (exit_m.get('client_name') or '').lower()
                    search_words = client_name.lower().split()
                    if not all(word in exit_client for word in search_words):
                        continue
                if shipping_line and exit_m.get('shipping_line') != shipping_line:
                    continue
                if min_days is not None and days_in_yard < min_days:
                    continue
                # Filtro por período: incluir a saída se houve sobreposição com o período
                # (entrada antes do date_to E saída depois do date_from)
                if date_from_dt and exit_date < date_from_dt:
                    continue
                if date_to_dt and entry_date > date_to_dt:
                    continue
                
                # Evitar duplicatas - verificar se já foi adicionado
                existing = next((c for c in yard_containers if c['id'] == exit_m['id']), None)
                if existing:
                    continue
                
                yard_containers.append({
                    'id': exit_m['id'],
                    'transaction_id': exit_m.get('transaction_id', 0),
                    'container_number': container,
                    'status': exit_m['status'],
                    'size_type': exit_m['size_type'],
                    'shipping_line': exit_m['shipping_line'],
                    'client_name': exit_m.get('client_name'),
                    'booking': exit_m.get('booking'),
                    'entry_date': entry_date.isoformat() if entry_date else None,
                    'exit_date': exit_date.isoformat(),
                    'days_in_yard': days_in_yard,
                    'tare': exit_m.get('tare'),
                    'seal': exit_m.get('seal'),
                    'service_type': exit_m.get('service_type'),
                    'operation_type': 'SAIDA',
                    'in_stock': False
                })
    
    # Ordenar por dias no pátio (maior primeiro)
    yard_containers.sort(key=lambda x: x['days_in_yard'], reverse=True)
    
    # Estatísticas
    total_containers = len(yard_containers)
    total_empty = sum(1 for c in yard_containers if c['status'] == 'VAZIO')
    total_full = sum(1 for c in yard_containers if c['status'] == 'CHEIO')
    avg_days = sum(c['days_in_yard'] for c in yard_containers) / total_containers if total_containers > 0 else 0
    max_days = max((c['days_in_yard'] for c in yard_containers), default=0)
    
    # Containers com mais de 30 dias (alerta)
    over_30_days = sum(1 for c in yard_containers if c['days_in_yard'] > 30)
    over_60_days = sum(1 for c in yard_containers if c['days_in_yard'] > 60)
    over_90_days = sum(1 for c in yard_containers if c['days_in_yard'] > 90)
    
    # Estoque por Cliente
    stock_by_client = {}
    for c in yard_containers:
        client = c.get('client_name') or 'Sem Cliente'
        if client not in stock_by_client:
            stock_by_client[client] = {'total': 0, 'empty': 0, 'full': 0}
        stock_by_client[client]['total'] += 1
        if c['status'] == 'VAZIO':
            stock_by_client[client]['empty'] += 1
        else:
            stock_by_client[client]['full'] += 1
    
    by_client = [{'client': k, **v} for k, v in sorted(stock_by_client.items(), key=lambda x: x[1]['total'], reverse=True)]
    
    # Estoque por Armador
    stock_by_shipping = {}
    for c in yard_containers:
        shipping = c.get('shipping_line') or 'Sem Armador'
        if shipping not in stock_by_shipping:
            stock_by_shipping[shipping] = {'total': 0, 'empty': 0, 'full': 0}
        stock_by_shipping[shipping]['total'] += 1
        if c['status'] == 'VAZIO':
            stock_by_shipping[shipping]['empty'] += 1
        else:
            stock_by_shipping[shipping]['full'] += 1
    
    by_shipping = [{'shipping_line': k, **v} for k, v in sorted(stock_by_shipping.items(), key=lambda x: x[1]['total'], reverse=True)]
    
    return {
        'containers': yard_containers,
        'stats': {
            'total': total_containers,
            'empty': total_empty,
            'full': total_full,
            'avg_days': round(avg_days, 1),
            'max_days': max_days,
            'over_30_days': over_30_days,
            'over_60_days': over_60_days,
            'over_90_days': over_90_days
        },
        'by_client': by_client,
        'by_shipping': by_shipping
    }

from pydantic import BaseModel as PydanticBaseModel

class QuickExitRequest(PydanticBaseModel):
    entry_movement_id: str
    driver_id: Optional[str] = None
    driver_name: Optional[str] = None
    vehicle_plate: Optional[str] = None
    transport_company_id: Optional[str] = None
    transport_company_name: Optional[str] = None
    observations: Optional[str] = None

@api_router.post("/yard-control/quick-exit")
async def register_quick_exit(
    data: QuickExitRequest,
    current_user: dict = Depends(get_current_user)
):
    """Registra saída rápida de container do pátio"""
    # Buscar movimentação de entrada
    entry_movement = await db.movements.find_one({"id": data.entry_movement_id}, {"_id": 0})
    if not entry_movement:
        raise HTTPException(status_code=404, detail="Movimentação de entrada não encontrada")
    
    # Verificar se já existe saída
    existing_exit = await db.movements.find_one({
        "container_number": entry_movement["container_number"],
        "operation_type": "SAIDA",
        "created_at": {"$gt": entry_movement["created_at"]}
    }, {"_id": 0})
    
    if existing_exit:
        raise HTTPException(status_code=400, detail="Container já possui saída registrada")
    
    # Gerar número sequencial
    counter = await db.counters.find_one_and_update(
        {"_id": "transaction_id"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    transaction_id = counter["seq"]
    
    # Criar movimentação de saída baseada na entrada
    exit_movement = {
        "id": str(uuid.uuid4()),
        "transaction_id": transaction_id,
        "operation_type": "SAIDA",
        "container_number": entry_movement["container_number"],
        "status": entry_movement["status"],
        "size_type": entry_movement["size_type"],
        "shipping_line": entry_movement["shipping_line"],
        "client_id": entry_movement.get("client_id"),
        "client_name": entry_movement.get("client_name"),
        "booking": entry_movement.get("booking"),
        "bl_number": entry_movement.get("bl_number"),
        "tare": entry_movement.get("tare"),
        "seal": entry_movement.get("seal"),
        "service_type": entry_movement.get("service_type"),
        "driver_id": data.driver_id,
        "driver_name": data.driver_name,
        "vehicle_plate": data.vehicle_plate,
        "transport_company_id": data.transport_company_id,
        "transport_company_name": data.transport_company_name,
        "observations": data.observations or f"Saída rápida - Entrada #{entry_movement.get('transaction_id')}",
        "created_by": current_user["sub"],
        "created_by_name": current_user["name"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": None
    }
    
    await db.movements.insert_one(exit_movement)
    
    # Remover _id se foi adicionado pelo MongoDB
    exit_movement.pop('_id', None)
    
    return {
        "message": "Saída registrada com sucesso",
        "movement": exit_movement,
        "transaction_id": transaction_id
    }

@api_router.get("/yard-control/excel")
async def download_yard_control_excel(
    status_filter: Optional[str] = None,
    client_name: Optional[str] = None,
    shipping_line: Optional[str] = None,
    min_days: Optional[int] = None,
    movement_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Gera relatório Excel de containers no pátio"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    # Buscar dados com todos os filtros
    yard_data = await get_yard_control(status_filter, client_name, shipping_line, min_days, movement_type, date_from, date_to, current_user)
    containers = yard_data['containers']
    stats = yard_data['stats']
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Controle de Pátio"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )
    warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alert_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    danger_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    entrada_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    saida_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    
    # Título
    ws.merge_cells('A1:J1')
    title_cell = ws.cell(row=1, column=1, value="CONTROLE DE CONTAINERS NO PÁTIO")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Filtros aplicados
    filter_text = "Filtros: "
    if movement_type:
        filter_text += f"Tipo: {movement_type} | "
    if status_filter:
        filter_text += f"Status: {status_filter} | "
    if date_from or date_to:
        filter_text += f"Período: {date_from or 'início'} a {date_to or 'atual'} | "
    ws.merge_cells('A2:J2')
    ws.cell(row=2, column=1, value=filter_text.rstrip(' | ') if filter_text != "Filtros: " else "Filtros: Todos")
    
    # Estatísticas
    ws.cell(row=4, column=1, value=f"Total de Registros: {stats['total']}")
    ws.cell(row=4, column=3, value=f"Vazios: {stats['empty']}")
    ws.cell(row=4, column=5, value=f"Cheios: {stats['full']}")
    ws.cell(row=5, column=1, value=f"Média de Dias: {stats['avg_days']}")
    ws.cell(row=5, column=3, value=f"Máximo de Dias: {stats['max_days']}")
    ws.cell(row=5, column=5, value=f">30 dias: {stats['over_30_days']} | >60 dias: {stats['over_60_days']} | >90 dias: {stats['over_90_days']}")
    
    # Cabeçalhos
    headers = ["Nº Container", "Tipo", "Status", "Tamanho", "Armador", "Cliente", "Data Entrada", "Data Saída", "Dias no Pátio", "Booking"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Congelar
    ws.freeze_panes = 'A8'
    
    # Dados
    for row, c in enumerate(containers, 8):
        ws.cell(row=row, column=1, value=c['container_number']).border = thin_border
        
        tipo_cell = ws.cell(row=row, column=2, value=c.get('operation_type', 'ENTRADA'))
        tipo_cell.border = thin_border
        if c.get('operation_type') == 'ENTRADA' or c.get('in_stock', True):
            tipo_cell.fill = entrada_fill
        else:
            tipo_cell.fill = saida_fill
        
        ws.cell(row=row, column=3, value=c['status']).border = thin_border
        ws.cell(row=row, column=4, value=c['size_type']).border = thin_border
        ws.cell(row=row, column=5, value=c['shipping_line']).border = thin_border
        ws.cell(row=row, column=6, value=c['client_name'] or '-').border = thin_border
        
        # Data Entrada
        if c.get('entry_date'):
            try:
                entry_date = datetime.fromisoformat(c['entry_date'].replace('Z', '+00:00'))
                ws.cell(row=row, column=7, value=entry_date.strftime('%d/%m/%Y')).border = thin_border
            except:
                ws.cell(row=row, column=7, value='-').border = thin_border
        else:
            ws.cell(row=row, column=7, value='-').border = thin_border
        
        # Data Saída
        if c.get('exit_date'):
            try:
                exit_date = datetime.fromisoformat(c['exit_date'].replace('Z', '+00:00'))
                ws.cell(row=row, column=8, value=exit_date.strftime('%d/%m/%Y')).border = thin_border
            except:
                ws.cell(row=row, column=8, value='-').border = thin_border
        else:
            ws.cell(row=row, column=8, value='-').border = thin_border
        
        days_cell = ws.cell(row=row, column=9, value=c['days_in_yard'])
        days_cell.border = thin_border
        days_cell.alignment = Alignment(horizontal='center')
        
        # Colorir baseado nos dias
        if c['days_in_yard'] > 90:
            days_cell.fill = danger_fill
        elif c['days_in_yard'] > 60:
            days_cell.fill = warning_fill
        elif c['days_in_yard'] > 30:
            days_cell.fill = alert_fill
        
        ws.cell(row=row, column=10, value=c['booking'] or '-').border = thin_border
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    
    # Salvar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"controle_patio_{now_brt().strftime('%d-%m-%Y_%H-%M')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/reports/pdf")
async def download_pdf_report(
    operation_type: Optional[str] = None,
    status_filter: Optional[str] = None,
    client_name: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    # Caso especial: Estoque Atual
    if operation_type == "ESTOQUE":
        all_movements = await db.movements.find({}, {"_id": 0}).sort("created_at", -1).to_list(10000)
        
        # Contar entradas e saídas por container
        container_counts = {}
        for m in sorted(all_movements, key=lambda x: parse_datetime_value(x['created_at'])):
            container = m['container_number']
            if container not in container_counts:
                container_counts[container] = {'entries': 0, 'exits': 0, 'last_entry': None}
            
            if m['operation_type'] == 'ENTRADA':
                container_counts[container]['entries'] += 1
                container_counts[container]['last_entry'] = m
            elif m['operation_type'] == 'SAIDA':
                container_counts[container]['exits'] += 1
        
        # Coletar containers em estoque
        movements = []
        for container, counts in container_counts.items():
            if counts['entries'] > counts['exits'] and counts['last_entry']:
                m = counts['last_entry']
                if status_filter and m['status'] != status_filter:
                    continue
                if client_name and m.get('client_name') != client_name:
                    continue
                movements.append(m)
        
        movements.sort(key=lambda x: parse_datetime_value(x['created_at']), reverse=True)
    else:
        query = {}
        if operation_type:
            query['operation_type'] = operation_type
        if status_filter:
            query['status'] = status_filter
        if client_name:
            query['client_name'] = client_name
        
        movements = await db.movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(None)
    
    # Filtrar por data se fornecido
    if date_from or date_to:
        filtered_movements = []
        for m in movements:
            try:
                movement_date = parse_datetime_value(m['created_at']).date()
                if date_from and movement_date < datetime.fromisoformat(date_from).date():
                    continue
                if date_to and movement_date > datetime.fromisoformat(date_to).date():
                    continue
                filtered_movements.append(m)
            except:
                filtered_movements.append(m)
        movements = filtered_movements
    
    # Gerar título do relatório baseado no filtro
    report_title = "Relatório de Movimentações"
    if operation_type == "ENTRADA":
        report_title = "Relatório de Entradas"
    elif operation_type == "SAIDA":
        report_title = "Relatório de Saídas"
    elif operation_type == "ESTOQUE":
        report_title = "Relatório de Estoque Atual"
    
    if client_name:
        report_title += f" - Cliente: {client_name}"
    
    pdf_buffer = generate_pdf_report(movements, report_title)
    
    return StreamingResponse(
        io.BytesIO(pdf_buffer),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=relatorio_movimentacoes.pdf"}
    )

@api_router.get("/reports/excel")
async def download_excel_report(
    operation_type: Optional[str] = None,
    status_filter: Optional[str] = None,
    client_name: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    # Caso especial: Estoque Atual
    if operation_type == "ESTOQUE":
        all_movements = await db.movements.find({}, {"_id": 0}).sort("created_at", -1).to_list(10000)
        
        # Contar entradas e saídas por container
        container_counts = {}
        for m in sorted(all_movements, key=lambda x: parse_datetime_value(x['created_at'])):
            container = m['container_number']
            if container not in container_counts:
                container_counts[container] = {'entries': 0, 'exits': 0, 'last_entry': None}
            
            if m['operation_type'] == 'ENTRADA':
                container_counts[container]['entries'] += 1
                container_counts[container]['last_entry'] = m
            elif m['operation_type'] == 'SAIDA':
                container_counts[container]['exits'] += 1
        
        # Coletar containers em estoque
        movements = []
        for container, counts in container_counts.items():
            if counts['entries'] > counts['exits'] and counts['last_entry']:
                m = counts['last_entry']
                if status_filter and m['status'] != status_filter:
                    continue
                if client_name and m.get('client_name') != client_name:
                    continue
                movements.append(m)
        
        movements.sort(key=lambda x: parse_datetime_value(x['created_at']), reverse=True)
    else:
        query = {}
        if operation_type:
            query['operation_type'] = operation_type
        if status_filter:
            query['status'] = status_filter
        if client_name:
            query['client_name'] = client_name
        
        movements = await db.movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(None)
    
    # Filtrar por data se fornecido
    if date_from or date_to:
        filtered_movements = []
        for m in movements:
            try:
                movement_date = parse_datetime_value(m['created_at']).date()
                if date_from and movement_date < datetime.fromisoformat(date_from).date():
                    continue
                if date_to and movement_date > datetime.fromisoformat(date_to).date():
                    continue
                filtered_movements.append(m)
            except:
                filtered_movements.append(m)
        movements = filtered_movements
    
    # Gerar título do relatório baseado no filtro
    report_title = "Relatório de Movimentações"
    if operation_type == "ENTRADA":
        report_title = "Relatório de Entradas"
    elif operation_type == "SAIDA":
        report_title = "Relatório de Saídas"
    elif operation_type == "ESTOQUE":
        report_title = "Relatório de Estoque Atual"
    
    if client_name:
        report_title += f" - Cliente: {client_name}"
    
    excel_buffer = generate_excel_report(movements, report_title)
    
    return StreamingResponse(
        io.BytesIO(excel_buffer),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=relatorio_movimentacoes.xlsx"}
    )

# ==================== RELATÓRIO DE FATURAMENTO ====================

@api_router.get("/reports/billing/pdf")
async def download_billing_pdf_report(
    operation_type: Optional[str] = None,
    status_filter: Optional[str] = None,
    client_name: Optional[str] = None,
    billed_filter: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if operation_type and operation_type != 'all':
        query['operation_type'] = operation_type
    if status_filter and status_filter != 'all':
        query['status'] = status_filter
    if client_name and client_name != 'all':
        query['client_name'] = client_name
    if billed_filter == 'billed':
        query['billed'] = True
    elif billed_filter == 'unbilled':
        query['billed'] = {"$ne": True}

    movements = await db.movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)

    if date_from or date_to:
        filtered = []
        for m in movements:
            try:
                movement_date = parse_datetime_value(m['created_at']).date()
                if date_from and movement_date < datetime.fromisoformat(date_from).date():
                    continue
                if date_to and movement_date > datetime.fromisoformat(date_to).date():
                    continue
                filtered.append(m)
            except:
                filtered.append(m)
        movements = filtered

    report_title = "Relatório de Faturamento"
    if client_name and client_name != 'all':
        report_title += f" - Cliente: {client_name}"

    pdf_buffer = generate_billing_pdf_report(movements, report_title)

    return StreamingResponse(
        io.BytesIO(pdf_buffer),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=relatorio_faturamento.pdf"}
    )


@api_router.get("/reports/billing/excel")
async def download_billing_excel_report(
    operation_type: Optional[str] = None,
    status_filter: Optional[str] = None,
    client_name: Optional[str] = None,
    billed_filter: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if operation_type and operation_type != 'all':
        query['operation_type'] = operation_type
    if status_filter and status_filter != 'all':
        query['status'] = status_filter
    if client_name and client_name != 'all':
        query['client_name'] = client_name
    if billed_filter == 'billed':
        query['billed'] = True
    elif billed_filter == 'unbilled':
        query['billed'] = {"$ne": True}

    movements = await db.movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)

    if date_from or date_to:
        filtered = []
        for m in movements:
            try:
                movement_date = parse_datetime_value(m['created_at']).date()
                if date_from and movement_date < datetime.fromisoformat(date_from).date():
                    continue
                if date_to and movement_date > datetime.fromisoformat(date_to).date():
                    continue
                filtered.append(m)
            except:
                filtered.append(m)
        movements = filtered

    report_title = "Relatório de Faturamento"
    if client_name and client_name != 'all':
        report_title += f" - Cliente: {client_name}"

    excel_buffer = generate_billing_excel(movements)

    return StreamingResponse(
        io.BytesIO(excel_buffer),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=relatorio_faturamento.xlsx"}
    )


from pydantic import BaseModel

class BillingRequest(BaseModel):
    movement_ids: List[str]

@api_router.post("/billing/report")
async def generate_billing_report(
    request: BillingRequest,
    current_user: dict = Depends(get_current_user)
):
    from reports import generate_billing_excel
    
    # Buscar movimentações pelos IDs
    movements = await db.movements.find(
        {"id": {"$in": request.movement_ids}},
        {"_id": 0}
    ).to_list(None)
    
    if not movements:
        raise HTTPException(status_code=404, detail="Nenhuma movimentação encontrada")
    
    # Marcar movimentações como faturadas
    billed_at = datetime.now(timezone.utc).isoformat()
    await db.movements.update_many(
        {"id": {"$in": request.movement_ids}},
        {"$set": {"billed": True, "billed_at": billed_at}}
    )
    
    # Ordenar por transaction_id
    movements.sort(key=lambda x: x.get('transaction_id', 0), reverse=True)
    
    excel_buffer = generate_billing_excel(movements)
    
    return StreamingResponse(
        io.BytesIO(excel_buffer),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=faturamento.xlsx"}
    )

# ==================== INVOICES (FATURAS) ====================

# Função para obter próximo invoice_number (sequencial)
async def get_next_invoice_number():
    """Obtém o próximo invoice_number usando um contador atômico"""
    result = await db.counters.find_one_and_update(
        {"_id": "invoice_number"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    return result["seq"]

# Função para registrar histórico de alterações
async def log_invoice_history(invoice_id: str, invoice_number: int, action: str, changes: dict, user_id: str, user_name: str):
    """Registra uma alteração no histórico da fatura"""
    history = InvoiceHistory(
        invoice_id=invoice_id,
        invoice_number=invoice_number,
        action=action,
        changes=changes,
        user_id=user_id,
        user_name=user_name
    )
    doc = history.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.invoice_history.insert_one(doc)

@api_router.post("/invoices", response_model=InvoiceResponse)
async def create_invoice(
    invoice_input: InvoiceCreate,
    current_user: dict = Depends(get_current_user)
):
    """Cria uma nova fatura a partir das movimentações selecionadas"""
    
    # Verificar se há movimentações
    if not invoice_input.movement_ids:
        raise HTTPException(status_code=400, detail="Selecione pelo menos uma movimentação")
    
    # Buscar movimentações pelos IDs
    movements = await db.movements.find(
        {"id": {"$in": invoice_input.movement_ids}},
        {"_id": 0}
    ).to_list(None)
    
    if not movements:
        raise HTTPException(status_code=404, detail="Nenhuma movimentação encontrada")
    
    # Verificar se alguma movimentação já foi faturada
    already_billed = [m for m in movements if m.get('billed', False)]
    if already_billed:
        billed_ids = [str(m.get('transaction_id', m['id'])) for m in already_billed]
        raise HTTPException(
            status_code=400, 
            detail=f"As seguintes movimentações já foram faturadas: {', '.join(billed_ids)}"
        )
    
    # Calcular valor total (arredondado para 2 casas decimais)
    total_value = round_money(sum(m.get('service_value', 0) or 0 for m in movements))
    
    # Obter próximo número de fatura
    invoice_number = await get_next_invoice_number()
    
    # Criar a fatura
    invoice = Invoice(
        invoice_number=invoice_number,
        client_name=invoice_input.client_name,
        client_cnpj=invoice_input.client_cnpj,
        movement_ids=invoice_input.movement_ids,
        total_value=total_value,
        notes=invoice_input.notes,
        created_by=current_user['sub'],
        user_name=current_user['name']
    )
    
    doc = invoice.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.invoices.insert_one(doc)
    
    # Marcar movimentações como faturadas
    billed_at = datetime.now(timezone.utc).isoformat()
    await db.movements.update_many(
        {"id": {"$in": invoice_input.movement_ids}},
        {"$set": {"billed": True, "billed_at": billed_at, "invoice_id": invoice.id}}
    )
    
    # Registrar no histórico
    await log_invoice_history(
        invoice_id=invoice.id,
        invoice_number=invoice.invoice_number,
        action="CREATED",
        changes={
            "client_name": invoice.client_name,
            "client_cnpj": invoice.client_cnpj,
            "total_value": total_value,
            "movements_count": len(invoice_input.movement_ids),
            "movement_ids": invoice_input.movement_ids
        },
        user_id=current_user['sub'],
        user_name=current_user['name']
    )
    
    return InvoiceResponse(
        id=invoice.id,
        invoice_number=invoice.invoice_number,
        client_name=invoice.client_name,
        client_cnpj=invoice.client_cnpj,
        movement_ids=invoice.movement_ids,
        total_value=invoice.total_value,
        notes=invoice.notes,
        created_at=invoice.created_at,
        user_name=invoice.user_name
    )

@api_router.get("/invoices", response_model=List[InvoiceResponse])
async def get_invoices(
    page: int = 1,
    per_page: int = 15,
    client_name: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lista todas as faturas com paginação"""
    query = {}
    if client_name:
        query['client_name'] = {"$regex": client_name, "$options": "i"}
    
    skip = (page - 1) * per_page
    
    invoices = await db.invoices.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page).to_list(per_page)
    
    return [
        InvoiceResponse(
            id=inv['id'],
            invoice_number=inv['invoice_number'],
            client_name=inv['client_name'],
            client_cnpj=inv.get('client_cnpj'),
            movement_ids=inv['movement_ids'],
            total_value=inv['total_value'],
            notes=inv.get('notes'),
            created_at=datetime.fromisoformat(inv['created_at']),
            user_name=inv['user_name']
        )
        for inv in invoices
    ]

@api_router.get("/invoices/count")
async def get_invoices_count(
    client_name: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Retorna a contagem total de faturas"""
    query = {}
    if client_name:
        query['client_name'] = {"$regex": client_name, "$options": "i"}
    
    count = await db.invoices.count_documents(query)
    return {"count": count}

@api_router.get("/invoices/{invoice_id}", response_model=InvoiceResponse)
async def get_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
    """Busca uma fatura específica por ID"""
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Fatura não encontrada")
    
    return InvoiceResponse(
        id=invoice['id'],
        invoice_number=invoice['invoice_number'],
        client_name=invoice['client_name'],
        client_cnpj=invoice.get('client_cnpj'),
        movement_ids=invoice['movement_ids'],
        total_value=invoice['total_value'],
        notes=invoice.get('notes'),
        created_at=datetime.fromisoformat(invoice['created_at']),
        user_name=invoice['user_name']
    )

@api_router.put("/invoices/{invoice_id}", response_model=InvoiceResponse)
async def update_invoice(invoice_id: str, invoice_update: InvoiceUpdate, current_user: dict = Depends(get_current_user)):
    """Atualiza uma fatura existente (dados e movimentações)"""
    # Buscar fatura existente
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Fatura não encontrada")
    
    # Preparar dados para atualização e rastrear mudanças
    update_data = {}
    changes = {}
    
    # Atualizar campos básicos se fornecidos
    if invoice_update.client_name is not None and invoice_update.client_name != invoice.get('client_name'):
        changes['client_name'] = {'from': invoice.get('client_name'), 'to': invoice_update.client_name}
        update_data['client_name'] = invoice_update.client_name
    if invoice_update.client_cnpj is not None and invoice_update.client_cnpj != invoice.get('client_cnpj'):
        changes['client_cnpj'] = {'from': invoice.get('client_cnpj'), 'to': invoice_update.client_cnpj}
        update_data['client_cnpj'] = invoice_update.client_cnpj
    if invoice_update.notes is not None and invoice_update.notes != invoice.get('notes'):
        changes['notes'] = {'from': invoice.get('notes'), 'to': invoice_update.notes}
        update_data['notes'] = invoice_update.notes
    
    # Processar movimentações
    current_movement_ids = set(invoice['movement_ids'])
    
    # Remover movimentações
    if invoice_update.movement_ids_to_remove:
        for mov_id in invoice_update.movement_ids_to_remove:
            current_movement_ids.discard(mov_id)
        # Desmarcar movimentações como faturadas
        await db.movements.update_many(
            {"id": {"$in": invoice_update.movement_ids_to_remove}},
            {"$set": {"billed": False, "billed_at": None, "invoice_id": None}}
        )
        changes['movements_removed'] = invoice_update.movement_ids_to_remove
    
    # Adicionar movimentações
    if invoice_update.movement_ids_to_add:
        # Verificar se as movimentações existem e não estão faturadas
        movements_to_add = await db.movements.find(
            {"id": {"$in": invoice_update.movement_ids_to_add}},
            {"_id": 0, "id": 1, "billed": 1, "service_value": 1}
        ).to_list(None)
        
        for mov in movements_to_add:
            if mov.get('billed', False):
                raise HTTPException(
                    status_code=400,
                    detail=f"Movimentação {mov['id']} já está faturada em outra fatura"
                )
            current_movement_ids.add(mov['id'])
        
        # Marcar novas movimentações como faturadas
        billed_at = datetime.now(timezone.utc).isoformat()
        await db.movements.update_many(
            {"id": {"$in": invoice_update.movement_ids_to_add}},
            {"$set": {"billed": True, "billed_at": billed_at, "invoice_id": invoice_id}}
        )
        changes['movements_added'] = invoice_update.movement_ids_to_add
    
    # Verificar se resta pelo menos uma movimentação
    if len(current_movement_ids) == 0:
        raise HTTPException(status_code=400, detail="A fatura deve ter pelo menos uma movimentação")
    
    # Atualizar lista de movimentações
    update_data['movement_ids'] = list(current_movement_ids)
    
    # Recalcular valor total (arredondado para 2 casas decimais)
    movements = await db.movements.find(
        {"id": {"$in": list(current_movement_ids)}},
        {"_id": 0, "service_value": 1}
    ).to_list(None)
    new_total = round_money(sum(m.get('service_value', 0) or 0 for m in movements))
    
    if new_total != invoice.get('total_value'):
        changes['total_value'] = {'from': invoice.get('total_value'), 'to': new_total}
    update_data['total_value'] = new_total
    
    # Aplicar atualização
    await db.invoices.update_one({"id": invoice_id}, {"$set": update_data})
    
    # Registrar no histórico se houve mudanças
    if changes:
        await log_invoice_history(
            invoice_id=invoice_id,
            invoice_number=invoice['invoice_number'],
            action="UPDATED",
            changes=changes,
            user_id=current_user['sub'],
            user_name=current_user['name']
        )
    
    # Buscar fatura atualizada
    updated_invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    
    return InvoiceResponse(
        id=updated_invoice['id'],
        invoice_number=updated_invoice['invoice_number'],
        client_name=updated_invoice['client_name'],
        client_cnpj=updated_invoice.get('client_cnpj'),
        movement_ids=updated_invoice['movement_ids'],
        total_value=updated_invoice['total_value'],
        notes=updated_invoice.get('notes'),
        created_at=datetime.fromisoformat(updated_invoice['created_at']),
        user_name=updated_invoice['user_name']
    )

@api_router.get("/invoices/{invoice_id}/movements", response_model=List[InvoiceMovementDetail])
async def get_invoice_movements(invoice_id: str, current_user: dict = Depends(get_current_user)):
    """Retorna detalhes das movimentações de uma fatura"""
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Fatura não encontrada")
    
    movements = await db.movements.find(
        {"id": {"$in": invoice['movement_ids']}},
        {"_id": 0}
    ).to_list(None)
    
    return [
        InvoiceMovementDetail(
            id=m['id'],
            transaction_id=m.get('transaction_id', 0),
            container_number=m['container_number'],
            operation_type=m['operation_type'],
            service_type=m.get('service_type'),
            service_value=m.get('service_value'),
            created_at=parse_datetime_value(m['created_at'])
        )
        for m in movements
    ]

@api_router.get("/invoices/{invoice_id}/pdf")
async def download_invoice_pdf(invoice_id: str, current_user: dict = Depends(get_current_user)):
    """Gera PDF da fatura para download"""
    from reports import generate_invoice_pdf
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Fatura não encontrada")
    
    movements = await db.movements.find(
        {"id": {"$in": invoice['movement_ids']}},
        {"_id": 0}
    ).to_list(None)
    
    pdf_buffer = generate_invoice_pdf(invoice, movements)
    
    filename = f"fatura_{invoice['invoice_number']}.pdf"
    
    return StreamingResponse(
        io.BytesIO(pdf_buffer),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/invoices/{invoice_id}/excel")
async def download_invoice_excel(invoice_id: str, current_user: dict = Depends(get_current_user)):
    """Gera Excel da fatura para download"""
    from reports import generate_invoice_excel
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Fatura não encontrada")
    
    movements = await db.movements.find(
        {"id": {"$in": invoice['movement_ids']}},
        {"_id": 0}
    ).to_list(None)
    
    excel_buffer = generate_invoice_excel(invoice, movements)
    
    filename = f"fatura_{invoice['invoice_number']}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(excel_buffer),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.delete("/invoices/{invoice_id}")
async def delete_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
    """Exclui uma fatura e desmarca as movimentações como faturadas"""
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Fatura não encontrada")
    
    # Registrar no histórico antes de deletar
    await log_invoice_history(
        invoice_id=invoice_id,
        invoice_number=invoice['invoice_number'],
        action="DELETED",
        changes={
            "client_name": invoice.get('client_name'),
            "total_value": invoice.get('total_value'),
            "movements_count": len(invoice.get('movement_ids', []))
        },
        user_id=current_user['sub'],
        user_name=current_user['name']
    )
    
    # Desmarcar movimentações como faturadas
    await db.movements.update_many(
        {"id": {"$in": invoice['movement_ids']}},
        {"$set": {"billed": False, "billed_at": None}, "$unset": {"invoice_id": ""}}
    )
    
    # Excluir a fatura
    await db.invoices.delete_one({"id": invoice_id})
    
    return {"message": "Fatura excluída com sucesso"}

@api_router.get("/invoices/{invoice_id}/history", response_model=List[InvoiceHistoryResponse])
async def get_invoice_history(invoice_id: str, current_user: dict = Depends(get_current_user)):
    """Retorna o histórico de alterações de uma fatura"""
    history = await db.invoice_history.find(
        {"invoice_id": invoice_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return [
        InvoiceHistoryResponse(
            id=h['id'],
            invoice_id=h['invoice_id'],
            invoice_number=h['invoice_number'],
            action=h['action'],
            changes=h['changes'],
            user_name=h['user_name'],
            created_at=datetime.fromisoformat(h['created_at'])
        )
        for h in history
    ]

# ===== REGISTRO FOTOGRÁFICO =====

async def get_next_registry_number():
    """Obtém o próximo registry_number usando um contador atômico"""
    result = await db.counters.find_one_and_update(
        {"_id": "registry_number"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    return result["seq"]

@api_router.get("/photo-registries")
async def list_photo_registries(
    page: int = 1,
    page_size: int = 20,
    search: Optional[str] = None,
    client_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lista registros fotográficos com paginação"""
    query = {}
    
    if search:
        query["$or"] = [
            {"container_number": {"$regex": search, "$options": "i"}},
            {"booking": {"$regex": search, "$options": "i"}},
            {"client_name": {"$regex": search, "$options": "i"}},
        ]
    
    if client_id:
        query["client_id"] = client_id
    
    total = await db.photo_registries.count_documents(query)
    skip = (page - 1) * page_size
    
    registries = await db.photo_registries.find(
        query, {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    
    return {
        "items": registries,
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": (total + page_size - 1) // page_size
    }

@api_router.get("/photo-registries/{registry_id}", response_model=PhotoRegistryResponse)
async def get_photo_registry(registry_id: str, current_user: dict = Depends(get_current_user)):
    """Obtém um registro fotográfico específico"""
    registry = await db.photo_registries.find_one({"id": registry_id}, {"_id": 0})
    if not registry:
        raise HTTPException(status_code=404, detail="Registro fotográfico não encontrado")
    return PhotoRegistryResponse(**registry)

@api_router.post("/photo-registries", response_model=PhotoRegistryResponse)
async def create_photo_registry(
    data: PhotoRegistryCreate,
    current_user: dict = Depends(get_current_user)
):
    """Cria um novo registro fotográfico"""
    registry_number = await get_next_registry_number()
    
    # Buscar nome do cliente se informado
    client_name = None
    if data.client_id:
        client = await db.clients.find_one({"id": data.client_id}, {"_id": 0, "name": 1})
        if client:
            client_name = client.get("name")
    
    # Buscar nome do armador se informado
    shipping_line_name = None
    if data.shipping_line_id:
        shipping_line = await db.shipping_lines.find_one({"id": data.shipping_line_id}, {"_id": 0, "name": 1})
        if shipping_line:
            shipping_line_name = shipping_line.get("name")
    
    registry = PhotoRegistry(
        registry_number=registry_number,
        container_number=data.container_number,
        container_seal=data.container_seal,
        collection_terminal=data.collection_terminal,
        booking=data.booking,
        client_id=data.client_id,
        client_name=client_name,
        shipping_line_id=data.shipping_line_id,
        shipping_line_name=shipping_line_name,
        created_by=current_user["sub"],
        created_by_name=current_user["name"]
    )
    
    registry_dict = registry.model_dump()
    registry_dict["created_at"] = registry_dict["created_at"].isoformat()
    
    await db.photo_registries.insert_one(registry_dict)
    
    return PhotoRegistryResponse(**registry_dict)

@api_router.put("/photo-registries/{registry_id}", response_model=PhotoRegistryResponse)
async def update_photo_registry(
    registry_id: str,
    data: PhotoRegistryUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Atualiza um registro fotográfico"""
    registry = await db.photo_registries.find_one({"id": registry_id}, {"_id": 0})
    if not registry:
        raise HTTPException(status_code=404, detail="Registro fotográfico não encontrado")
    
    update_data = {}
    
    if data.container_number is not None:
        update_data["container_number"] = data.container_number
    
    if data.container_seal is not None:
        update_data["container_seal"] = data.container_seal
    
    if data.collection_terminal is not None:
        update_data["collection_terminal"] = data.collection_terminal
    
    if data.booking is not None:
        update_data["booking"] = data.booking
    
    if data.client_id is not None:
        update_data["client_id"] = data.client_id
        if data.client_id:
            client = await db.clients.find_one({"id": data.client_id}, {"_id": 0, "name": 1})
            update_data["client_name"] = client.get("name") if client else None
        else:
            update_data["client_name"] = None
    
    if data.shipping_line_id is not None:
        update_data["shipping_line_id"] = data.shipping_line_id
        if data.shipping_line_id:
            shipping_line = await db.shipping_lines.find_one({"id": data.shipping_line_id}, {"_id": 0, "name": 1})
            update_data["shipping_line_name"] = shipping_line.get("name") if shipping_line else None
        else:
            update_data["shipping_line_name"] = None
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.photo_registries.update_one({"id": registry_id}, {"$set": update_data})
    
    updated = await db.photo_registries.find_one({"id": registry_id}, {"_id": 0})
    return PhotoRegistryResponse(**updated)

@api_router.post("/photo-registries/{registry_id}/upload-photo")
async def upload_photo_registry_photo(
    registry_id: str,
    position: str,  # front, back, left, right
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload de foto para um registro fotográfico"""
    if position not in ["front", "back", "left", "right"]:
        raise HTTPException(status_code=400, detail="Posição inválida. Use: front, back, left, right")
    
    registry = await db.photo_registries.find_one({"id": registry_id}, {"_id": 0})
    if not registry:
        raise HTTPException(status_code=404, detail="Registro fotográfico não encontrado")
    
    # Criar diretório para fotos de registro
    photo_dir = UPLOADS_DIR / "photo_registries" / registry_id
    photo_dir.mkdir(parents=True, exist_ok=True)
    
    # Gerar nome do arquivo
    file_ext = Path(file.filename).suffix.lower() if file.filename else ".jpg"
    if file_ext not in [".jpg", ".jpeg", ".png", ".webp"]:
        file_ext = ".jpg"
    
    filename = f"{position}{file_ext}"
    file_path = photo_dir / filename
    
    # Salvar arquivo
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # URL relativa do arquivo
    photo_url = f"/api/uploads/photo_registries/{registry_id}/{filename}"
    
    # Atualizar registro no banco
    field_name = f"photo_{position}"
    await db.photo_registries.update_one(
        {"id": registry_id},
        {"$set": {field_name: photo_url, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"url": photo_url, "position": position}

@api_router.delete("/photo-registries/{registry_id}/photo/{position}")
async def delete_photo_registry_photo(
    registry_id: str,
    position: str,
    current_user: dict = Depends(get_current_user)
):
    """Remove uma foto de um registro fotográfico"""
    if position not in ["front", "back", "left", "right"]:
        raise HTTPException(status_code=400, detail="Posição inválida")
    
    registry = await db.photo_registries.find_one({"id": registry_id}, {"_id": 0})
    if not registry:
        raise HTTPException(status_code=404, detail="Registro fotográfico não encontrado")
    
    field_name = f"photo_{position}"
    photo_url = registry.get(field_name)
    
    if photo_url:
        # Tentar remover arquivo físico
        try:
            file_path = UPLOADS_DIR / photo_url.replace("/api/uploads/", "")
            if file_path.exists():
                file_path.unlink()
        except Exception:
            pass
        
        # Remover referência do banco
        await db.photo_registries.update_one(
            {"id": registry_id},
            {"$set": {field_name: None, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    return {"message": "Foto removida com sucesso"}

@api_router.delete("/photo-registries/{registry_id}")
async def delete_photo_registry(
    registry_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Exclui um registro fotográfico"""
    registry = await db.photo_registries.find_one({"id": registry_id}, {"_id": 0})
    if not registry:
        raise HTTPException(status_code=404, detail="Registro fotográfico não encontrado")
    
    # Remover diretório de fotos
    try:
        photo_dir = UPLOADS_DIR / "photo_registries" / registry_id
        if photo_dir.exists():
            shutil.rmtree(photo_dir)
    except Exception:
        pass
    
    await db.photo_registries.delete_one({"id": registry_id})
    
    return {"message": "Registro fotográfico excluído com sucesso"}

# ========== VISTORIA DE CONTAINER ==========

@api_router.get("/container-inspections")
async def list_container_inspections(
    page: int = 1,
    per_page: int = 20,
    current_user: dict = Depends(get_current_user)
):
    """Lista todas as vistorias de container"""
    skip = (page - 1) * per_page
    
    total = await db.container_inspections.count_documents({})
    inspections = await db.container_inspections.find({}, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page).to_list(per_page)
    
    return {
        "items": inspections,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": (total + per_page - 1) // per_page
    }

@api_router.get("/container-inspections/{inspection_id}", response_model=ContainerInspectionResponse)
async def get_container_inspection(inspection_id: str, current_user: dict = Depends(get_current_user)):
    """Obtém uma vistoria de container pelo ID"""
    inspection = await db.container_inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not inspection:
        raise HTTPException(status_code=404, detail="Vistoria de container não encontrada")
    return ContainerInspectionResponse(**inspection)

@api_router.post("/container-inspections", response_model=ContainerInspectionResponse)
async def create_container_inspection(
    data: ContainerInspectionCreate,
    current_user: dict = Depends(get_current_user)
):
    """Cria uma nova vistoria de container"""
    # Gerar número sequencial
    counter = await db.counters.find_one_and_update(
        {"_id": "inspection_number"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    inspection_number = counter["seq"]
    
    # Buscar nomes de cliente e armador
    client_name = None
    shipping_line_name = None
    
    if data.client_id:
        client = await db.clients.find_one({"id": data.client_id}, {"_id": 0, "name": 1})
        client_name = client.get("name") if client else None
    
    if data.shipping_line_id:
        shipping_line = await db.shipping_lines.find_one({"id": data.shipping_line_id}, {"_id": 0, "name": 1})
        shipping_line_name = shipping_line.get("name") if shipping_line else None
    
    inspection = ContainerInspection(
        inspection_number=inspection_number,
        container_number=data.container_number,
        container_seal=data.container_seal,
        collection_terminal=data.collection_terminal,
        booking=data.booking,
        client_id=data.client_id,
        client_name=client_name,
        shipping_line_id=data.shipping_line_id,
        shipping_line_name=shipping_line_name,
        observations=data.observations,
        created_by=current_user["sub"],
        created_by_name=current_user["name"]
    )
    
    inspection_dict = inspection.model_dump()
    inspection_dict["created_at"] = inspection_dict["created_at"].isoformat()
    
    await db.container_inspections.insert_one(inspection_dict)
    
    return ContainerInspectionResponse(**inspection_dict)

@api_router.put("/container-inspections/{inspection_id}", response_model=ContainerInspectionResponse)
async def update_container_inspection(
    inspection_id: str,
    data: ContainerInspectionUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Atualiza uma vistoria de container"""
    inspection = await db.container_inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not inspection:
        raise HTTPException(status_code=404, detail="Vistoria de container não encontrada")
    
    update_data = {}
    
    if data.container_number is not None:
        update_data["container_number"] = data.container_number
    
    if data.container_seal is not None:
        update_data["container_seal"] = data.container_seal
    
    if data.collection_terminal is not None:
        update_data["collection_terminal"] = data.collection_terminal
    
    if data.booking is not None:
        update_data["booking"] = data.booking
    
    if data.observations is not None:
        update_data["observations"] = data.observations
    
    if data.client_id is not None:
        update_data["client_id"] = data.client_id
        if data.client_id:
            client = await db.clients.find_one({"id": data.client_id}, {"_id": 0, "name": 1})
            update_data["client_name"] = client.get("name") if client else None
        else:
            update_data["client_name"] = None
    
    if data.shipping_line_id is not None:
        update_data["shipping_line_id"] = data.shipping_line_id
        if data.shipping_line_id:
            shipping_line = await db.shipping_lines.find_one({"id": data.shipping_line_id}, {"_id": 0, "name": 1})
            update_data["shipping_line_name"] = shipping_line.get("name") if shipping_line else None
        else:
            update_data["shipping_line_name"] = None
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.container_inspections.update_one({"id": inspection_id}, {"$set": update_data})
    
    updated = await db.container_inspections.find_one({"id": inspection_id}, {"_id": 0})
    return ContainerInspectionResponse(**updated)

@api_router.post("/container-inspections/{inspection_id}/upload-photo")
async def upload_container_inspection_photo(
    inspection_id: str,
    position: str = Query(..., regex="^(front|back|left|right|internal)$"),
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Faz upload de uma foto para uma vistoria de container"""
    inspection = await db.container_inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not inspection:
        raise HTTPException(status_code=404, detail="Vistoria de container não encontrada")
    
    # Criar diretório
    photo_dir = UPLOADS_DIR / "container_inspections" / inspection_id
    photo_dir.mkdir(parents=True, exist_ok=True)
    
    # Salvar arquivo
    file_ext = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    file_path = photo_dir / f"{position}.{file_ext}"
    
    with open(file_path, "wb") as buffer:
        content = await file.read()
        buffer.write(content)
    
    # Atualizar banco
    photo_url = f"/api/uploads/container_inspections/{inspection_id}/{position}.{file_ext}"
    await db.container_inspections.update_one(
        {"id": inspection_id},
        {"$set": {f"photo_{position}": photo_url, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"url": photo_url, "position": position}

@api_router.delete("/container-inspections/{inspection_id}/photo/{position}")
async def delete_container_inspection_photo(
    inspection_id: str,
    position: str,
    current_user: dict = Depends(get_current_user)
):
    """Remove uma foto de uma vistoria de container"""
    inspection = await db.container_inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not inspection:
        raise HTTPException(status_code=404, detail="Vistoria de container não encontrada")
    
    photo_field = f"photo_{position}"
    if photo_field not in inspection or not inspection[photo_field]:
        raise HTTPException(status_code=404, detail="Foto não encontrada")
    
    # Remover arquivo
    photo_dir = UPLOADS_DIR / "container_inspections" / inspection_id
    for ext in ["jpg", "jpeg", "png", "webp"]:
        file_path = photo_dir / f"{position}.{ext}"
        if file_path.exists():
            file_path.unlink()
    
    # Atualizar banco
    await db.container_inspections.update_one(
        {"id": inspection_id},
        {"$set": {photo_field: None, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "Foto removida com sucesso"}

@api_router.delete("/container-inspections/{inspection_id}")
async def delete_container_inspection(
    inspection_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Exclui uma vistoria de container"""
    inspection = await db.container_inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not inspection:
        raise HTTPException(status_code=404, detail="Vistoria de container não encontrada")
    
    # Remover diretório de fotos
    try:
        photo_dir = UPLOADS_DIR / "container_inspections" / inspection_id
        if photo_dir.exists():
            shutil.rmtree(photo_dir)
    except Exception:
        pass
    
    await db.container_inspections.delete_one({"id": inspection_id})
    
    return {"message": "Vistoria de container excluída com sucesso"}

# ========== FLEX TANK (CONTROLE DE ESTOQUE DE BOLSAS) ==========

@api_router.get("/flex-tank/movements")
async def list_flex_tank_movements(
    page: int = 1,
    per_page: int = 20,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    client_id: Optional[str] = None,
    movement_number: Optional[int] = None,
    movement_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lista todas as movimentações de Flex Tank com filtros"""
    query = {}
    
    # Filtro por data
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter["$gte"] = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        if end_date:
            date_filter["$lte"] = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        if date_filter:
            query["movement_date"] = date_filter
    
    # Filtro por cliente
    if client_id:
        query["client_id"] = client_id
    
    # Filtro por número de registro
    if movement_number:
        query["movement_number"] = movement_number
    
    # Filtro por tipo
    if movement_type:
        query["movement_type"] = movement_type
    
    skip = (page - 1) * per_page
    
    total = await db.flex_tank_movements.count_documents(query)
    movements = await db.flex_tank_movements.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page).to_list(per_page)
    
    return {
        "items": movements,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": (total + per_page - 1) // per_page
    }

@api_router.get("/flex-tank/stock")
async def get_flex_tank_stock(
    client_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Retorna o resumo do estoque de Flex Tank"""
    query = {}
    if client_id:
        query["client_id"] = client_id
    
    # Buscar todas as movimentações
    movements = await db.flex_tank_movements.find(query, {"_id": 0}).to_list(10000)
    
    # Calcular totais
    total_entries = sum(1 for m in movements if m.get("movement_type") == "ENTRADA")
    total_exits = sum(1 for m in movements if m.get("movement_type") == "SAIDA")
    total_bags = total_entries - total_exits
    
    # Agrupar por cliente
    client_stock = {}
    for m in movements:
        cid = m.get("client_id") or "sem_cliente"
        cname = m.get("client_name") or "Sem Cliente"
        if cid not in client_stock:
            client_stock[cid] = {"client_id": cid, "client_name": cname, "entries": 0, "exits": 0, "stock": 0}
        if m.get("movement_type") == "ENTRADA":
            client_stock[cid]["entries"] += 1
        else:
            client_stock[cid]["exits"] += 1
        client_stock[cid]["stock"] = client_stock[cid]["entries"] - client_stock[cid]["exits"]
    
    # Agrupar por tamanho
    size_stock = {}
    for m in movements:
        size = m.get("bag_size") or "Não informado"
        if size not in size_stock:
            size_stock[size] = {"size": size, "entries": 0, "exits": 0, "stock": 0}
        if m.get("movement_type") == "ENTRADA":
            size_stock[size]["entries"] += 1
        else:
            size_stock[size]["exits"] += 1
        size_stock[size]["stock"] = size_stock[size]["entries"] - size_stock[size]["exits"]
    
    return {
        "total_bags": total_bags,
        "total_entries": total_entries,
        "total_exits": total_exits,
        "by_client": list(client_stock.values()),
        "by_size": list(size_stock.values())
    }

@api_router.get("/flex-tank/movements/{movement_id}", response_model=FlexTankMovementResponse)
async def get_flex_tank_movement(movement_id: str, current_user: dict = Depends(get_current_user)):
    """Obtém uma movimentação de Flex Tank pelo ID"""
    movement = await db.flex_tank_movements.find_one({"id": movement_id}, {"_id": 0})
    if not movement:
        raise HTTPException(status_code=404, detail="Movimentação não encontrada")
    return FlexTankMovementResponse(**movement)

@api_router.post("/flex-tank/movements", response_model=FlexTankMovementResponse)
async def create_flex_tank_movement(
    data: FlexTankMovementCreate,
    current_user: dict = Depends(get_current_user)
):
    """Cria uma nova movimentação de Flex Tank"""
    # Gerar número sequencial
    counter = await db.counters.find_one_and_update(
        {"_id": "flex_tank_movement_number"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    movement_number = counter["seq"]
    
    # Buscar nome do cliente
    client_name = None
    if data.client_id:
        client = await db.clients.find_one({"id": data.client_id}, {"_id": 0, "name": 1})
        client_name = client.get("name") if client else None
    
    # Buscar nome do cliente destino
    destination_client_name = None
    if data.destination_client_id:
        dest_client = await db.clients.find_one({"id": data.destination_client_id}, {"_id": 0, "name": 1})
        destination_client_name = dest_client.get("name") if dest_client else None
    
    movement = FlexTankMovement(
        movement_number=movement_number,
        bag_number=data.bag_number,
        bag_size=data.bag_size,
        movement_date=data.movement_date,
        movement_type=data.movement_type,
        client_id=data.client_id,
        client_name=client_name,
        destination_client_id=data.destination_client_id,
        destination_client_name=destination_client_name,
        container_number=data.container_number,
        observations=data.observations,
        created_by=current_user["sub"],
        created_by_name=current_user["name"]
    )
    
    movement_dict = movement.model_dump()
    movement_dict["created_at"] = movement_dict["created_at"].isoformat()
    movement_dict["movement_date"] = movement_dict["movement_date"].isoformat()
    
    await db.flex_tank_movements.insert_one(movement_dict)
    
    return FlexTankMovementResponse(**movement_dict)

@api_router.put("/flex-tank/movements/{movement_id}", response_model=FlexTankMovementResponse)
async def update_flex_tank_movement(
    movement_id: str,
    data: FlexTankMovementUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Atualiza uma movimentação de Flex Tank"""
    movement = await db.flex_tank_movements.find_one({"id": movement_id}, {"_id": 0})
    if not movement:
        raise HTTPException(status_code=404, detail="Movimentação não encontrada")
    
    update_data = {}
    
    if data.bag_number is not None:
        update_data["bag_number"] = data.bag_number
    
    if data.bag_size is not None:
        update_data["bag_size"] = data.bag_size
    
    if data.movement_date is not None:
        update_data["movement_date"] = data.movement_date.isoformat()
    
    if data.movement_type is not None:
        update_data["movement_type"] = data.movement_type
    
    if data.container_number is not None:
        update_data["container_number"] = data.container_number
    
    if data.observations is not None:
        update_data["observations"] = data.observations
    
    if data.client_id is not None:
        update_data["client_id"] = data.client_id
        if data.client_id:
            client = await db.clients.find_one({"id": data.client_id}, {"_id": 0, "name": 1})
            update_data["client_name"] = client.get("name") if client else None
        else:
            update_data["client_name"] = None
    
    if data.destination_client_id is not None:
        update_data["destination_client_id"] = data.destination_client_id
        if data.destination_client_id:
            dest_client = await db.clients.find_one({"id": data.destination_client_id}, {"_id": 0, "name": 1})
            update_data["destination_client_name"] = dest_client.get("name") if dest_client else None
        else:
            update_data["destination_client_name"] = None
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.flex_tank_movements.update_one({"id": movement_id}, {"$set": update_data})
    
    updated = await db.flex_tank_movements.find_one({"id": movement_id}, {"_id": 0})
    return FlexTankMovementResponse(**updated)

@api_router.delete("/flex-tank/movements/{movement_id}")
async def delete_flex_tank_movement(
    movement_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Exclui uma movimentação de Flex Tank"""
    movement = await db.flex_tank_movements.find_one({"id": movement_id}, {"_id": 0})
    if not movement:
        raise HTTPException(status_code=404, detail="Movimentação não encontrada")
    
    await db.flex_tank_movements.delete_one({"id": movement_id})
    
    return {"message": "Movimentação excluída com sucesso"}

@api_router.get("/flex-tank/report/excel")
async def download_flex_tank_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    client_id: Optional[str] = None,
    movement_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Gera relatório Excel das movimentações de Flex Tank"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    query = {}
    
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter["$gte"] = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        if end_date:
            date_filter["$lte"] = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        if date_filter:
            query["movement_date"] = date_filter
    
    if client_id:
        query["client_id"] = client_id
    
    if movement_type:
        query["movement_type"] = movement_type
    
    movements = await db.flex_tank_movements.find(query, {"_id": 0}).sort("movement_date", -1).to_list(10000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimentações Flex Tank"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    entrada_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    saida_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Cabeçalhos - adicionado Cliente Destino
    headers = ["Nº Registro", "Nº Bolsa", "Tamanho", "Data", "Tipo", "Cliente", "Cliente Destino", "Container", "Observações"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Congelar primeira linha
    ws.freeze_panes = 'A2'
    
    # Dados
    for row, m in enumerate(movements, 2):
        # Nº Registro
        cell = ws.cell(row=row, column=1, value=m.get("movement_number"))
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        
        # Nº Bolsa
        cell = ws.cell(row=row, column=2, value=m.get("bag_number"))
        cell.border = thin_border
        
        # Tamanho
        cell = ws.cell(row=row, column=3, value=m.get("bag_size"))
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        
        # Data
        movement_date = m.get("movement_date")
        if isinstance(movement_date, str):
            movement_date = datetime.fromisoformat(movement_date.replace('Z', '+00:00'))
        cell = ws.cell(row=row, column=4, value=movement_date.strftime("%d/%m/%Y") if movement_date else "")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        
        # Tipo (com formatação condicional)
        tipo = m.get("movement_type")
        cell = ws.cell(row=row, column=5, value=tipo)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if tipo == "ENTRADA":
            cell.fill = entrada_fill
        elif tipo == "SAIDA":
            cell.fill = saida_fill
        
        # Cliente
        cell = ws.cell(row=row, column=6, value=m.get("client_name") or "-")
        cell.border = thin_border
        
        # Cliente Destino
        cell = ws.cell(row=row, column=7, value=m.get("destination_client_name") or "-")
        cell.border = thin_border
        
        # Container
        cell = ws.cell(row=row, column=8, value=m.get("container_number") or "-")
        cell.border = thin_border
        
        # Observações
        cell = ws.cell(row=row, column=9, value=m.get("observations") or "-")
        cell.border = thin_border
        
        # Linha alternada
        if row % 2 == 0:
            for col in [1, 2, 3, 4, 6, 7, 8, 9]:  # Não aplicar na coluna Tipo (5)
                ws.cell(row=row, column=col).fill = alt_fill
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 35
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 35
    
    # Salvar em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"relatorio_flex_tank_{now_brt().strftime('%d-%m-%Y_%H-%M')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== FROTA - CADASTRO DE VEÍCULOS ====================

from models import Vehicle, VehicleCreate, VehicleUpdate, VehicleResponse

@api_router.get("/vehicles")
async def get_vehicles(
    search: Optional[str] = None,
    vehicle_type: Optional[str] = None,
    status: Optional[str] = None,
    page: int = 1,
    per_page: int = 20,
    current_user: dict = Depends(get_current_user)
):
    """Lista todos os veículos cadastrados"""
    query = {}
    
    if search:
        query["$or"] = [
            {"plate": {"$regex": search, "$options": "i"}},
            {"model": {"$regex": search, "$options": "i"}},
            {"brand": {"$regex": search, "$options": "i"}}
        ]
    
    if vehicle_type:
        query["vehicle_type"] = vehicle_type
    
    if status:
        query["status"] = status
    
    total = await db.vehicles.count_documents(query)
    skip = (page - 1) * per_page
    
    cursor = db.vehicles.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page)
    vehicles = await cursor.to_list(length=per_page)
    
    return {
        "items": vehicles,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page
    }


@api_router.get("/vehicles/{vehicle_id}", response_model=VehicleResponse)
async def get_vehicle(vehicle_id: str, current_user: dict = Depends(get_current_user)):
    """Busca veículo por ID"""
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Veículo não encontrado")
    return vehicle


@api_router.post("/vehicles", response_model=VehicleResponse)
async def create_vehicle(data: VehicleCreate, current_user: dict = Depends(get_current_user)):
    """Cadastra novo veículo"""
    # Verificar se placa já existe
    existing = await db.vehicles.find_one({"plate": data.plate.upper()})
    if existing:
        raise HTTPException(status_code=400, detail="Placa já cadastrada")
    
    vehicle_data = {
        "id": str(uuid.uuid4()),
        "plate": data.plate.upper(),
        "model": data.model,
        "brand": data.brand,
        "year": data.year,
        "vehicle_type": data.vehicle_type.upper(),
        "status": data.status.upper(),
        "observations": data.observations,
        "created_by": current_user["sub"],
        "created_by_name": current_user.get("name", "Sistema"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": None
    }
    
    await db.vehicles.insert_one(vehicle_data)
    vehicle_data.pop("_id", None)
    
    return vehicle_data


@api_router.put("/vehicles/{vehicle_id}", response_model=VehicleResponse)
async def update_vehicle(vehicle_id: str, data: VehicleUpdate, current_user: dict = Depends(get_current_user)):
    """Atualiza veículo"""
    vehicle = await db.vehicles.find_one({"id": vehicle_id})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Veículo não encontrado")
    
    update_data = {}
    if data.plate is not None:
        # Verificar se placa já existe em outro veículo
        existing = await db.vehicles.find_one({"plate": data.plate.upper(), "id": {"$ne": vehicle_id}})
        if existing:
            raise HTTPException(status_code=400, detail="Placa já cadastrada em outro veículo")
        update_data["plate"] = data.plate.upper()
    if data.model is not None:
        update_data["model"] = data.model
    if data.brand is not None:
        update_data["brand"] = data.brand
    if data.year is not None:
        update_data["year"] = data.year
    if data.vehicle_type is not None:
        update_data["vehicle_type"] = data.vehicle_type.upper()
    if data.status is not None:
        update_data["status"] = data.status.upper()
    if data.observations is not None:
        update_data["observations"] = data.observations
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.vehicles.update_one({"id": vehicle_id}, {"$set": update_data})
    
    updated = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    return updated


@api_router.delete("/vehicles/{vehicle_id}")
async def delete_vehicle(vehicle_id: str, current_user: dict = Depends(get_current_user)):
    """Exclui veículo"""
    vehicle = await db.vehicles.find_one({"id": vehicle_id})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Veículo não encontrado")
    
    await db.vehicles.delete_one({"id": vehicle_id})
    return {"message": "Veículo excluído com sucesso"}


@api_router.get("/vehicles/types/list")
async def get_vehicle_types(current_user: dict = Depends(get_current_user)):
    """Lista tipos de veículos disponíveis"""
    return [
        {"value": "CAMINHÃO", "label": "Caminhão"},
        {"value": "CARRETA", "label": "Carreta"},
        {"value": "CAVALO", "label": "Cavalo Mecânico"},
        {"value": "EMPILHADEIRA", "label": "Empilhadeira"},
        {"value": "GUINDASTE", "label": "Guindaste"},
        {"value": "REACH_STACKER", "label": "Reach Stacker"},
        {"value": "EQUIPAMENTO", "label": "Outro Equipamento"},
    ]


# ==================== FROTA - CONTROLE DE REVISÃO ====================

@api_router.get("/vehicle-revisions")
async def get_vehicle_revisions(
    vehicle_plate: Optional[str] = None,
    page: int = 1,
    per_page: int = 20,
    current_user: dict = Depends(get_current_user)
):
    """Lista todas as revisões de veículos"""
    query = {}
    if vehicle_plate:
        query["vehicle_plate"] = {"$regex": vehicle_plate.upper(), "$options": "i"}
    
    skip = (page - 1) * per_page
    
    total = await db.vehicle_revisions.count_documents(query)
    revisions = await db.vehicle_revisions.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page).to_list(per_page)
    
    return {
        "items": revisions,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page
    }

@api_router.get("/vehicle-revisions/{revision_id}", response_model=VehicleRevisionResponse)
async def get_vehicle_revision(
    revision_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Busca uma revisão específica"""
    revision = await db.vehicle_revisions.find_one({"id": revision_id}, {"_id": 0})
    if not revision:
        raise HTTPException(status_code=404, detail="Revisão não encontrada")
    return VehicleRevisionResponse(**revision)

@api_router.post("/vehicle-revisions", response_model=VehicleRevisionResponse)
async def create_vehicle_revision(
    data: VehicleRevisionCreate,
    current_user: dict = Depends(get_current_user)
):
    """Cria uma nova revisão de veículo"""
    counter = await db.counters.find_one_and_update(
        {"_id": "vehicle_revision_number"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    revision_number = counter["seq"]
    
    revision = VehicleRevision(
        revision_number=revision_number,
        vehicle_plate=data.vehicle_plate.upper(),
        vehicle_model=data.vehicle_model,
        revision_date=data.revision_date,
        oil_used=data.oil_used,
        current_km=data.current_km,
        next_oil_motor_km=data.next_oil_motor_km,
        next_oil_filter_km=data.next_oil_filter_km,
        next_air_filter_km=data.next_air_filter_km,
        next_ac_filter_km=data.next_ac_filter_km,
        next_fuel_filter_km=data.next_fuel_filter_km,
        next_racor_filter_km=data.next_racor_filter_km,
        next_apu_filter_km=data.next_apu_filter_km,
        next_hydraulic_filter_km=data.next_hydraulic_filter_km,
        next_gearbox_oil_km=data.next_gearbox_oil_km,
        next_differential_oil_km=data.next_differential_oil_km,
        next_lubrication_km=data.next_lubrication_km,
        next_washing_km=data.next_washing_km,
        mechanic_name=data.mechanic_name,
        performed_by=data.performed_by,
        observations=data.observations,
        created_by=current_user["sub"],
        created_by_name=current_user["name"]
    )
    
    revision_dict = revision.model_dump()
    revision_dict["revision_date"] = revision_dict["revision_date"].isoformat()
    revision_dict["created_at"] = revision_dict["created_at"].isoformat()
    
    await db.vehicle_revisions.insert_one(revision_dict)
    revision_dict.pop('_id', None)
    
    return VehicleRevisionResponse(**revision_dict)

@api_router.delete("/vehicle-revisions/{revision_id}")
async def delete_vehicle_revision(
    revision_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Exclui uma revisão"""
    result = await db.vehicle_revisions.delete_one({"id": revision_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Revisão não encontrada")
    return {"message": "Revisão excluída com sucesso"}


@api_router.put("/vehicle-revisions/{revision_id}", response_model=VehicleRevisionResponse)
async def update_vehicle_revision(
    revision_id: str,
    data: VehicleRevisionCreate,
    current_user: dict = Depends(get_current_user)
):
    """Atualiza uma revisão existente."""
    existing = await db.vehicle_revisions.find_one({"id": revision_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Revisão não encontrada")

    update_data = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}

    # Atualizar dados do veículo (placa/modelo) caso vehicle_id tenha mudado
    if 'vehicle_id' in update_data and update_data['vehicle_id']:
        vehicle = await db.vehicles.find_one({"id": update_data['vehicle_id']}, {"_id": 0})
        if vehicle:
            update_data['vehicle_plate'] = vehicle.get('plate', existing.get('vehicle_plate'))
            update_data['vehicle_model'] = vehicle.get('model', existing.get('vehicle_model'))

    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    await db.vehicle_revisions.update_one({"id": revision_id}, {"$set": update_data})

    updated = await db.vehicle_revisions.find_one({"id": revision_id}, {"_id": 0})
    return VehicleRevisionResponse(**updated)


@api_router.get("/vehicle-revisions/{revision_id}/pdf")
async def generate_revision_pdf(
    revision_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Gera PDF da revisão - Layout profissional igual ao comprovante de movimentação"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import requests
    
    revision = await db.vehicle_revisions.find_one({"id": revision_id}, {"_id": 0})
    if not revision:
        raise HTTPException(status_code=404, detail="Revisão não encontrada")
    
    buffer = io.BytesIO()
    
    # Cores corporativas (igual aos outros relatórios)
    PRIMARY_COLOR = "008B7B"
    HEADER_BG_COLOR = "E8F4F5"
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15*mm,
        leftMargin=15*mm,
        topMargin=15*mm,
        bottomMargin=15*mm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # ========== DOWNLOAD LOGO ==========
    logo_buffer = None
    LOGO_URL = os.environ.get('LOGO_URL', "https://customer-assets.emergentagent.com/job_da181895-6b28-4daf-bef5-4444909581e8/artifacts/i8vfweuv_logo.png")
    try:
        response = requests.get(LOGO_URL, timeout=5)
        if response.status_code == 200:
            logo_buffer = io.BytesIO(response.content)
    except:
        pass
    
    # ========== HEADER SECTION ==========
    company_style = ParagraphStyle(
        'CompanyName',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.black,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
        leading=16
    )
    address_style = ParagraphStyle(
        'Address',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.black,
        alignment=TA_CENTER,
        leading=10
    )
    
    logo_cell = ""
    if logo_buffer:
        try:
            logo_cell = Image(logo_buffer, width=50, height=50)
        except:
            pass
    
    company_info = [
        Paragraph("J.A LOGÍSTICA E ARMAZENAGEM LTDA", company_style),
        Paragraph("CNPJ: 58.180.321/0001-03", address_style),
        Paragraph("Rodovia CE-155, 16226 - Distrito Industrial", address_style),
        Paragraph("São Gonçalo do Amarante - CE", address_style),
        Paragraph("operacional@jalogisticas.com | (85) 9 9175-1472", address_style),
    ]
    
    header_data = [[logo_cell, company_info, ""]]
    header_table = Table(header_data, colWidths=[60, 400, 60])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 5))
    
    # Linha separadora
    line_table = Table([[""]], colWidths=[520])
    line_table.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (-1, 0), 2, colors.HexColor(f'#{PRIMARY_COLOR}')),
    ]))
    elements.append(line_table)
    elements.append(Spacer(1, 10))
    
    # ========== TÍTULO ==========
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Normal'],
        fontSize=16,
        textColor=colors.HexColor(f'#{PRIMARY_COLOR}'),
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
        spaceAfter=15
    )
    elements.append(Paragraph("CONTROLE DE REVISÃO", title_style))
    
    # ========== INFO BAR ==========
    rev_date = parse_datetime_value(revision['revision_date'])
    info_text = f"Revisão Nº {revision['revision_number']}  |  Veículo: {revision['vehicle_plate']}  |  Data: {rev_date.strftime('%d/%m/%Y')}"
    
    info_style = ParagraphStyle(
        'InfoBar',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor(f'#{PRIMARY_COLOR}'),
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    info_data = [[Paragraph(info_text, info_style)]]
    info_table = Table(info_data, colWidths=[520])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(f'#{HEADER_BG_COLOR}')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor(f'#{PRIMARY_COLOR}')),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 15))
    
    # ========== DADOS DO VEÍCULO ==========
    label_style = ParagraphStyle('Label', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold')
    value_style = ParagraphStyle('Value', parent=styles['Normal'], fontSize=9, fontName='Helvetica')
    
    km = f"{revision['current_km']:,}".replace(",", ".")
    
    vehicle_data = [
        [Paragraph("VEÍCULO:", label_style), Paragraph(revision['vehicle_plate'], value_style),
         Paragraph("MODELO:", label_style), Paragraph(revision.get('vehicle_model') or '-', value_style)],
        [Paragraph("ÓLEO UTILIZADO:", label_style), Paragraph(revision['oil_used'], value_style),
         Paragraph("KM ATUAL:", label_style), Paragraph(f"{km} KM", value_style)],
        [Paragraph("MECÂNICO:", label_style), Paragraph(revision['mechanic_name'], value_style),
         Paragraph("REALIZADO POR:", label_style), Paragraph(revision.get('performed_by') or '-', value_style)],
    ]
    
    vehicle_table = Table(vehicle_data, colWidths=[100, 160, 100, 160])
    vehicle_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F5F5F5')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#F5F5F5')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(vehicle_table)
    elements.append(Spacer(1, 20))
    
    # ========== PRÓXIMA REVISÃO - TÍTULO ==========
    section_title_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.white,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    section_data = [[Paragraph("PRÓXIMA REVISÃO - QUILOMETRAGEM", section_title_style)]]
    section_table = Table(section_data, colWidths=[520])
    section_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(f'#{PRIMARY_COLOR}')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(section_table)
    
    # ========== TABELA DE PRÓXIMAS REVISÕES ==========
    def format_km(value):
        if value:
            return f"{value:,} KM".replace(",", ".")
        return "-"
    
    next_rev_data = [
        ["ITEM", "PRÓXIMA TROCA (KM)"],
        ["Óleo Motor", format_km(revision.get('next_oil_motor_km'))],
        ["Filtro de Óleo", format_km(revision.get('next_oil_filter_km'))],
        ["Filtro de Ar", format_km(revision.get('next_air_filter_km'))],
        ["Filtro Ar Condicionado", format_km(revision.get('next_ac_filter_km'))],
        ["Filtro de Combustível", format_km(revision.get('next_fuel_filter_km'))],
        ["Filtro Racor", format_km(revision.get('next_racor_filter_km'))],
        ["Filtro APU", format_km(revision.get('next_apu_filter_km'))],
        ["Filtro Hidráulico", format_km(revision.get('next_hydraulic_filter_km'))],
        ["Óleo Caixa de Marcha", format_km(revision.get('next_gearbox_oil_km'))],
        ["Óleo Diferencial", format_km(revision.get('next_differential_oil_km'))],
        ["Lubrificação", format_km(revision.get('next_lubrication_km'))],
        ["Lavagem", format_km(revision.get('next_washing_km'))],
    ]
    
    next_table = Table(next_rev_data, colWidths=[300, 220])
    next_table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{HEADER_BG_COLOR}')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor(f'#{PRIMARY_COLOR}')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        # Borders and padding
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor(f'#{PRIMARY_COLOR}')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        # Alternating rows
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F8F8')]),
    ]))
    elements.append(next_table)
    elements.append(Spacer(1, 20))
    
    # ========== OBSERVAÇÕES ==========
    if revision.get('observations'):
        obs_title_style = ParagraphStyle(
            'ObsTitle',
            parent=styles['Normal'],
            fontSize=10,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor(f'#{PRIMARY_COLOR}')
        )
        obs_style = ParagraphStyle(
            'Obs',
            parent=styles['Normal'],
            fontSize=9,
            fontName='Helvetica'
        )
        elements.append(Paragraph("OBSERVAÇÕES:", obs_title_style))
        elements.append(Spacer(1, 5))
        elements.append(Paragraph(revision.get('observations', ''), obs_style))
        elements.append(Spacer(1, 15))
    
    # ========== RODAPÉ COM ASSINATURA ==========
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    
    created_at = parse_datetime_value(revision['created_at']).astimezone(timezone(timedelta(hours=-3)))
    elements.append(Spacer(1, 30))

    # Linha de assinatura
    sig_data = [
        ["_" * 50, "_" * 50],
        ["Responsável pela Revisão", "Conferido por"]
    ]
    sig_table = Table(sig_data, colWidths=[260, 260])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, 1), 5),
    ]))
    elements.append(sig_table)
    elements.append(Spacer(1, 20))
    
    elements.append(Paragraph(f"Registrado por: {revision['created_by_name']} em {created_at.strftime('%d/%m/%Y às %H:%M')}", footer_style))
    elements.append(Paragraph("ContainerLogix - J.A Logística", footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"revisao_{revision['vehicle_plate']}_{revision['revision_number']}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})

@api_router.get("/vehicles/plates")
async def get_vehicle_plates(current_user: dict = Depends(get_current_user)):
    """Lista todas as placas de veículos"""
    movements_plates = await db.movements.distinct("truck_plate")
    revisions_plates = await db.vehicle_revisions.distinct("vehicle_plate")
    all_plates = list(set([p for p in movements_plates if p] + [p for p in revisions_plates if p]))
    all_plates.sort()
    return all_plates


# ==================== OPERACIONAL - PROGRAMAÇÃO DE CARREGAMENTO ====================

from models import LoadingSchedule, LoadingScheduleCreate, LoadingScheduleResponse, LoadingScheduleItem

@api_router.get("/loading-schedules")
async def get_loading_schedules(
    search: Optional[str] = None,
    status: Optional[str] = None,
    page: int = 1,
    per_page: int = 20,
    current_user: dict = Depends(get_current_user)
):
    """Lista todas as programações de carregamento"""
    query = {}
    
    if search:
        query["$or"] = [
            {"destination_client_name": {"$regex": search, "$options": "i"}},
            {"contracting_client_name": {"$regex": search, "$options": "i"}},
            {"items.driver_name": {"$regex": search, "$options": "i"}},
            {"items.container_number": {"$regex": search, "$options": "i"}}
        ]
    
    if status:
        query["status"] = status
    
    total = await db.loading_schedules.count_documents(query)
    skip = (page - 1) * per_page
    
    cursor = db.loading_schedules.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page)
    schedules = await cursor.to_list(length=per_page)
    
    return {
        "items": schedules,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page
    }


@api_router.get("/loading-schedules/{schedule_id}", response_model=LoadingScheduleResponse)
async def get_loading_schedule(schedule_id: str, current_user: dict = Depends(get_current_user)):
    """Busca programação por ID"""
    schedule = await db.loading_schedules.find_one({"id": schedule_id}, {"_id": 0})
    if not schedule:
        raise HTTPException(status_code=404, detail="Programação não encontrada")
    return schedule


@api_router.post("/loading-schedules", response_model=LoadingScheduleResponse)
async def create_loading_schedule(data: LoadingScheduleCreate, current_user: dict = Depends(get_current_user)):
    """Cria nova programação de carregamento"""
    counter = await db.counters.find_one_and_update(
        {"_id": "loading_schedule_number"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    schedule_number = counter["seq"]
    
    schedule_data = {
        "id": str(uuid.uuid4()),
        "schedule_number": schedule_number,
        "destination_client_id": data.destination_client_id,
        "destination_client_name": data.destination_client_name,
        "contracting_client_id": data.contracting_client_id,
        "contracting_client_name": data.contracting_client_name,
        "booking": data.booking,
        "voyage": data.voyage,
        "items": [item.model_dump() for item in data.items],
        "status": "ATIVO",
        "observations": data.observations,
        "created_by": current_user["sub"],
        "created_by_name": current_user.get("name", "Sistema"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": None
    }
    
    await db.loading_schedules.insert_one(schedule_data)
    schedule_data.pop("_id", None)
    
    return schedule_data


@api_router.put("/loading-schedules/{schedule_id}", response_model=LoadingScheduleResponse)
async def update_loading_schedule(schedule_id: str, data: LoadingScheduleCreate, current_user: dict = Depends(get_current_user)):
    """Atualiza programação de carregamento"""
    schedule = await db.loading_schedules.find_one({"id": schedule_id})
    if not schedule:
        raise HTTPException(status_code=404, detail="Programação não encontrada")
    
    update_data = {
        "destination_client_id": data.destination_client_id,
        "destination_client_name": data.destination_client_name,
        "contracting_client_id": data.contracting_client_id,
        "contracting_client_name": data.contracting_client_name,
        "booking": data.booking,
        "voyage": data.voyage,
        "items": [item.model_dump() for item in data.items],
        "observations": data.observations,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.loading_schedules.update_one({"id": schedule_id}, {"$set": update_data})
    updated = await db.loading_schedules.find_one({"id": schedule_id}, {"_id": 0})
    return updated


@api_router.delete("/loading-schedules/{schedule_id}")
async def delete_loading_schedule(schedule_id: str, current_user: dict = Depends(get_current_user)):
    """Exclui programação de carregamento"""
    schedule = await db.loading_schedules.find_one({"id": schedule_id})
    if not schedule:
        raise HTTPException(status_code=404, detail="Programação não encontrada")
    
    await db.loading_schedules.delete_one({"id": schedule_id})
    return {"message": "Programação excluída com sucesso"}


@api_router.put("/loading-schedules/{schedule_id}/update-status")
async def update_loading_schedule_status(schedule_id: str, new_status: str, current_user: dict = Depends(get_current_user)):
    """Atualiza status da programação"""
    if new_status not in ["ATIVO", "CONCLUIDO", "CANCELADO"]:
        raise HTTPException(status_code=400, detail="Status inválido")
    
    result = await db.loading_schedules.update_one(
        {"id": schedule_id},
        {"$set": {"status": new_status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Programação não encontrada")
    
    return {"message": "Status atualizado"}


@api_router.get("/loading-schedules/{schedule_id}/pdf")
async def generate_loading_schedule_pdf(schedule_id: str, current_user: dict = Depends(get_current_user)):
    """Gera PDF da programação de carregamento - Layout similar ao comprovante de movimentação"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.graphics.barcode import code128
    from reportlab.graphics.shapes import Drawing
    import requests
    
    schedule = await db.loading_schedules.find_one({"id": schedule_id}, {"_id": 0})
    if not schedule:
        raise HTTPException(status_code=404, detail="Programação não encontrada")
    
    buffer = io.BytesIO()
    
    # Cores
    BLACK = colors.black
    BORDER_COLOR = colors.black
    HEADER_BG = colors.HexColor('#F5F5F5')
    PRIMARY_GREEN = colors.HexColor('#008B7B')
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=12*mm,
        leftMargin=12*mm,
        topMargin=10*mm,
        bottomMargin=10*mm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Download logo
    logo_buffer = None
    LOGO_URL = os.environ.get('LOGO_URL', "https://customer-assets.emergentagent.com/job_da181895-6b28-4daf-bef5-4444909581e8/artifacts/i8vfweuv_logo.png")
    try:
        response = requests.get(LOGO_URL, timeout=5)
        if response.status_code == 200:
            logo_buffer = io.BytesIO(response.content)
    except:
        pass
    
    # ========== HEADER com Logo, Empresa, Código de Barras ==========
    company_style = ParagraphStyle('Company', parent=styles['Normal'], fontSize=18, fontName='Helvetica-Bold', alignment=TA_CENTER, textColor=PRIMARY_GREEN, leading=20)
    slogan_style = ParagraphStyle('Slogan', parent=styles['Normal'], fontSize=9, fontName='Helvetica', alignment=TA_CENTER, textColor=PRIMARY_GREEN, leading=11)
    address_style = ParagraphStyle('Address', parent=styles['Normal'], fontSize=8, fontName='Helvetica', alignment=TA_CENTER, textColor=BLACK, leading=10)
    info_right_style = ParagraphStyle('InfoRight', parent=styles['Normal'], fontSize=8, fontName='Helvetica', alignment=TA_RIGHT, textColor=BLACK)
    
    # Logo
    logo_cell = ""
    if logo_buffer:
        try:
            logo_cell = Image(logo_buffer, width=45, height=45)
        except:
            pass
    
    # Informações da empresa (centro)
    company_text = Paragraph("J.A LOGÍSTICA", company_style)
    slogan_text = Paragraph("LOGÍSTICA E ARMAZENAGEM", slogan_style)
    address_text = Paragraph("Rodovia CE-155, 16226 - Industrial - CEP: 61668-150 - Caucaia/CE", address_style)
    
    center_content = [[company_text], [slogan_text], [address_text]]
    center_table = Table(center_content, colWidths=[400])
    center_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    
    # Código de barras e informações (direita)
    barcode_value = f"PROG{schedule['schedule_number']:06d}"
    barcode = code128.Code128(barcode_value, barWidth=1.2, barHeight=30)
    
    # Converter para horário de Brasília (UTC-3)
    from zoneinfo import ZoneInfo
    created_at = parse_datetime_value(schedule['created_at'])
    brasilia_tz = ZoneInfo('America/Sao_Paulo')
    created_at_brasilia = created_at.astimezone(brasilia_tz)
    date_str = created_at_brasilia.strftime('%d/%m/%Y')
    time_str = created_at_brasilia.strftime('%H:%M')
    
    # Abreviar nome do criador (primeiro e segundo nome, ignorando preposições)
    full_creator_name = schedule.get('created_by_name', 'Sistema')
    if full_creator_name:
        name_parts = full_creator_name.strip().split()
        preposicoes = ['DE', 'DA', 'DO', 'DOS', 'DAS', 'E']
        nomes_filtrados = [p for p in name_parts if p.upper() not in preposicoes]
        if len(nomes_filtrados) >= 2:
            creator_short_name = f"{nomes_filtrados[0]} {nomes_filtrados[1]}"
        elif len(nomes_filtrados) == 1:
            creator_short_name = nomes_filtrados[0]
        else:
            creator_short_name = ' '.join(name_parts[:2]) if len(name_parts) >= 2 else name_parts[0] if name_parts else 'Sistema'
    else:
        creator_short_name = 'Sistema'
    
    barcode_info = Paragraph(f"<b>Nº {schedule['schedule_number']}</b>", ParagraphStyle('BarcodeNum', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', alignment=TA_CENTER))
    date_info = Paragraph(f"Data: {date_str}", info_right_style)
    user_info = Paragraph(f"Criado por: {creator_short_name}", info_right_style)
    
    right_content = [[barcode], [barcode_info], [date_info], [user_info]]
    right_table = Table(right_content, colWidths=[150])
    right_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    
    # Montar header completo
    header_data = [[logo_cell, center_table, right_table]]
    header_table = Table(header_data, colWidths=[55, 450, 160])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
    ]))
    elements.append(header_table)
    
    # Linha separadora verde
    elements.append(Spacer(1, 5))
    line_data = [[""]]
    line_table = Table(line_data, colWidths=[700])
    line_table.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, -1), 2, PRIMARY_GREEN),
    ]))
    elements.append(line_table)
    elements.append(Spacer(1, 10))
    
    # ========== TÍTULO: Box com borda ==========
    title_style = ParagraphStyle('Title', parent=styles['Normal'], fontSize=12, fontName='Helvetica-Bold', alignment=TA_CENTER, textColor=PRIMARY_GREEN)
    
    title_content = [[Paragraph("PROGRAMAÇÃO DE CARREGAMENTO", title_style)]]
    title_table = Table(title_content, colWidths=[700])
    title_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 2, BORDER_COLOR),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(title_table)
    elements.append(Spacer(1, 10))
    
    # ========== BOX 1: Informações dos Clientes ==========
    label_style = ParagraphStyle('Label', parent=styles['Normal'], fontSize=8, fontName='Helvetica', textColor=BLACK)
    value_style = ParagraphStyle('Value', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', textColor=BLACK)
    section_title = ParagraphStyle('SectionTitle', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', textColor=BLACK)
    
    # Header da seção
    client_header = [[Paragraph("Informações dos Clientes", section_title)]]
    client_header_table = Table(client_header, colWidths=[700])
    client_header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), HEADER_BG),
        ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(client_header_table)
    
    # Conteúdo - Linha 1: Clientes
    client_row1 = [
        [Paragraph("Cliente Contratante", label_style), Paragraph(schedule['contracting_client_name'], value_style)],
        [Paragraph("Cliente Destino", label_style), Paragraph(schedule['destination_client_name'], value_style)]
    ]
    # Conteúdo - Linha 2: Booking e Viagem
    booking_value = schedule.get('booking') or '-'
    voyage_value = schedule.get('voyage') or '-'
    client_row2 = [
        [Paragraph("Booking", label_style), Paragraph(booking_value, value_style)],
        [Paragraph("Viagem", label_style), Paragraph(voyage_value, value_style)]
    ]
    
    client_content = [client_row1, client_row2]
    client_table = Table(client_content, colWidths=[350, 350])
    client_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
        ('LINEAFTER', (0, 0), (0, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('LINEBELOW', (0, 0), (-1, 0), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(client_table)
    elements.append(Spacer(1, 10))
    
    # ========== BOX 2: Tabela de Programações ==========
    prog_header = [[Paragraph("Itens da Programação", section_title)]]
    prog_header_table = Table(prog_header, colWidths=[700])
    prog_header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), HEADER_BG),
        ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(prog_header_table)
    
    # Tabela de dados
    table_header = ["#", "MOTORISTA", "CPF", "CAVALO", "CARRETA", "LOCAL DE CARREG.", "DATA", "CONTAINER", "LACRE"]
    table_data = [table_header]
    
    for idx, item in enumerate(schedule['items'], 1):
        loading_date = item.get('loading_date', '')
        if loading_date:
            try:
                dt = datetime.fromisoformat(loading_date.replace('Z', '+00:00'))
                loading_date = dt.strftime('%d/%m/%Y')
            except:
                pass
        
        # Pegar primeiro e segundo nome do motorista (ignorando preposições)
        driver_full_name = item.get('driver_name', '-')
        if driver_full_name and driver_full_name != '-':
            name_parts = driver_full_name.strip().split()
            # Filtrar preposições comuns
            preposicoes = ['DE', 'DA', 'DO', 'DOS', 'DAS', 'E']
            nomes_filtrados = [p for p in name_parts if p.upper() not in preposicoes]
            
            if len(nomes_filtrados) >= 2:
                driver_display_name = f"{nomes_filtrados[0]} {nomes_filtrados[1]}"
            elif len(nomes_filtrados) == 1:
                driver_display_name = nomes_filtrados[0]
            else:
                # Se só tem preposições, usa os dois primeiros
                driver_display_name = ' '.join(name_parts[:2]) if len(name_parts) >= 2 else name_parts[0] if name_parts else '-'
        else:
            driver_display_name = '-'
        
        row = [
            str(idx),
            driver_display_name,
            item.get('driver_cpf', '-') or '-',
            item.get('cavalo_plate', '-'),
            item.get('carreta_plate', '-') or '-',
            item.get('loading_location', '-'),
            loading_date or '-',
            item.get('container_number', '-') or '-',
            item.get('seal_number', '-') or '-'
        ]
        table_data.append(row)
    
    col_widths = [20, 110, 80, 60, 60, 125, 65, 100, 80]  # Total = 700 para alinhar com cabeçalho
    main_table = Table(table_data, colWidths=col_widths)
    main_table.setStyle(TableStyle([
        # Header row - verde padrão
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY_GREEN),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # # column
        ('ALIGN', (6, 1), (6, -1), 'CENTER'),  # Data column
        # Borders
        ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
    ]))
    elements.append(main_table)
    elements.append(Spacer(1, 12))
    
    # ========== BOX 3: Observações (se houver) ==========
    if schedule.get('observations'):
        obs_header = [[Paragraph("Observações", section_title)]]
        obs_header_table = Table(obs_header, colWidths=[700])
        obs_header_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), HEADER_BG),
            ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(obs_header_table)
        
        obs_content_style = ParagraphStyle('ObsContent', parent=styles['Normal'], fontSize=9, fontName='Helvetica', textColor=BLACK)
        obs_content = [[Paragraph(schedule['observations'], obs_content_style)]]
        obs_table = Table(obs_content, colWidths=[700])
        obs_table.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(obs_table)
        elements.append(Spacer(1, 12))
    
    # ========== RODAPÉ ==========
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
    elements.append(Spacer(1, 15))
    elements.append(Paragraph("ContainerLogix - J.A Logística", footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"programacao_carregamento_{schedule['schedule_number']}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})


# ==================== STATUS DE ENTREGA ENDPOINTS ====================

from models import DeliveryStatus, DeliveryStatusCreate, DeliveryStatusResponse, DeliveryStatusItem

@api_router.get("/delivery-status")
async def get_delivery_statuses(
    page: int = 1,
    per_page: int = 20,
    status: Optional[str] = None,
    schedule_number: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lista todos os status de entrega com paginação e filtros"""
    query = {}
    if status:
        query["status"] = status
    if schedule_number:
        query["schedule_number"] = schedule_number
    
    total = await db.delivery_statuses.count_documents(query)
    skip = (page - 1) * per_page
    
    statuses = await db.delivery_statuses.find(query, {"_id": 0}).sort("status_number", -1).skip(skip).limit(per_page).to_list(per_page)
    
    return {
        "items": statuses,
        "total": total,
        "page": page,
        "pages": (total + per_page - 1) // per_page
    }

@api_router.get("/delivery-status/schedule/{schedule_number}")
async def get_schedule_for_delivery_status(schedule_number: int, current_user: dict = Depends(get_current_user)):
    """Busca uma programação de carregamento pelo número para criar um status de entrega"""
    schedule = await db.loading_schedules.find_one({"schedule_number": schedule_number}, {"_id": 0})
    if not schedule:
        raise HTTPException(status_code=404, detail="Programação não encontrada")
    return schedule

@api_router.get("/delivery-status/{status_id}", response_model=DeliveryStatusResponse)
async def get_delivery_status(status_id: str, current_user: dict = Depends(get_current_user)):
    """Busca um status de entrega pelo ID"""
    status = await db.delivery_statuses.find_one({"id": status_id}, {"_id": 0})
    if not status:
        raise HTTPException(status_code=404, detail="Status de entrega não encontrado")
    return status

@api_router.post("/delivery-status", response_model=DeliveryStatusResponse)
async def create_delivery_status(data: DeliveryStatusCreate, current_user: dict = Depends(get_current_user)):
    """Cria um novo status de entrega baseado em uma programação"""
    # Buscar a programação de carregamento
    schedule = await db.loading_schedules.find_one({"schedule_number": data.schedule_number}, {"_id": 0})
    if not schedule:
        raise HTTPException(status_code=404, detail="Programação não encontrada")
    
    # Gerar próximo número sequencial
    last = await db.delivery_statuses.find_one({}, {"_id": 0, "status_number": 1}, sort=[("status_number", -1)])
    next_number = (last["status_number"] + 1) if last and "status_number" in last else 1
    
    # Criar o status com dados da programação
    status = DeliveryStatus(
        status_number=next_number,
        schedule_id=schedule["id"],
        schedule_number=schedule["schedule_number"],
        destination_client_name=schedule["destination_client_name"],
        contracting_client_name=schedule["contracting_client_name"],
        booking=schedule.get("booking"),
        voyage=schedule.get("voyage"),
        status_date=data.status_date,
        items=[item.model_dump() if hasattr(item, 'model_dump') else item for item in data.items],
        observations=data.observations,
        created_by=current_user["sub"],
        created_by_name=current_user["name"]
    )
    
    await db.delivery_statuses.insert_one(status.model_dump())
    
    result = await db.delivery_statuses.find_one({"id": status.id}, {"_id": 0})
    return result

@api_router.put("/delivery-status/{status_id}", response_model=DeliveryStatusResponse)
async def update_delivery_status(status_id: str, data: DeliveryStatusCreate, current_user: dict = Depends(get_current_user)):
    """Atualiza um status de entrega existente"""
    existing = await db.delivery_statuses.find_one({"id": status_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Status de entrega não encontrado")
    
    update_data = {
        "status_date": data.status_date,
        "items": [item.model_dump() for item in data.items],
        "observations": data.observations,
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.delivery_statuses.update_one({"id": status_id}, {"$set": update_data})
    
    result = await db.delivery_statuses.find_one({"id": status_id}, {"_id": 0})
    return result

@api_router.delete("/delivery-status/{status_id}")
async def delete_delivery_status(status_id: str, current_user: dict = Depends(get_current_user)):
    """Deleta um status de entrega"""
    result = await db.delivery_statuses.delete_one({"id": status_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Status de entrega não encontrado")
    return {"message": "Status de entrega deletado com sucesso"}

@api_router.put("/delivery-status/{status_id}/update-status")
async def update_delivery_status_status(status_id: str, new_status: str, current_user: dict = Depends(get_current_user)):
    """Atualiza o status (ATIVO, CONCLUIDO, CANCELADO)"""
    if new_status not in ["ATIVO", "CONCLUIDO", "CANCELADO"]:
        raise HTTPException(status_code=400, detail="Status inválido")
    
    result = await db.delivery_statuses.update_one(
        {"id": status_id},
        {"$set": {"status": new_status, "updated_at": datetime.now(timezone.utc)}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Status de entrega não encontrado")
    return {"message": "Status atualizado com sucesso"}

@api_router.get("/delivery-status/{status_id}/pdf")
async def generate_delivery_status_pdf(status_id: str, current_user: dict = Depends(get_current_user)):
    """Gera PDF do status de entrega - Layout similar à programação de carregamento"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.graphics.barcode import code128
    import requests
    
    delivery_status = await db.delivery_statuses.find_one({"id": status_id}, {"_id": 0})
    if not delivery_status:
        raise HTTPException(status_code=404, detail="Status de entrega não encontrado")
    
    buffer = io.BytesIO()
    
    # Cores
    BLACK = colors.black
    BORDER_COLOR = colors.black
    HEADER_BG = colors.HexColor('#F5F5F5')
    PRIMARY_GREEN = colors.HexColor('#008B7B')
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=12*mm,
        leftMargin=12*mm,
        topMargin=10*mm,
        bottomMargin=10*mm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Download logo
    logo_buffer = None
    LOGO_URL = os.environ.get('LOGO_URL', "https://customer-assets.emergentagent.com/job_da181895-6b28-4daf-bef5-4444909581e8/artifacts/i8vfweuv_logo.png")
    try:
        response = requests.get(LOGO_URL, timeout=5)
        if response.status_code == 200:
            logo_buffer = io.BytesIO(response.content)
    except:
        pass
    
    # ========== HEADER ==========
    company_style = ParagraphStyle('Company', parent=styles['Normal'], fontSize=18, fontName='Helvetica-Bold', alignment=TA_CENTER, textColor=PRIMARY_GREEN, leading=20)
    slogan_style = ParagraphStyle('Slogan', parent=styles['Normal'], fontSize=9, fontName='Helvetica', alignment=TA_CENTER, textColor=PRIMARY_GREEN, leading=11)
    address_style = ParagraphStyle('Address', parent=styles['Normal'], fontSize=8, fontName='Helvetica', alignment=TA_CENTER, textColor=BLACK, leading=10)
    info_right_style = ParagraphStyle('InfoRight', parent=styles['Normal'], fontSize=8, fontName='Helvetica', alignment=TA_RIGHT, textColor=BLACK)
    
    # Logo
    logo_cell = ""
    if logo_buffer:
        try:
            logo_cell = Image(logo_buffer, width=45, height=45)
        except:
            pass
    
    # Informações da empresa (centro)
    company_text = Paragraph("J.A LOGÍSTICA", company_style)
    slogan_text = Paragraph("LOGÍSTICA E ARMAZENAGEM", slogan_style)
    address_text = Paragraph("Rodovia CE-155, 16226 - Industrial - CEP: 61668-150 - Caucaia/CE", address_style)
    
    center_content = [[company_text], [slogan_text], [address_text]]
    center_table = Table(center_content, colWidths=[400])
    center_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    
    # Código de barras e informações (direita)
    barcode_value = f"ENTR{delivery_status['status_number']:06d}"
    barcode = code128.Code128(barcode_value, barWidth=1.2, barHeight=30)
    
    # Converter para horário de Brasília
    from zoneinfo import ZoneInfo
    created_at = parse_datetime_value(delivery_status['created_at'])
    brasilia_tz = ZoneInfo('America/Sao_Paulo')
    created_at_brasilia = created_at.astimezone(brasilia_tz)
    date_str = created_at_brasilia.strftime('%d/%m/%Y')
    
    # Abreviar nome do criador
    full_creator_name = delivery_status.get('created_by_name', 'Sistema')
    if full_creator_name:
        name_parts = full_creator_name.strip().split()
        preposicoes = ['DE', 'DA', 'DO', 'DOS', 'DAS', 'E']
        nomes_filtrados = [p for p in name_parts if p.upper() not in preposicoes]
        if len(nomes_filtrados) >= 2:
            creator_short_name = f"{nomes_filtrados[0]} {nomes_filtrados[1]}"
        elif len(nomes_filtrados) == 1:
            creator_short_name = nomes_filtrados[0]
        else:
            creator_short_name = ' '.join(name_parts[:2]) if len(name_parts) >= 2 else name_parts[0] if name_parts else 'Sistema'
    else:
        creator_short_name = 'Sistema'
    
    barcode_info = Paragraph(f"<b>Nº {delivery_status['status_number']}</b>", ParagraphStyle('BarcodeNum', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', alignment=TA_CENTER))
    date_info = Paragraph(f"Data: {date_str}", info_right_style)
    user_info = Paragraph(f"Criado por: {creator_short_name}", info_right_style)
    
    right_content = [[barcode], [barcode_info], [date_info], [user_info]]
    right_table = Table(right_content, colWidths=[150])
    right_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    
    # Montar header completo
    header_data = [[logo_cell, center_table, right_table]]
    header_table = Table(header_data, colWidths=[55, 450, 160])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
    ]))
    elements.append(header_table)
    
    # Linha separadora verde
    elements.append(Spacer(1, 5))
    line_data = [[""]]
    line_table = Table(line_data, colWidths=[700])
    line_table.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, -1), 2, PRIMARY_GREEN),
    ]))
    elements.append(line_table)
    elements.append(Spacer(1, 10))
    
    # ========== TÍTULO ==========
    title_style = ParagraphStyle('Title', parent=styles['Normal'], fontSize=12, fontName='Helvetica-Bold', alignment=TA_CENTER, textColor=PRIMARY_GREEN)
    
    title_content = [[Paragraph("STATUS DE ENTREGA", title_style)]]
    title_table = Table(title_content, colWidths=[700])
    title_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 2, BORDER_COLOR),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(title_table)
    elements.append(Spacer(1, 10))
    
    # ========== BOX 1: Informações da Programação ==========
    label_style = ParagraphStyle('Label', parent=styles['Normal'], fontSize=8, fontName='Helvetica', textColor=BLACK)
    value_style = ParagraphStyle('Value', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', textColor=BLACK)
    section_title = ParagraphStyle('SectionTitle', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', textColor=BLACK)
    
    # Largura padrão para todas as seções (700px para alinhar com o título)
    SECTION_WIDTH = 700
    
    # Header da seção
    info_header = [[Paragraph("Informações da Programação", section_title)]]
    info_header_table = Table(info_header, colWidths=[SECTION_WIDTH])
    info_header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), HEADER_BG),
        ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(info_header_table)
    
    # Linha 1: Programação Ref e Data do Status
    status_date_value = delivery_status.get('status_date', '')
    if status_date_value:
        try:
            dt = datetime.fromisoformat(status_date_value.replace('Z', '+00:00'))
            status_date_value = dt.strftime('%d/%m/%Y')
        except:
            pass
    
    info_row1 = [
        [Paragraph("Programação Ref.", label_style), Paragraph(f"Nº {delivery_status['schedule_number']}", value_style)],
        [Paragraph("Data do Status", label_style), Paragraph(status_date_value, value_style)]
    ]
    # Linha 2: Clientes
    info_row2 = [
        [Paragraph("Cliente Contratante", label_style), Paragraph(delivery_status['contracting_client_name'], value_style)],
        [Paragraph("Cliente Destino", label_style), Paragraph(delivery_status['destination_client_name'], value_style)]
    ]
    # Linha 3: Booking e Viagem
    booking_value = delivery_status.get('booking') or '-'
    voyage_value = delivery_status.get('voyage') or '-'
    info_row3 = [
        [Paragraph("Booking", label_style), Paragraph(booking_value, value_style)],
        [Paragraph("Viagem", label_style), Paragraph(voyage_value, value_style)]
    ]
    
    info_content = [info_row1, info_row2, info_row3]
    info_table = Table(info_content, colWidths=[SECTION_WIDTH/2, SECTION_WIDTH/2])
    info_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
        ('LINEAFTER', (0, 0), (0, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('LINEBELOW', (0, 0), (-1, 1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 10))
    
    # ========== BOX 2: Tabela de Status de Entrega ==========
    status_header = [[Paragraph("Status de Entrega por Motorista", section_title)]]
    status_header_table = Table(status_header, colWidths=[SECTION_WIDTH])
    status_header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), HEADER_BG),
        ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(status_header_table)
    
    # Tabela de dados - headers como strings simples (igual ao PDF de referência)
    table_header = ["#", "MOTORISTA", "CPF", "CAVALO", "CONTAINER", "LOCAL", "CHEGADA", "INÍCIO", "TÉRMINO", "SAÍDA", "AGEND.", "ENTREGA"]
    table_data = [table_header]
    
    for idx, item in enumerate(delivery_status['items'], 1):
        # Abreviar nome do motorista
        driver_full_name = item.get('driver_name', '-')
        if driver_full_name and driver_full_name != '-':
            name_parts = driver_full_name.strip().split()
            preposicoes = ['DE', 'DA', 'DO', 'DOS', 'DAS', 'E']
            nomes_filtrados = [p for p in name_parts if p.upper() not in preposicoes]
            if len(nomes_filtrados) >= 2:
                driver_display_name = f"{nomes_filtrados[0]} {nomes_filtrados[1]}"
            elif len(nomes_filtrados) == 1:
                driver_display_name = nomes_filtrados[0]
            else:
                driver_display_name = ' '.join(name_parts[:2]) if len(name_parts) >= 2 else name_parts[0] if name_parts else '-'
        else:
            driver_display_name = '-'
        
        row = [
            str(idx),
            driver_display_name,
            item.get('driver_cpf', '-') or '-',
            item.get('cavalo_plate', '-') or '-',
            item.get('container_number', '-') or '-',
            item.get('loading_location', '-') or '-',
            item.get('arrival_time', '-') or '-',
            item.get('loading_start_time', '-') or '-',
            item.get('loading_end_time', '-') or '-',
            item.get('departure_time', '-') or '-',
            item.get('port_schedule_time', '-') or '-',
            item.get('delivery_completed', '-') or '-'
        ]
        table_data.append(row)
    
    # Larguras ajustadas para 700px total (12 colunas - igual ao padrão do PDF de Programação)
    # Total = 700px para alinhar perfeitamente com o cabeçalho e demais seções
    col_widths = [20, 85, 70, 50, 80, 85, 52, 52, 52, 52, 52, 50]  # Total = 700
    data_table = Table(table_data, colWidths=col_widths)
    data_table.setStyle(TableStyle([
        # Header row - verde padrão (igual ao PDF de referência)
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY_GREEN),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # # column
        ('ALIGN', (6, 1), (-1, -1), 'CENTER'),  # Horários columns
        # Borders
        ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
    ]))
    elements.append(data_table)
    elements.append(Spacer(1, 8))
    
    # ========== Observações ==========
    if delivery_status.get('observations'):
        obs_header = [[Paragraph("Observações", section_title)]]
        obs_header_table = Table(obs_header, colWidths=[SECTION_WIDTH])
        obs_header_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), HEADER_BG),
            ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(obs_header_table)
        
        obs_content_style = ParagraphStyle('ObsContent', parent=styles['Normal'], fontSize=9, fontName='Helvetica', textColor=BLACK)
        obs_content = [[Paragraph(delivery_status['observations'], obs_content_style)]]
        obs_table = Table(obs_content, colWidths=[SECTION_WIDTH])
        obs_table.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(obs_table)
        elements.append(Spacer(1, 12))
    
    # ========== Rodapé ==========
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
    elements.append(Spacer(1, 15))
    elements.append(Paragraph("ContainerLogix - J.A Logística", footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"status_entrega_{delivery_status['status_number']}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})



# ==================== SEGREGAÇÃO DE UNIDADE ENDPOINTS ====================

from models import UnitSegregation, UnitSegregationCreate, UnitSegregationUpdate, UnitSegregationResponse, UnitSegregationItem

@api_router.get("/unit-segregations")
async def get_unit_segregations(
    page: int = 1,
    per_page: int = 20,
    status: Optional[str] = None,
    client_id: Optional[str] = None,
    container_number: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lista todas as segregações de unidade com filtros"""
    query = {}
    
    if status:
        query["status"] = status
    if client_id:
        query["client_id"] = client_id
    if container_number:
        # Buscar nos itens
        query["items.container_number"] = {"$regex": container_number, "$options": "i"}
    
    total = await db.unit_segregations.count_documents(query)
    skip = (page - 1) * per_page
    
    cursor = db.unit_segregations.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page)
    items = await cursor.to_list(length=per_page)
    
    return {
        "items": items,
        "total": total,
        "page": page,
        "pages": (total + per_page - 1) // per_page
    }


@api_router.get("/unit-segregations/{segregation_id}")
async def get_unit_segregation(segregation_id: str, current_user: dict = Depends(get_current_user)):
    """Busca uma segregação específica"""
    segregation = await db.unit_segregations.find_one({"id": segregation_id}, {"_id": 0})
    if not segregation:
        raise HTTPException(status_code=404, detail="Segregação não encontrada")
    return segregation


@api_router.post("/unit-segregations", response_model=UnitSegregationResponse)
async def create_unit_segregation(data: UnitSegregationCreate, current_user: dict = Depends(get_current_user)):
    """Cria uma nova segregação de unidade com múltiplos containers"""
    
    if not data.items or len(data.items) == 0:
        raise HTTPException(status_code=400, detail="Pelo menos um container deve ser informado")
    
    # Verificar se algum container já está segregado (ativo)
    for item in data.items:
        existing = await db.unit_segregations.find_one({
            "items.container_number": item.container_number.upper(),
            "status": "ATIVO"
        })
        if existing:
            raise HTTPException(status_code=400, detail=f"Container {item.container_number} já está segregado para o cliente {existing['client_name']}")
    
    # Buscar nome do cliente
    client = await db.clients.find_one({"id": data.client_id}, {"_id": 0, "name": 1})
    if not client:
        raise HTTPException(status_code=400, detail="Cliente não encontrado")
    
    # Processar itens - buscar nomes dos armadores
    processed_items = []
    for item in data.items:
        shipowner = await db.shipping_lines.find_one({"id": item.shipping_line}, {"_id": 0, "name": 1})
        shipping_line_name = shipowner["name"] if shipowner else item.shipping_line
        processed_items.append({
            "container_number": item.container_number.upper(),
            "tare": item.tare,
            "shipping_line": item.shipping_line,
            "shipping_line_name": shipping_line_name
        })
    
    # Gerar número sequencial
    last = await db.unit_segregations.find_one(sort=[("segregation_number", -1)])
    next_number = (last["segregation_number"] + 1) if last else 1
    
    segregation = UnitSegregation(
        segregation_number=next_number,
        client_id=data.client_id,
        client_name=client["name"],
        items=processed_items,
        observations=data.observations,
        created_by=current_user["sub"],
        created_by_name=current_user["name"]
    )
    
    await db.unit_segregations.insert_one(segregation.model_dump())
    
    result = segregation.model_dump()
    return result


@api_router.put("/unit-segregations/{segregation_id}")
async def update_unit_segregation(segregation_id: str, data: UnitSegregationUpdate, current_user: dict = Depends(get_current_user)):
    """Atualiza uma segregação de unidade"""
    segregation = await db.unit_segregations.find_one({"id": segregation_id})
    if not segregation:
        raise HTTPException(status_code=404, detail="Segregação não encontrada")
    
    update_data = data.model_dump(exclude_unset=True)
    
    # Se mudou o cliente, buscar o nome
    if "client_id" in update_data:
        client = await db.clients.find_one({"id": update_data["client_id"]}, {"_id": 0, "name": 1})
        if not client:
            raise HTTPException(status_code=400, detail="Cliente não encontrado")
        update_data["client_name"] = client["name"]
    
    # Se atualizou os itens, buscar nomes dos armadores
    if "items" in update_data and update_data["items"]:
        processed_items = []
        for item in update_data["items"]:
            item_dict = item if isinstance(item, dict) else item.model_dump() if hasattr(item, 'model_dump') else dict(item)
            shipowner = await db.shipping_lines.find_one({"id": item_dict.get("shipping_line")}, {"_id": 0, "name": 1})
            shipping_line_name = shipowner["name"] if shipowner else item_dict.get("shipping_line")
            processed_items.append({
                "container_number": item_dict.get("container_number", "").upper(),
                "tare": item_dict.get("tare"),
                "shipping_line": item_dict.get("shipping_line"),
                "shipping_line_name": shipping_line_name
            })
        update_data["items"] = processed_items
    
    # Se está liberando a segregação
    if update_data.get("status") == "LIBERADO":
        update_data["released_at"] = datetime.now(timezone.utc)
        update_data["released_by"] = current_user["sub"]
        update_data["released_by_name"] = current_user["name"]
    
    await db.unit_segregations.update_one(
        {"id": segregation_id},
        {"$set": update_data}
    )
    
    updated = await db.unit_segregations.find_one({"id": segregation_id}, {"_id": 0})
    return updated


@api_router.delete("/unit-segregations/{segregation_id}")
async def delete_unit_segregation(segregation_id: str, current_user: dict = Depends(get_current_user)):
    """Exclui uma segregação de unidade"""
    result = await db.unit_segregations.delete_one({"id": segregation_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Segregação não encontrada")
    return {"message": "Segregação excluída com sucesso"}


@api_router.post("/unit-segregations/{segregation_id}/release")
async def release_unit_segregation(segregation_id: str, current_user: dict = Depends(get_current_user)):
    """Libera uma segregação de unidade"""
    segregation = await db.unit_segregations.find_one({"id": segregation_id})
    if not segregation:
        raise HTTPException(status_code=404, detail="Segregação não encontrada")
    
    if segregation["status"] != "ATIVO":
        raise HTTPException(status_code=400, detail="Segregação já foi liberada ou cancelada")
    
    await db.unit_segregations.update_one(
        {"id": segregation_id},
        {"$set": {
            "status": "LIBERADO",
            "released_at": datetime.now(timezone.utc),
            "released_by": current_user["sub"],
            "released_by_name": current_user["name"]
        }}
    )
    
    updated = await db.unit_segregations.find_one({"id": segregation_id}, {"_id": 0})
    return updated


@api_router.get("/unit-segregations/{segregation_id}/pdf")
async def get_unit_segregation_pdf(segregation_id: str, current_user: dict = Depends(get_current_user)):
    """Gera PDF da segregação de unidade - Formato Horizontal (Landscape)"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.graphics.barcode import code128
    
    segregation = await db.unit_segregations.find_one({"id": segregation_id}, {"_id": 0})
    if not segregation:
        raise HTTPException(status_code=404, detail="Segregação não encontrada")
    
    buffer = io.BytesIO()
    # Usar landscape para orientação horizontal
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=40, rightMargin=40, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    # Largura total disponível em landscape (A4 landscape = 842 x 595 pontos)
    PAGE_WIDTH = landscape(A4)[0] - 80  # 842 - 80 = 762
    SECTION_WIDTH = PAGE_WIDTH
    
    # Cores
    PRIMARY_GREEN = colors.HexColor('#047857')
    HEADER_BG = colors.HexColor('#F3F4F6')
    BORDER_COLOR = colors.HexColor('#E5E7EB')
    BLACK = colors.HexColor('#1F2937')
    
    # ========== CABEÇALHO ==========
    # Logo
    # ========== DOWNLOAD LOGO ==========
    import requests
    logo_buffer = None
    LOGO_URL = os.environ.get('LOGO_URL', "https://customer-assets.emergentagent.com/job_da181895-6b28-4daf-bef5-4444909581e8/artifacts/i8vfweuv_logo.png")
    try:
        response = requests.get(LOGO_URL, timeout=5)
        if response.status_code == 200:
            logo_buffer = io.BytesIO(response.content)
    except:
        pass
    
    # Logo
    logo_cell = ""
    if logo_buffer:
        try:
            logo_cell = Image(logo_buffer, width=50, height=50)
        except:
            logo_cell = Paragraph("", styles['Normal'])
    else:
        logo_cell = Paragraph("", styles['Normal'])
    
    # Informações centrais
    company_style = ParagraphStyle('Company', parent=styles['Normal'], fontSize=12, fontName='Helvetica-Bold', textColor=PRIMARY_GREEN, alignment=TA_CENTER)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=9, fontName='Helvetica', textColor=BLACK, alignment=TA_CENTER)
    
    center_content = [
        [Paragraph("J.A LOGÍSTICA", company_style)],
        [Paragraph("LOGÍSTICA E ARMAZENAGEM", subtitle_style)],
        [Paragraph("Rodovia CE-155, 16226 - Industrial - CEP: 61668-150 - Caucaia/CE", subtitle_style)]
    ]
    center_table = Table(center_content, colWidths=[450])
    center_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    # Informações direita (código de barras)
    info_right_style = ParagraphStyle('InfoRight', parent=styles['Normal'], fontSize=8, textColor=BLACK, alignment=TA_CENTER)
    
    barcode_value = f"SEG{segregation['segregation_number']:06d}"
    barcode = code128.Code128(barcode_value, barWidth=1.2, barHeight=30)
    
    from zoneinfo import ZoneInfo
    created_at = parse_datetime_value(segregation['created_at'])
    brasilia_tz = ZoneInfo('America/Sao_Paulo')
    created_at_brasilia = created_at.astimezone(brasilia_tz)
    date_str = created_at_brasilia.strftime('%d/%m/%Y')
    
    # Abreviar nome do criador
    full_creator_name = segregation.get('created_by_name', 'Sistema')
    if full_creator_name:
        name_parts = full_creator_name.strip().split()
        preposicoes = ['DE', 'DA', 'DO', 'DOS', 'DAS', 'E']
        nomes_filtrados = [p for p in name_parts if p.upper() not in preposicoes]
        if len(nomes_filtrados) >= 2:
            creator_short_name = f"{nomes_filtrados[0]} {nomes_filtrados[1]}"
        elif len(nomes_filtrados) == 1:
            creator_short_name = nomes_filtrados[0]
        else:
            creator_short_name = ' '.join(name_parts[:2]) if len(name_parts) >= 2 else name_parts[0] if name_parts else 'Sistema'
    else:
        creator_short_name = 'Sistema'
    
    barcode_info = Paragraph(f"<b>Nº {segregation['segregation_number']}</b>", ParagraphStyle('BarcodeNum', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', alignment=TA_CENTER))
    date_info = Paragraph(f"Data: {date_str}", info_right_style)
    user_info = Paragraph(f"Criado por: {creator_short_name}", info_right_style)
    
    right_content = [[barcode], [barcode_info], [date_info], [user_info]]
    right_table = Table(right_content, colWidths=[150])
    right_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    
    # Montar header completo
    header_data = [[logo_cell, center_table, right_table]]
    header_table = Table(header_data, colWidths=[55, SECTION_WIDTH - 215, 160])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
    ]))
    elements.append(header_table)
    
    # Linha separadora verde
    elements.append(Spacer(1, 5))
    line_data = [[""]]
    line_table = Table(line_data, colWidths=[SECTION_WIDTH])
    line_table.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, -1), 2, PRIMARY_GREEN),
    ]))
    elements.append(line_table)
    elements.append(Spacer(1, 10))
    
    # ========== TÍTULO ==========
    title_style = ParagraphStyle('Title', parent=styles['Normal'], fontSize=12, fontName='Helvetica-Bold', alignment=TA_CENTER, textColor=PRIMARY_GREEN)
    
    title_content = [[Paragraph("SEGREGAÇÃO DE UNIDADE", title_style)]]
    title_table = Table(title_content, colWidths=[SECTION_WIDTH])
    title_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 2, BORDER_COLOR),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(title_table)
    elements.append(Spacer(1, 10))
    
    # ========== INFORMAÇÕES DA SEGREGAÇÃO ==========
    label_style = ParagraphStyle('Label', parent=styles['Normal'], fontSize=8, fontName='Helvetica', textColor=BLACK)
    value_style = ParagraphStyle('Value', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', textColor=BLACK)
    section_title = ParagraphStyle('SectionTitle', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', textColor=BLACK)
    
    
    # Header da seção
    info_header = [[Paragraph("Informações da Segregação", section_title)]]
    info_header_table = Table(info_header, colWidths=[SECTION_WIDTH])
    info_header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), HEADER_BG),
        ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(info_header_table)
    
    # Linha 1: Cliente Reservado e Status
    status_value = segregation.get('status', 'ATIVO')
    items_count = len(segregation.get('items', []))
    info_row1 = [
        [Paragraph("Cliente Reservado", label_style), Paragraph(segregation['client_name'], value_style)],
        [Paragraph("Status", label_style), Paragraph(status_value, value_style)]
    ]
    # Linha 2: Quantidade de containers
    info_row2 = [
        [Paragraph("Qtd. de Containers", label_style), Paragraph(str(items_count), value_style)],
        [Paragraph("", label_style), Paragraph("", value_style)]
    ]
    
    info_content = [info_row1, info_row2]
    info_table = Table(info_content, colWidths=[SECTION_WIDTH/2, SECTION_WIDTH/2])
    info_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
        ('LINEAFTER', (0, 0), (0, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('LINEBELOW', (0, 0), (-1, 0), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 10))
    
    # ========== TABELA DE CONTAINERS ==========
    items_header = [[Paragraph("Unidades Segregadas", section_title)]]
    items_header_table = Table(items_header, colWidths=[SECTION_WIDTH])
    items_header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), HEADER_BG),
        ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(items_header_table)
    
    # Cabeçalho da tabela
    table_data = [["#", "CONTAINER", "TARA", "ARMADOR"]]
    
    for idx, item in enumerate(segregation.get('items', []), 1):
        table_data.append([
            str(idx),
            item.get('container_number', '-'),
            item.get('tare', '-') or '-',
            item.get('shipping_line_name', '') or item.get('shipping_line', '-')
        ])
    
    # Se não houver itens, mostrar mensagem
    if len(table_data) == 1:
        table_data.append(['', 'Nenhum container cadastrado', '', ''])
    
    col_widths = [40, 250, 120, SECTION_WIDTH - 410]  # Ajustado para landscape
    data_table = Table(table_data, colWidths=col_widths)
    data_table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY_GREEN),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        # Borders
        ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
    ]))
    elements.append(data_table)
    elements.append(Spacer(1, 10))
    
    # ========== Observações ==========
    if segregation.get('observations'):
        obs_header = [[Paragraph("Observações", section_title)]]
        obs_header_table = Table(obs_header, colWidths=[SECTION_WIDTH])
        obs_header_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), HEADER_BG),
            ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(obs_header_table)
        
        obs_content_style = ParagraphStyle('ObsContent', parent=styles['Normal'], fontSize=9, fontName='Helvetica', textColor=BLACK)
        obs_content = [[Paragraph(segregation['observations'], obs_content_style)]]
        obs_table = Table(obs_content, colWidths=[SECTION_WIDTH])
        obs_table.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(obs_table)
        elements.append(Spacer(1, 12))
    
    # ========== Rodapé ==========
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
    elements.append(Spacer(1, 15))
    elements.append(Paragraph("ContainerLogix - J.A Logística", footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"segregacao_unidade_{segregation['segregation_number']}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})


@api_router.get("/check-segregation/{container_number}")
async def check_container_segregation(container_number: str, current_user: dict = Depends(get_current_user)):
    """Verifica se um container está segregado"""
    segregation = await db.unit_segregations.find_one({
        "items.container_number": container_number.upper(),
        "status": "ATIVO"
    }, {"_id": 0})
    
    if segregation:
        return {
            "is_segregated": True,
            "segregation": segregation
        }
    return {"is_segregated": False}



# ==================== INVOICE ENDPOINTS ====================

# Dados fixos do recebedor (J.A Logística)
RECEIVER_DATA = {
    "company": "J.A LOGÍSTICA E ARMAZENAGEM LTDA",
    "cnpj": "58.180.321/0001-03",
    "email": "operacional@jalogisticas.com",
    "phone": "(85) 9 9175-1472",
    "address": "Rodovia CE-155, 16226 - Distrito Industrial",
    "city_state": "São Gonçalo do Amarante - CE",
    "zip": "62670-000",
    "complement": ""
}

@api_router.get("/intl-invoices")
async def get_intl_invoices(
    page: int = 1,
    per_page: int = 20,
    status: Optional[str] = None,
    currency: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Lista todas as invoices internacionais"""
    query = {}
    if status:
        query["status"] = status
    if currency:
        query["currency"] = currency
    
    total = await db.intl_invoices.count_documents(query)
    skip = (page - 1) * per_page
    
    invoices = await db.intl_invoices.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page).to_list(per_page)
    
    return {
        "items": invoices,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page
    }

@api_router.get("/intl-invoices/receiver-data")
async def get_intl_receiver_data(current_user: dict = Depends(get_current_user)):
    """Retorna dados pré-preenchidos do recebedor"""
    return RECEIVER_DATA

@api_router.get("/intl-invoices/movement/{transaction_id}")
async def get_movement_for_invoice(transaction_id: str, current_user: dict = Depends(get_current_user)):
    """Busca uma movimentação pelo número para adicionar como item na invoice"""
    # Tentar converter para inteiro se possível
    try:
        trans_id_int = int(transaction_id)
        movement = await db.movements.find_one({"transaction_id": trans_id_int}, {"_id": 0})
    except ValueError:
        movement = await db.movements.find_one({"transaction_id": transaction_id}, {"_id": 0})
    
    if not movement:
        raise HTTPException(status_code=404, detail="Movimentação não encontrada")
    
    return {
        "id": movement.get("id"),
        "transaction_id": movement.get("transaction_id"),
        "container_number": movement.get("container_number"),
        "service_type": movement.get("service_type"),
        "service_value": movement.get("service_value") or 0,
        "currency": movement.get("currency", "BRL"),
        "client_name": movement.get("client_name"),
        "operation_type": movement.get("operation_type"),
        "size_type": movement.get("size_type"),
        "shipping_line": movement.get("shipping_line"),
    }

@api_router.post("/intl-invoices")
async def create_intl_invoice(
    data: IntlInvoiceCreate,
    current_user: dict = Depends(get_current_user)
):
    """Cria uma nova invoice internacional"""
    
    # Gerar número sequencial
    counter = await db.counters.find_one_and_update(
        {"_id": "intl_invoice_number"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    invoice_number = counter.get("seq", 1)
    
    # Calcular totais
    subtotal = sum(item.quantity * item.unit_price for item in data.items)
    total = subtotal
    
    # Preparar itens
    items_data = []
    for item in data.items:
        items_data.append({
            "description": item.description,
            "quantity": item.quantity,
            "unit_price": item.unit_price,
            "total": item.quantity * item.unit_price
        })
    
    invoice_data = {
        "id": str(uuid.uuid4()),
        "invoice_number": invoice_number,
        "receiver_company": RECEIVER_DATA["company"],
        "receiver_cnpj": RECEIVER_DATA["cnpj"],
        "receiver_email": RECEIVER_DATA["email"],
        "receiver_phone": RECEIVER_DATA["phone"],
        "receiver_address": RECEIVER_DATA["address"],
        "receiver_city_state": RECEIVER_DATA["city_state"],
        "receiver_zip": RECEIVER_DATA["zip"],
        "receiver_complement": RECEIVER_DATA["complement"],
        "payer_client_id": data.payer_client_id,
        "payer_company": data.payer_company,
        "payer_cnpj": data.payer_cnpj,
        "payer_contact": data.payer_contact,
        "payer_email": data.payer_email,
        "payer_address": data.payer_address,
        "issue_date": data.issue_date,
        "due_date": data.due_date,
        "currency": data.currency,
        "items": items_data,
        "subtotal": subtotal,
        "total": total,
        "notes": data.notes,
        "status": "EMITIDA",
        "created_by": current_user["sub"],
        "created_by_name": current_user.get("name", "Sistema"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": None
    }
    
    await db.intl_invoices.insert_one(invoice_data)
    
    # Remover _id inserido pelo MongoDB antes de retornar
    invoice_data.pop("_id", None)
    
    return invoice_data

@api_router.get("/intl-invoices/{invoice_id}")
async def get_intl_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
    """Retorna uma invoice internacional específica"""
    invoice = await db.intl_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice não encontrada")
    return invoice

@api_router.put("/intl-invoices/{invoice_id}/status")
async def update_intl_invoice_status(
    invoice_id: str,
    status: str,
    current_user: dict = Depends(get_current_user)
):
    """Atualiza o status de uma invoice internacional"""
    if status not in ["EMITIDA", "PAGA", "CANCELADA"]:
        raise HTTPException(status_code=400, detail="Status inválido")
    
    result = await db.intl_invoices.update_one(
        {"id": invoice_id},
        {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Invoice não encontrada")
    
    return {"message": "Status atualizado com sucesso"}

@api_router.delete("/intl-invoices/{invoice_id}")
async def delete_intl_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
    """Exclui uma invoice internacional"""
    result = await db.intl_invoices.delete_one({"id": invoice_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Invoice não encontrada")
    return {"message": "Invoice excluída com sucesso"}

@api_router.put("/intl-invoices/{invoice_id}")
async def update_intl_invoice(
    invoice_id: str,
    data: IntlInvoiceCreate,
    current_user: dict = Depends(get_current_user)
):
    """Atualiza uma invoice internacional"""
    invoice = await db.intl_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice não encontrada")
    
    # Calcular totais
    subtotal = sum(item.quantity * item.unit_price for item in data.items)
    total = subtotal
    
    # Preparar itens
    items_data = []
    for item in data.items:
        items_data.append({
            "description": item.description,
            "quantity": item.quantity,
            "unit_price": item.unit_price,
            "total": item.quantity * item.unit_price
        })
    
    update_data = {
        "payer_client_id": data.payer_client_id,
        "payer_company": data.payer_company,
        "payer_cnpj": data.payer_cnpj,
        "payer_contact": data.payer_contact,
        "payer_email": data.payer_email,
        "payer_address": data.payer_address,
        "issue_date": data.issue_date,
        "due_date": data.due_date,
        "currency": data.currency,
        "items": items_data,
        "subtotal": subtotal,
        "total": total,
        "notes": data.notes,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.intl_invoices.update_one(
        {"id": invoice_id},
        {"$set": update_data}
    )
    
    # Buscar invoice atualizada
    updated_invoice = await db.intl_invoices.find_one({"id": invoice_id}, {"_id": 0})
    return updated_invoice

@api_router.get("/intl-invoices/{invoice_id}/pdf")
async def generate_intl_invoice_pdf(invoice_id: str, current_user: dict = Depends(get_current_user)):
    """Gera PDF da invoice internacional"""
    from reports import generate_intl_invoice_pdf as gen_pdf
    
    invoice = await db.intl_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice não encontrada")
    
    pdf_buffer = gen_pdf(invoice)
    
    return StreamingResponse(
        io.BytesIO(pdf_buffer),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=invoice_{invoice['invoice_number']}.pdf"}
    )


# ==================== FATURAS DOMÉSTICAS (BILLING) ====================
# Os endpoints de faturas domésticas permanecem em /api/billing-invoices


# ==================== RPA TERCEIRO ====================
from models import RPATerceiro, RPATerceiroCreate, RPATerceiroUpdate, RPATerceiroResponse, RPAServiceItem


def _rpa_calc_balance(rpa: dict) -> float:
    """Calcula SALDO A RECEBER: service_value + daily + fuel + others - advance - discounts"""
    sv = float(rpa.get('service_value') or 0)
    dr = float(rpa.get('daily_rate') or 0)
    fu = float(rpa.get('fuel') or 0)
    ad = float(rpa.get('advance') or 0)
    ot = float(rpa.get('others') or 0)
    de = float(rpa.get('discounts') or 0)
    return round(sv + dr + fu + ot - ad - de, 2)


def _rpa_serialize(rpa: dict) -> dict:
    """Adiciona balance calculado ao dict de RPA para resposta."""
    out = {**rpa}
    out['balance'] = _rpa_calc_balance(rpa)
    return out


@api_router.get("/rpa-terceiro", response_model=List[RPATerceiroResponse])
async def list_rpa_terceiro(
    search: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    rpa_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Listar RPAs (todos, sem paginação)."""
    query = {}
    if rpa_type:
        # Aceita 'terceiro' ou 'agregado'. Documentos antigos sem campo são tratados como 'terceiro'.
        if rpa_type == "terceiro":
            query["$or"] = [{"rpa_type": "terceiro"}, {"rpa_type": {"$exists": False}}]
        else:
            query["rpa_type"] = rpa_type
    if search:
        search_or = [
            {"driver_name": {"$regex": search, "$options": "i"}},
            {"client_name": {"$regex": search, "$options": "i"}},
            {"container_number": {"$regex": search, "$options": "i"}},
            {"truck_plate": {"$regex": search, "$options": "i"}},
            {"bank_beneficiary": {"$regex": search, "$options": "i"}},
        ]
        if "$or" in query:
            # Combinar com filtro de tipo
            existing_or = query.pop("$or")
            query["$and"] = [{"$or": existing_or}, {"$or": search_or}]
        else:
            query["$or"] = search_or
    if date_from:
        query.setdefault("created_at", {})["$gte"] = date_from
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"

    rpas = await db.rpa_terceiro.find(query, {"_id": 0}).sort("rpa_number", -1).to_list(None)
    return [_rpa_serialize(r) for r in rpas]


@api_router.get("/rpa-terceiro/next-number")
async def get_next_rpa_number(
    rpa_type: Optional[str] = "terceiro",
    current_user: dict = Depends(get_current_user)
):
    """Próximo número sequencial do RPA (separado por tipo)."""
    if rpa_type == "terceiro":
        type_filter = {"$or": [{"rpa_type": "terceiro"}, {"rpa_type": {"$exists": False}}]}
    else:
        type_filter = {"rpa_type": rpa_type}
    last = await db.rpa_terceiro.find_one(type_filter, sort=[("rpa_number", -1)])
    next_num = (last["rpa_number"] + 1) if last else 1
    return {"next_number": next_num}


@api_router.get("/rpa-terceiro/driver-info/{driver_id}")
async def get_rpa_driver_info(driver_id: str, current_user: dict = Depends(get_current_user)):
    """Retorna info do motorista + última movimentação para autopreencher RPA."""
    driver = await db.drivers.find_one({"id": driver_id}, {"_id": 0})
    if not driver:
        raise HTTPException(status_code=404, detail="Motorista não encontrado")

    last_mov = await db.movements.find_one(
        {"driver_name": driver["name"]},
        {"_id": 0},
        sort=[("created_at", -1)]
    )

    return {
        "driver_name": driver.get("name"),
        "driver_cpf": driver.get("cpf"),
        "driver_phone": driver.get("phone"),
        "truck_plate": (last_mov or {}).get("truck_plate"),
        "trailer_plate": (last_mov or {}).get("trailer_plate_1"),
        "truck_owner": (last_mov or {}).get("transport_company"),
    }


@api_router.get("/rpa-terceiro/{rpa_id}", response_model=RPATerceiroResponse)
async def get_rpa_terceiro(rpa_id: str, current_user: dict = Depends(get_current_user)):
    rpa = await db.rpa_terceiro.find_one({"id": rpa_id}, {"_id": 0})
    if not rpa:
        raise HTTPException(status_code=404, detail="RPA não encontrado")
    return _rpa_serialize(rpa)


@api_router.post("/rpa-terceiro", response_model=RPATerceiroResponse)
async def create_rpa_terceiro(data: RPATerceiroCreate, current_user: dict = Depends(get_current_user)):
    # Próximo número - separado por tipo (terceiro / agregado)
    rpa_type = data.rpa_type or "terceiro"
    if rpa_type == "terceiro":
        type_filter = {"$or": [{"rpa_type": "terceiro"}, {"rpa_type": {"$exists": False}}]}
    else:
        type_filter = {"rpa_type": rpa_type}
    last = await db.rpa_terceiro.find_one(type_filter, sort=[("rpa_number", -1)])
    next_num = (last["rpa_number"] + 1) if last else 1

    rpa = RPATerceiro(
        rpa_number=next_num,
        **data.model_dump(),
        created_by=current_user["sub"],
        created_by_name=current_user["name"]
    )
    rpa_dict = rpa.model_dump()
    rpa_dict["created_at"] = rpa.created_at.isoformat()
    await db.rpa_terceiro.insert_one(rpa_dict)
    return _rpa_serialize(rpa_dict)


@api_router.put("/rpa-terceiro/{rpa_id}", response_model=RPATerceiroResponse)
async def update_rpa_terceiro(rpa_id: str, data: RPATerceiroUpdate, current_user: dict = Depends(get_current_user)):
    existing = await db.rpa_terceiro.find_one({"id": rpa_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="RPA não encontrado")

    update_data = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.rpa_terceiro.update_one({"id": rpa_id}, {"$set": update_data})

    updated = await db.rpa_terceiro.find_one({"id": rpa_id}, {"_id": 0})
    return _rpa_serialize(updated)


@api_router.delete("/rpa-terceiro/{rpa_id}")
async def delete_rpa_terceiro(rpa_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.rpa_terceiro.delete_one({"id": rpa_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="RPA não encontrado")
    return {"message": "RPA removido"}


@api_router.get("/rpa-terceiro/{rpa_id}/pdf")
async def download_rpa_terceiro_pdf(rpa_id: str, current_user: dict = Depends(get_current_user)):
    """Gera PDF do RPA seguindo o modelo da J.A Logística."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reports import download_logo, PRIMARY_COLOR

    rpa = await db.rpa_terceiro.find_one({"id": rpa_id}, {"_id": 0})
    if not rpa:
        raise HTTPException(status_code=404, detail="RPA não encontrado")

    rpa['balance'] = _rpa_calc_balance(rpa)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=8 * mm, bottomMargin=8 * mm
    )

    elements = []
    styles = getSampleStyleSheet()

    def money(v):
        try:
            return f"R$ {float(v or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except Exception:
            return "R$ 0,00"

    def fmt_date_br(iso_str):
        if not iso_str:
            return "-"
        try:
            return datetime.fromisoformat(iso_str).strftime('%d/%m/%Y')
        except Exception:
            return str(iso_str)

    # ============================================================
    # LAYOUT NO ESTILO EIR - Centralizado, limpo, com linhas finas
    # ============================================================
    PRIMARY = colors.HexColor(f'#{PRIMARY_COLOR}')
    RULE_COLOR = colors.HexColor('#2D3748')   # cinza-preto para linhas
    LABEL_COLOR = colors.HexColor('#4A5568')  # cinza médio para labels
    VALUE_COLOR = colors.HexColor('#1A202C')  # quase preto para valores
    MUTED_COLOR = colors.HexColor('#718096')  # cinza para textos secundários

    # ===== HEADER CENTRALIZADO (logo + nome + tagline + título) =====
    logo_buffer = download_logo()
    if logo_buffer:
        logo_img = Image(logo_buffer, width=18 * mm, height=18 * mm)
        logo_img.hAlign = 'CENTER'
        elements.append(logo_img)

    company_name_style = ParagraphStyle(
        'CompanyName', parent=styles['Normal'], fontSize=14, leading=15,
        alignment=TA_CENTER, fontName='Helvetica-Bold', textColor=PRIMARY
    )
    elements.append(Paragraph("J.A LOGÍSTICA", company_name_style))

    tagline_style = ParagraphStyle(
        'Tagline', parent=styles['Normal'], fontSize=7, leading=9,
        alignment=TA_CENTER, fontName='Helvetica-Oblique', textColor=MUTED_COLOR
    )
    elements.append(Paragraph("Logística e Armazenagem", tagline_style))
    elements.append(Spacer(1, 3))

    doc_title_style = ParagraphStyle(
        'DocTitle', parent=styles['Normal'], fontSize=11, leading=13,
        alignment=TA_CENTER, fontName='Helvetica-Bold', textColor=VALUE_COLOR
    )
    elements.append(Paragraph("RPA - RECIBO DE PAGAMENTO A AUTÔNOMO", doc_title_style))

    rpa_id_style = ParagraphStyle(
        'RPAId', parent=styles['Normal'], fontSize=8, leading=10,
        alignment=TA_CENTER, fontName='Helvetica', textColor=MUTED_COLOR
    )
    elements.append(Paragraph(f"RPA Nº #{rpa['rpa_number']}", rpa_id_style))
    elements.append(Spacer(1, 5))

    # ===== Estilos das seções (estilo EIR) =====
    section_title_style = ParagraphStyle(
        'SectionTitle', parent=styles['Normal'], fontSize=10, leading=12,
        alignment=TA_LEFT, fontName='Helvetica-Bold', textColor=VALUE_COLOR,
        spaceBefore=0, spaceAfter=2
    )
    label_style = ParagraphStyle(
        'Label', parent=styles['Normal'], fontSize=7, leading=9,
        fontName='Helvetica', textColor=LABEL_COLOR
    )
    value_style = ParagraphStyle(
        'Value', parent=styles['Normal'], fontSize=9, leading=11,
        fontName='Helvetica-Bold', textColor=VALUE_COLOR
    )

    def section_header(title: str):
        """Linha horizontal escura + título da seção (estilo EIR)"""
        rule = Table([[""]], colWidths=[190 * mm], rowHeights=[0.7])
        rule.setStyle(TableStyle([
            ('LINEBELOW', (0, 0), (-1, -1), 1.2, RULE_COLOR),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(rule)
        elements.append(Spacer(1, 2))
        elements.append(Paragraph(title, section_title_style))
        elements.append(Spacer(1, 3))

    def field_cell(label, val):
        """Célula com label em cima (cinza pequeno) e valor abaixo (preto bold)"""
        return Paragraph(
            f"<font color='#4A5568' size='7'>{label}</font><br/>"
            f"<font color='#1A202C' size='9'><b>{val if val else '-'}</b></font>",
            ParagraphStyle('Cell', parent=styles['Normal'], leading=11)
        )

    fields_grid_style = TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ])

    # ===== INFORMAÇÕES DO AUTÔNOMO =====
    section_header("Informações do Autônomo")
    auto_data = [
        [field_cell("Motorista", rpa.get('driver_name')),
         field_cell("CPF", rpa.get('driver_cpf')),
         field_cell("Telefone", rpa.get('driver_phone'))],
        [field_cell("Placa Cavalo", rpa.get('truck_plate')),
         field_cell("Renavan", rpa.get('truck_renavam')),
         field_cell("Proprietário", rpa.get('truck_owner'))],
        [field_cell("Placa Carreta", rpa.get('trailer_plate')),
         field_cell("Renavan Carreta", rpa.get('trailer_renavam')),
         field_cell("Proprietário Carreta", rpa.get('trailer_owner'))],
    ]
    auto_table = Table(auto_data, colWidths=[80 * mm, 50 * mm, 60 * mm])
    auto_table.setStyle(fields_grid_style)
    elements.append(auto_table)
    elements.append(Spacer(1, 2))

    # ===== INFORMAÇÕES DO SERVIÇO PRESTADO =====
    section_header("Informações do Serviço Prestado")
    serv_data = [
        [field_cell("Local", rpa.get('service_local')),
         field_cell("Data", fmt_date_br(rpa.get('service_date'))),
         field_cell("Serviço", rpa.get('service_type'))],
        [field_cell("Tipo (LS/RODO)", rpa.get('service_modality')),
         field_cell("Origem", rpa.get('origin')),
         field_cell("Destino", rpa.get('destination'))],
        [field_cell("CTE", rpa.get('cte')),
         field_cell("Peso", rpa.get('weight')),
         field_cell("Nº Container", rpa.get('container_number'))],
        [field_cell("Data Coleta", fmt_date_br(rpa.get('collection_date'))),
         field_cell("Data Entrega", fmt_date_br(rpa.get('delivery_date'))),
         field_cell("Cliente", rpa.get('client_name'))],
    ]
    serv_table = Table(serv_data, colWidths=[63 * mm, 63 * mm, 64 * mm])
    serv_table.setStyle(fields_grid_style)
    elements.append(serv_table)
    elements.append(Spacer(1, 2))

    # ===== DEMONSTRATIVO DOS SERVIÇOS PRESTADOS =====
    section_header("Demonstrativo dos Serviços Prestados")
    services = rpa.get('services') or []
    demo_header = [
        Paragraph("<font size='8'><b>Descrição</b></font>", styles['Normal']),
        Paragraph("<font size='8'><b>Valor</b></font>", styles['Normal']),
    ]
    demo_data = [demo_header]
    if not services:
        demo_data.append([Paragraph("<font size='9'>-</font>", styles['Normal']),
                          Paragraph(f"<font size='9'>{money(0)}</font>", styles['Normal'])])
    else:
        for s in services:
            demo_data.append([
                Paragraph(f"<font size='9'>{s.get('description') or '-'}</font>", styles['Normal']),
                Paragraph(f"<font size='9'>{money(s.get('value'))}</font>", styles['Normal']),
            ])
    demo_table = Table(demo_data, colWidths=[150 * mm, 40 * mm])
    demo_table.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (-1, 0), 0.5, RULE_COLOR),
        ('LINEBELOW', (0, 0), (-1, 0), 0.5, RULE_COLOR),
        ('LINEBELOW', (0, -1), (-1, -1), 0.5, RULE_COLOR),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(demo_table)
    elements.append(Spacer(1, 2))

    # ===== ESPECIFICAÇÃO DA REMUNERAÇÃO =====
    section_header("Especificação da Remuneração do Serviço")
    rem_data = [
        [field_cell("I. Valor do Serviço", money(rpa.get('service_value'))),
         field_cell("II. Diárias", money(rpa.get('daily_rate'))),
         field_cell("III. Abastecimento", money(rpa.get('fuel')))],
        [field_cell("IV. Adiantamento", money(rpa.get('advance'))),
         field_cell("VI. Outros", money(rpa.get('others'))),
         field_cell("Descontos", money(rpa.get('discounts')))],
    ]
    rem_table = Table(rem_data, colWidths=[63 * mm, 63 * mm, 64 * mm])
    rem_table.setStyle(fields_grid_style)
    elements.append(rem_table)
    elements.append(Spacer(1, 4))

    # ===== SALDO A RECEBER (destacado) =====
    saldo_para = Paragraph(
        f"<font size='13' color='#1A202C'><b>SALDO A RECEBER:</b></font>"
        f"<font size='14' color='#{PRIMARY_COLOR}'><b>&nbsp;&nbsp;{money(rpa['balance'])}</b></font>",
        ParagraphStyle('Saldo', parent=styles['Normal'], alignment=TA_CENTER)
    )
    saldo_table = Table([[saldo_para]], colWidths=[190 * mm])
    saldo_table.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (-1, -1), 1.2, RULE_COLOR),
        ('LINEBELOW', (0, 0), (-1, -1), 1.2, RULE_COLOR),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(saldo_table)
    elements.append(Spacer(1, 4))

    # ===== DADOS BANCÁRIOS =====
    section_header("Dados Bancários do Beneficiário")
    bank_data = [
        [field_cell("Beneficiário", rpa.get('bank_beneficiary')),
         field_cell("Nº Agência", rpa.get('bank_agency')),
         field_cell("Nº Conta", rpa.get('bank_account'))],
        [field_cell("Chave PIX", rpa.get('bank_pix')),
         Paragraph("", styles['Normal']),
         Paragraph("", styles['Normal'])],
    ]
    bank_table = Table(bank_data, colWidths=[80 * mm, 50 * mm, 60 * mm])
    bank_table.setStyle(fields_grid_style)
    elements.append(bank_table)
    elements.append(Spacer(1, 6))

    # ===== ASSINATURAS (estilo EIR: 2 colunas) =====
    sig_label_style = ParagraphStyle(
        'SigLabel', parent=styles['Normal'], fontSize=8, leading=10,
        alignment=TA_CENTER, fontName='Helvetica-Bold', textColor=VALUE_COLOR
    )
    sig_field_style = ParagraphStyle(
        'SigField', parent=styles['Normal'], fontSize=7, leading=9,
        alignment=TA_LEFT, fontName='Helvetica', textColor=LABEL_COLOR
    )

    sign_data = [
        [Paragraph("&nbsp;", styles['Normal']),
         Paragraph("&nbsp;", styles['Normal'])],
        [Paragraph("Assinatura do Motorista/Proprietário", sig_label_style),
         Paragraph("Local e Data", sig_label_style)],
        [Paragraph(
            f"<font size='7' color='#4A5568'>Nome:</font> <font size='8'><b>{rpa.get('driver_name') or '-'}</b></font><br/>"
            f"<font size='7' color='#4A5568'>CPF:</font> <font size='8'><b>{rpa.get('driver_cpf') or '-'}</b></font>",
            sig_field_style),
         Paragraph(
            f"<font size='7' color='#4A5568'>Data:</font> <font size='8'><b>{now_brt().strftime('%d/%m/%Y')}</b></font>",
            sig_field_style)],
    ]
    sign_table = Table(sign_data, colWidths=[95 * mm, 95 * mm], rowHeights=[18, None, None])
    sign_table.setStyle(TableStyle([
        ('LINEABOVE', (0, 1), (-1, 1), 0.5, RULE_COLOR),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(sign_table)
    elements.append(Spacer(1, 4))

    # ===== DECLARAÇÃO =====
    decl_style = ParagraphStyle(
        'Decl', parent=styles['Normal'], fontSize=7, leading=9,
        alignment=TA_CENTER, fontName='Helvetica-Oblique', textColor=MUTED_COLOR
    )
    elements.append(Paragraph(
        "Declaro para os devidos fins, que recebi da J.A LOGISTICA LTDA - CNPJ 58.180.321/0001-03, "
        "os valores descritos neste recibo referente aos serviços prestados por mim, sem mais nada a declarar.",
        decl_style
    ))
    elements.append(Spacer(1, 4))

    # ===== RODAPÉ (estilo EIR) =====
    footer_rule = Table([[""]], colWidths=[190 * mm], rowHeights=[0.5])
    footer_rule.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, RULE_COLOR),
    ]))
    elements.append(footer_rule)
    elements.append(Spacer(1, 3))

    # Linha do ID + obs
    note_text = rpa.get('observations') or "Somente efetuar pagamento mediante comprovante de exportação e pesagem"
    footer_id_style = ParagraphStyle(
        'FooterId', parent=styles['Normal'], fontSize=8, leading=10,
        alignment=TA_LEFT, fontName='Helvetica-Bold', textColor=VALUE_COLOR
    )
    elements.append(Paragraph(f"#{rpa['rpa_number']}", footer_id_style))

    footer_meta_style = ParagraphStyle(
        'FooterMeta', parent=styles['Normal'], fontSize=7, leading=9,
        alignment=TA_LEFT, fontName='Helvetica', textColor=MUTED_COLOR
    )
    elements.append(Paragraph(
        f"Usuário: {rpa.get('created_by_name') or '-'}", footer_meta_style
    ))
    elements.append(Paragraph(
        f"Data e hora da impressão: {now_brt().strftime('%d/%m/%Y %H:%M')}", footer_meta_style
    ))
    elements.append(Spacer(1, 4))

    elements.append(Paragraph(
        f"<i>* {note_text}</i>",
        ParagraphStyle('Note', parent=styles['Normal'], fontSize=7, leading=9,
                       alignment=TA_LEFT, textColor=MUTED_COLOR)
    ))
    elements.append(Spacer(1, 3))

    elements.append(Paragraph(
        "<b>J.A LOGÍSTICA</b> - Logística e Armazenagem | Este documento é válido como recibo de pagamento",
        ParagraphStyle('FooterCo', parent=styles['Normal'], fontSize=7, leading=9,
                       alignment=TA_CENTER, textColor=MUTED_COLOR)
    ))

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()

    filename = f"RPA_{rpa['rpa_number']}_{(rpa.get('driver_name') or 'motorista').upper().replace(' ', '_')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== ORDEM DE SERVIÇO ====================
from models import OrdemServico, OrdemServicoCreate, OrdemServicoUpdate, OrdemServicoResponse, OSItem


def _os_calc_item_total(item: dict) -> float:
    qty = float(item.get('quantity') or 0)
    unit_price = float(item.get('unit_price') or 0)
    discount = float(item.get('discount') or 0)
    return round(qty * unit_price - discount, 2)


def _os_serialize(os_doc: dict) -> dict:
    out = {**os_doc}
    products = out.get('products') or []
    services = out.get('services') or []
    for p in products:
        p['total'] = _os_calc_item_total(p)
    for s in services:
        s['total'] = _os_calc_item_total(s)
    out['products_total'] = round(sum(p.get('total', 0) for p in products), 2)
    out['services_total'] = round(sum(s.get('total', 0) for s in services), 2)
    out['grand_total'] = round(out['products_total'] + out['services_total'], 2)
    return out


@api_router.get("/ordem-servico", response_model=List[OrdemServicoResponse])
async def list_ordem_servico(
    search: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query['status'] = status
    if search:
        query["$or"] = [
            {"person_name": {"$regex": search, "$options": "i"}},
            {"equipment_plate": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}},
            {"category": {"$regex": search, "$options": "i"}},
        ]
    rows = await db.ordem_servico.find(query, {"_id": 0}).sort("os_number", -1).to_list(None)
    return [_os_serialize(r) for r in rows]


@api_router.get("/ordem-servico/next-number")
async def get_next_os_number(current_user: dict = Depends(get_current_user)):
    last = await db.ordem_servico.find_one({}, sort=[("os_number", -1)])
    return {"next_number": (last["os_number"] + 1) if last else 1}


@api_router.get("/ordem-servico/{os_id}", response_model=OrdemServicoResponse)
async def get_ordem_servico(os_id: str, current_user: dict = Depends(get_current_user)):
    doc = await db.ordem_servico.find_one({"id": os_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Ordem de Serviço não encontrada")
    return _os_serialize(doc)


@api_router.post("/ordem-servico", response_model=OrdemServicoResponse)
async def create_ordem_servico(data: OrdemServicoCreate, current_user: dict = Depends(get_current_user)):
    last = await db.ordem_servico.find_one({}, sort=[("os_number", -1)])
    next_num = (last["os_number"] + 1) if last else 1

    os_obj = OrdemServico(
        os_number=next_num,
        **data.model_dump(),
        created_by=current_user["sub"],
        created_by_name=current_user["name"]
    )
    doc = os_obj.model_dump()
    doc["created_at"] = os_obj.created_at.isoformat()
    await db.ordem_servico.insert_one(doc)
    return _os_serialize(doc)


@api_router.put("/ordem-servico/{os_id}", response_model=OrdemServicoResponse)
async def update_ordem_servico(os_id: str, data: OrdemServicoUpdate, current_user: dict = Depends(get_current_user)):
    existing = await db.ordem_servico.find_one({"id": os_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Ordem de Serviço não encontrada")
    update_data = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.ordem_servico.update_one({"id": os_id}, {"$set": update_data})
    updated = await db.ordem_servico.find_one({"id": os_id}, {"_id": 0})
    return _os_serialize(updated)


@api_router.delete("/ordem-servico/{os_id}")
async def delete_ordem_servico(os_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.ordem_servico.delete_one({"id": os_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ordem de Serviço não encontrada")
    return {"message": "Ordem de Serviço removida"}


@api_router.get("/ordem-servico/{os_id}/pdf")
async def download_ordem_servico_pdf(os_id: str, current_user: dict = Depends(get_current_user)):
    """Gera PDF da Ordem de Serviço seguindo o modelo Bsoft TMS."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    os_doc = await db.ordem_servico.find_one({"id": os_id}, {"_id": 0})
    if not os_doc:
        raise HTTPException(status_code=404, detail="Ordem de Serviço não encontrada")
    os_doc = _os_serialize(os_doc)

    def money(v):
        try:
            return f"{float(v or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except Exception:
            return "0,00"

    def fmt_dt(s):
        if not s:
            return ''
        try:
            return datetime.fromisoformat(s.replace('Z', '+00:00')).strftime('%d/%m/%Y %H:%M:%S')
        except Exception:
            return str(s)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=8 * mm, bottomMargin=8 * mm)
    elements = []
    styles = getSampleStyleSheet()

    BLACK = colors.HexColor('#000000')
    GRAY_BG = colors.HexColor('#E8E8E8')

    # ===== HEADER: Logo + Empresa (esquerda) + Título OS (direita) =====
    from reports import download_logo
    from reportlab.platypus import Image as RLImage
    logo_buffer = download_logo()
    if logo_buffer:
        logo_img = RLImage(logo_buffer, width=22 * mm, height=22 * mm)
    else:
        logo_img = Paragraph("", styles['Normal'])

    company_style = ParagraphStyle('CompHead', parent=styles['Normal'], fontSize=9, leading=11,
                                   fontName='Helvetica-Bold')
    company_para = Paragraph(
        "<b>J. A. LOGISTICA LTDA - ME</b><br/>"
        "<font size='8'>RODOVIA CE-155, 16226 - INDUSTRIAL<br/>"
        "CEP: 61668-150, CAUCAIA - CE, IE: 07224458-5 - Fone:<br/>"
        "CNPJ: 58.180.321/0001-03, e-mail:</font>", company_style)

    os_title_style = ParagraphStyle('OSTit', parent=styles['Normal'], fontSize=14, leading=16,
                                    alignment=TA_RIGHT, fontName='Helvetica-Bold')
    right_para = Paragraph(
        f"Ordem de Serviço<br/><font size='9'>O.S. Nro: <b>{os_doc['os_number']}</b></font>", os_title_style)

    header = Table([[logo_img, company_para, right_para]], colWidths=[24 * mm, 96 * mm, 70 * mm])
    header.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('LEFTPADDING', (1, 0), (1, 0), 4),
    ]))
    elements.append(header)
    elements.append(Spacer(1, 4))

    # ===== Linha de categoria + datas =====
    label_s = ParagraphStyle('Lbl', parent=styles['Normal'], fontSize=7, leading=9,
                             textColor=colors.HexColor('#555'))
    value_s = ParagraphStyle('Val', parent=styles['Normal'], fontSize=8, leading=10,
                             fontName='Helvetica-Bold')

    def field(label, val):
        return Paragraph(f"<font size='7' color='#555'>{label}</font><br/>"
                         f"<font size='8'><b>{val if val else '_____________'}</b></font>",
                         ParagraphStyle('F', parent=styles['Normal'], leading=11))

    info_row1 = [
        field("Categoria:", os_doc.get('category')),
        field("Data/Hora Recepção:", fmt_dt(os_doc.get('opened_at'))),
        field("Data de abertura:", fmt_dt(os_doc.get('opened_at'))),
    ]
    info_row2 = [
        field("Tipo:", os_doc.get('os_type')),
        field("Data de fechamento:", fmt_dt(os_doc.get('closed_at'))),
        field("Tempo de serviço:", ''),
    ]
    info_t = Table([info_row1, info_row2], colWidths=[80 * mm, 55 * mm, 55 * mm])
    info_t.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, BLACK),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(info_t)
    elements.append(Spacer(1, 4))

    # ===== Dados da O.S. (Clientes) =====
    def section_bar(title):
        t = Table([[Paragraph(f"<b>{title}</b>",
                              ParagraphStyle('SB', parent=styles['Normal'], fontSize=8.5))]],
                  colWidths=[190 * mm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), GRAY_BG),
            ('LINEBELOW', (0, 0), (-1, -1), 0.5, BLACK),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ]))
        return t

    elements.append(section_bar("Dados da O.S."))
    dados_t = Table([[
        field("Clientes:", os_doc.get('person_name')),
        field("Supervisor:", os_doc.get('supervisor_name')),
    ], [
        field("CPF/CNPJ:", os_doc.get('person_doc')),
        field("Técnico:", os_doc.get('technician_name')),
    ], [
        field("PT:", "Sim" if os_doc.get('requires_pt') else "Não"),
        field("Status:", os_doc.get('status')),
    ]], colWidths=[95 * mm, 95 * mm])
    dados_t.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, BLACK),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(dados_t)
    elements.append(Spacer(1, 4))

    # ===== Detalhes Operacionais =====
    elements.append(section_bar("Detalhes Operacionais"))
    det_t = Table([[
        field("Endereço:", os_doc.get('address')),
        field("Telefone:", os_doc.get('contact_value')),
    ], [
        field("Cidade/Estado:", os_doc.get('city_uf')),
        field("Data agenda:", ''),
    ], [
        field("Hora agenda:", ''),
        field("Prioridade:", os_doc.get('priority')),
    ]], colWidths=[95 * mm, 95 * mm])
    det_t.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, BLACK),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(det_t)
    elements.append(Spacer(1, 4))

    # ===== Identificação do Cliente (Equipamento) =====
    elements.append(section_bar("Identificação do Cliente"))
    eq_t = Table([[
        field("Equipamento (Placa):", os_doc.get('equipment_plate')),
        field("Medidor de abertura:", f"{os_doc.get('reading_initial') or 0:.0f}"),
    ], [
        field("Descrição:", os_doc.get('description')),
        Paragraph("", styles['Normal']),
    ]], colWidths=[95 * mm, 95 * mm])
    eq_t.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, BLACK),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(eq_t)
    elements.append(Spacer(1, 4))

    # ===== Detalhamento da Demanda =====
    elements.append(section_bar("Detalhamento da Demanda"))
    demand_t = Table([[Paragraph(
        f"<font size='9'>{(os_doc.get('description') or '').replace(chr(10), '<br/>')}</font>",
        styles['Normal']
    )]], colWidths=[190 * mm], rowHeights=[36])
    demand_t.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, BLACK),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(demand_t)
    elements.append(Spacer(1, 4))

    # ===== Parecer de Encerramento =====
    elements.append(section_bar("Parecer de Encerramento"))
    enc_t = Table([[Paragraph(
        f"<font size='8'>Uso no fechamento: ____________________________________________________________<br/>"
        f"<br/>{(os_doc.get('closure_remark') or '').replace(chr(10), '<br/>')}</font>",
        styles['Normal']
    )]], colWidths=[190 * mm], rowHeights=[28])
    enc_t.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, BLACK),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(enc_t)
    elements.append(Spacer(1, 4))

    # ===== Produtos =====
    elements.append(section_bar("Produtos"))
    prod_header = ['Código', 'Descrição', 'Qtd', 'Un', 'V. Unit.', 'V. Total', 'Desc.', 'V. c/ Desc.']
    prod_rows = [prod_header]
    for p in (os_doc.get('products') or []):
        prod_rows.append([
            p.get('code') or '-',
            p.get('description') or '-',
            f"{float(p.get('quantity') or 0):.2f}".replace('.', ','),
            p.get('unit') or 'UN',
            money(p.get('unit_price')),
            money(float(p.get('quantity') or 0) * float(p.get('unit_price') or 0)),
            money(p.get('discount')),
            money(p.get('total')),
        ])
    prod_rows.append(['', 'Total', '', '',
                      money(sum(float(p.get('unit_price') or 0) for p in (os_doc.get('products') or []))),
                      '', '', money(os_doc.get('products_total'))])
    prod_t = Table(prod_rows, colWidths=[16, 70, 14, 14, 18, 18, 14, 26], repeatRows=1)
    prod_t._argW = [16 * mm, 70 * mm, 14 * mm, 14 * mm, 18 * mm, 18 * mm, 14 * mm, 26 * mm]
    prod_t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), GRAY_BG),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F5F5F5')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(prod_t)
    elements.append(Spacer(1, 4))

    # ===== Serviços =====
    elements.append(section_bar("Serviços"))
    serv_header = ['Código', 'Descrição', 'Qtd', 'Unidade', 'V. Unit.', 'V. Total']
    serv_rows = [serv_header]
    for s in (os_doc.get('services') or []):
        serv_rows.append([
            s.get('code') or '-',
            s.get('description') or '-',
            f"{float(s.get('quantity') or 0):.2f}".replace('.', ','),
            s.get('unit') or 'quantidade',
            money(s.get('unit_price')),
            money(s.get('total')),
        ])
    serv_rows.append(['', 'Total', '', '', '', money(os_doc.get('services_total'))])
    serv_t = Table(serv_rows, colWidths=[20 * mm, 95 * mm, 18 * mm, 20 * mm, 18 * mm, 19 * mm], repeatRows=1)
    serv_t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), GRAY_BG),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F5F5F5')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(serv_t)
    elements.append(Spacer(1, 8))

    # ===== Chegada/Saída + Assinaturas =====
    sig_t = Table([
        [Paragraph("<font size='8'>Data / Hora da Chegada: _________________________</font>", styles['Normal']),
         Paragraph("<font size='8'>Data / Hora da Saída: _________________________</font>", styles['Normal'])],
        [Paragraph("&nbsp;", styles['Normal']), Paragraph("&nbsp;", styles['Normal'])],
        [Paragraph("<font size='8'><b>__________________________<br/>Técnico:</b></font>", styles['Normal']),
         Paragraph("<font size='8'><b>__________________________<br/>Cliente:</b></font>", styles['Normal'])],
    ], colWidths=[95 * mm, 95 * mm], rowHeights=[None, 14, None])
    sig_t.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(sig_t)
    elements.append(Spacer(1, 6))

    # ===== Declarações =====
    decl_style = ParagraphStyle('Decl', parent=styles['Normal'], fontSize=7.5, leading=10,
                                fontName='Helvetica-Oblique')
    decl_t = Table([
        [Paragraph("O serviço foi realizado e o cliente declara ter realizado os devidos testes de funcionamento do equipamento.", decl_style),
         Paragraph("O cliente não forneceu acesso ao equipamento para realização do serviço responsabilizando-se pelas implicações que esta ação pode gerar.", decl_style)]
    ], colWidths=[95 * mm, 95 * mm])
    decl_t.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.3, colors.grey),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(decl_t)
    elements.append(Spacer(1, 8))

    # ===== Rodapé =====
    elements.append(Paragraph(
        f"<font size='7' color='#888'>{now_brt().strftime('%d/%m/%Y %H:%M')} &nbsp;&nbsp; "
        f"J.A Logística - Sistema de Gestão</font>",
        ParagraphStyle('Footer', parent=styles['Normal'], alignment=TA_CENTER)
    ))

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    filename = f"OS_{os_doc['os_number']}.pdf"
    return StreamingResponse(io.BytesIO(pdf_bytes), media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


app.include_router(api_router)

# WebSocket endpoint para sincronização em tempo real
@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            # Mantém a conexão aberta e escuta mensagens (ping/pong)
            data = await websocket.receive_text()
            if data == "ping":
                await websocket.send_text("pong")
    except WebSocketDisconnect:
        manager.disconnect(websocket)
    except Exception:
        manager.disconnect(websocket)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


@app.on_event("startup")
async def startup_event():
    # Inicializar contador de transaction_id se não existir
    counter = await db.counters.find_one({"_id": "transaction_id"})
    if not counter:
        # Buscar o maior transaction_id existente
        last_movement = await db.movements.find_one({}, {"_id": 0, "transaction_id": 1}, sort=[("transaction_id", -1)])
        max_id = last_movement.get('transaction_id', 0) if last_movement else 0
        await db.counters.insert_one({"_id": "transaction_id", "seq": max_id})

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()