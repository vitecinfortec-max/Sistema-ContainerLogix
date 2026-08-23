from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from datetime import datetime, timezone, timedelta
from pathlib import Path
import io
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
import requests
from PIL import Image as PILImage
import logging
from models import VEHICLE_CHECKLIST_TEMPLATE, VEHICLE_CHECKLIST_SECTION_LABELS

logger = logging.getLogger(__name__)

# Fuso horário de Brasília (UTC-3). Independe do TZ do container.
BRT_TZ = timezone(timedelta(hours=-3))


def now_brt() -> datetime:
    """Retorna o horário atual em Brasília (UTC-3) com tzinfo."""
    return datetime.now(BRT_TZ)


def to_brt(dt_value) -> datetime | None:
    """Converte um datetime/ISO string para o fuso de Brasília.
    - Datetimes 'naive' (sem tz) são interpretados como UTC (padrão do banco).
    - Retorna None se não for possível converter.
    """
    if dt_value is None or dt_value == '':
        return None
    try:
        if isinstance(dt_value, str):
            # Suporta 'Z' e offset
            iso = dt_value.replace('Z', '+00:00')
            dt = datetime.fromisoformat(iso)
        elif isinstance(dt_value, datetime):
            dt = dt_value
        else:
            return None
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(BRT_TZ)
    except Exception:
        return None

def fmt_date(value):
    """Formata uma data 'YYYY-MM-DD' como 'DD/MM/YYYY'; devolve o valor original
    (ou '-') se não for possível interpretar."""
    if not value:
        return '-'
    try:
        return datetime.strptime(value, '%Y-%m-%d').strftime('%d/%m/%Y')
    except Exception:
        return value


def fmt_datetime(value):
    """Formata um datetime ISO como 'DD/MM/YYYY HH:MM'; devolve o valor original
    (ou '-') se não for possível interpretar."""
    if not value:
        return '-'
    try:
        return datetime.fromisoformat(value.replace('Z', '+00:00')).strftime('%d/%m/%Y %H:%M')
    except Exception:
        return value


# Sem valor padrão fixo: cada instância/cliente define seu próprio logo em "Dados
# da Empresa", ou via variável de ambiente se quiser um logo padrão pré-configurado.
LOGO_URL = os.environ.get('LOGO_URL')

PRIMARY_COLOR = "008B7B"
HEADER_BG_COLOR = "E8F4F5"

ROOT_DIR = Path(__file__).parent
UPLOADS_DIR = ROOT_DIR.parent / 'uploads'

# Valores usados quando a empresa ainda não configurou seus próprios dados em
# "Dados da Empresa" - propositalmente genéricos (nunca dados reais de um cliente
# específico), já que cada instância vendida usa esse mesmo fallback até configurar
# o seu. CNPJ/telefone/dados bancários/PIX ficam como placeholder óbvio em vez de
# vazios, pra não sair "Banco: " em branco nem, pior, os dados de outro cliente.
DEFAULT_COMPANY = {
    'name': 'Sua Empresa',
    'cnpj': '00.000.000/0000-00',
    'address': 'Configure em "Dados da Empresa"',
    'phone': '(00) 00000-0000',
    'email': 'contato@suaempresa.com',
    'bank_name': 'Configure em "Dados da Empresa"',
    'bank_agency': '-',
    'bank_account': '-',
    'pix_key': 'Configure em "Dados da Empresa"',
}


def format_currency(value, currency='BRL'):
    """Formata valor monetário com o símbolo da moeda. BRL usa separador
    decimal brasileiro (R$ 1.234,56); demais moedas usam o padrão
    internacional (US$ 1,234.56), que é o correto para elas."""
    value = value or 0
    symbol = {'USD': '$', 'EUR': '€', 'BRL': 'R$'}.get(currency, currency)
    if currency == 'BRL':
        formatted = f"{value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    else:
        formatted = f"{value:,.2f}"
    return f"{symbol} {formatted}"


def merge_company(company: dict = None) -> dict:
    """Combina os dados cadastrados em 'Dados da Empresa' com os valores padrão,
    usando o padrão para qualquer campo ainda não preenchido."""
    company = company or {}
    return {**DEFAULT_COMPANY, **{k: v for k, v in company.items() if v}}


def download_logo(company: dict = None):
    """Retorna o logo como file-like object.
    Prioriza o logo enviado pelo usuário em 'Dados da Empresa' (uploads/); se não
    houver e a variável de ambiente LOGO_URL estiver configurada (opcional, usada
    quando essa instância quer um logo padrão pré-definido), cai pra ela. Sem
    nenhum dos dois, o cabeçalho do documento simplesmente sai sem logo."""
    if company and company.get('logo_filename'):
        try:
            logo_path = UPLOADS_DIR / company['logo_filename']
            if logo_path.exists():
                return io.BytesIO(logo_path.read_bytes())
        except Exception as e:
            logger.error(f"Error reading uploaded logo: {e}")
    if not LOGO_URL:
        return None
    try:
        response = requests.get(LOGO_URL, timeout=5)
        if response.status_code == 200:
            return io.BytesIO(response.content)
    except Exception as e:
        logger.error(f"Error downloading logo: {e}")
    return None


def _build_pdf_header(styles, logo_buffer, report_title, generation_info=None, company=None, content_width=540):
    """
    Build standard PDF header with logo on left and company info centered.
    `content_width` deve ser a largura útil real do documento (doc.width) para
    que o cabeçalho nunca ultrapasse a margem nem fique desalinhado do resto do conteúdo.
    """
    c = merge_company(company)
    elements = []

    # ========== LOGO ==========
    # Redimensiona mantendo a proporção original (evita esticar/achatar a logo),
    # limitando à mesma caixa 70x70 usada antes.
    logo_cell = ""
    if logo_buffer:
        try:
            max_box = 70
            logo_buffer.seek(0)
            with PILImage.open(logo_buffer) as pil_img:
                orig_w, orig_h = pil_img.size
            if orig_w and orig_h:
                scale = min(max_box / orig_w, max_box / orig_h)
                logo_w, logo_h = orig_w * scale, orig_h * scale
            else:
                logo_w, logo_h = max_box, max_box
            logo_buffer.seek(0)
            logo_cell = Image(logo_buffer, width=logo_w, height=logo_h)
        except Exception as e:
            logger.error(f"Error adding logo to PDF: {e}")
    
    # ========== COMPANY INFO ==========
    company_style = ParagraphStyle(
        'CompanyName',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.black,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
        leading=16
    )
    address_style = ParagraphStyle(
        'Address',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.black,
        alignment=TA_CENTER,
        leading=12
    )
    
    address_lines = [line.strip() for line in c['address'].split('\n') if line.strip()]
    company_info_elements = [
        Paragraph(c['name'], company_style),
        Paragraph(f"CNPJ: {c['cnpj']}", address_style),
    ] + [
        Paragraph(line, address_style) for line in address_lines
    ] + [
        Paragraph(f"{c['email']} | {c['phone']}", address_style),
    ]
    
    # Inner header table: Logo | Company Info
    header_data = [[logo_cell, company_info_elements]]
    header_table = Table(header_data, colWidths=[80, 350])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'CENTER'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
    ]))
    
    # Outer table to center the entire header on the page: largura igual à área
    # útil real do documento, para nunca "vazar" da margem nem ficar deslocado.
    outer_table = Table([[header_table]], colWidths=[content_width])
    outer_table.hAlign = 'CENTER'
    outer_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(outer_table)
    elements.append(Spacer(1, 8))
    
    # Horizontal line separator
    elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor(f'#{PRIMARY_COLOR}'), spaceAfter=10))
    
    # ========== REPORT TITLE ==========
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor(f'#{PRIMARY_COLOR}'),
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
        spaceAfter=10
    )
    elements.append(Paragraph(report_title.upper(), title_style))
    
    return elements


def _make_pdf_footer(company_name):
    """Retorna uma função de rodapé (canvas, doc) -> None presa ao nome da empresa
    cadastrada em 'Dados da Empresa' — necessário porque o ReportLab só chama
    onFirstPage/onLaterPages com (canvas, doc), sem espaço para outros argumentos."""
    def _footer(canvas, doc):
        canvas.saveState()

        # Footer line
        canvas.setStrokeColor(colors.HexColor(f'#{PRIMARY_COLOR}'))
        canvas.setLineWidth(0.5)
        canvas.line(doc.leftMargin, 25, doc.width + doc.leftMargin, 25)

        # Page number (left)
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.grey)
        canvas.drawString(doc.leftMargin, 12, f"Página {doc.page}")

        # System info (right)
        canvas.drawRightString(doc.width + doc.leftMargin, 12, f"ContainerLogix - {company_name}")

        canvas.restoreState()
    return _footer


def generate_pdf_report(movements: list, report_title: str = "Relatório de Movimentações", company: dict = None) -> bytes:
    """
    Generate PDF report following the standard J.A LOGÍSTICA layout: logo + dados
    da empresa no topo (via _build_pdf_header), linha de estatísticas, tabela com
    cabeçalho teal e rodapé com numeração de página (via _make_pdf_footer) — mesmo
    padrão visual dos demais relatórios do sistema.
    """
    c = merge_company(company)
    buffer = io.BytesIO()

    # Use landscape orientation for A4
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=10*mm,
        leftMargin=10*mm,
        topMargin=15*mm,
        bottomMargin=15*mm
    )

    elements = []
    styles = getSampleStyleSheet()

    logo_buffer = download_logo(company)
    elements.extend(_build_pdf_header(styles, logo_buffer, report_title, company=company, content_width=doc.width))

    # ========== STATISTICS LINE ==========
    total_records = len(movements)
    total_entries = sum(1 for m in movements if m.get('operation_type') == 'ENTRADA')
    total_exits = sum(1 for m in movements if m.get('operation_type') == 'SAIDA')
    
    if "Estoque" in report_title:
        stats_text = f"Containers em Estoque: {total_records}"
    elif "Entradas" in report_title:
        stats_text = f"Total de Entradas: {total_records}"
    elif "Saídas" in report_title:
        stats_text = f"Total de Saídas: {total_records}"
    else:
        stats_text = f"Total: {total_records}  |  Entradas: {total_entries}  |  Saídas: {total_exits}"
    
    stats_style = ParagraphStyle(
        'StatsLine',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor(f'#{PRIMARY_COLOR}'),
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
        spaceBefore=6,
        spaceAfter=8
    )
    elements.append(Paragraph(stats_text, stats_style))
    
    # ========== GENERATION INFO ==========
    gen_style = ParagraphStyle(
        'GenInfo',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#808080'),
        alignment=TA_CENTER,
        spaceBefore=0,
        spaceAfter=12
    )
    gen_date = now_brt().strftime('%d/%m/%Y %H:%M')
    elements.append(Paragraph(f"Gerado em: {gen_date} | Fuso: UTC-3 (Brasília)", gen_style))
    
    # ========== TABLE SECTION ==========
    # 15 columns matching template
    data = [[
        'ID Trans.', 'Data/Hora', 'Tipo', 'Nº Container', 'Motorista', 'CPF',
        'Placa Cavalo', 'Placa Carreta', 'Transportadora',
        'Status', 'Tamanho', 'Tara', 'Armador', 'Booking'
    ]]

    for m in movements:
        dt_brt = to_brt(m.get('created_at'))
        created_at = dt_brt.strftime('%d/%m/%Y %H:%M') if dt_brt else str(m.get('created_at', '-'))

        data.append([
            str(m.get('transaction_id', '-')),
            created_at,
            "ENTRADA" if m.get('operation_type') == 'ENTRADA' else "SAÍDA",
            m.get('container_number', '-'),
            m.get('driver_name', '-'),
            m.get('driver_cpf', '-'),
            m.get('truck_plate', '-'),
            m.get('trailer_plate_1', '') or '-',
            m.get('transport_company', '-'),
            m.get('status', 'VAZIO'),
            m.get('size_type', '-'),
            str(m.get('tare', '')) if m.get('tare') else '-',
            m.get('shipping_line', '-'),
            m.get('booking', '') or '-'
        ])

    # Column widths for 14 columns in landscape A4 (~780 points available)
    col_widths = [30, 56, 42, 58, 75, 48, 46, 52, 72, 38, 38, 32, 52, 42]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    header_bg = colors.HexColor(f'#{PRIMARY_COLOR}')
    border_gray = colors.HexColor('#CCCCCC')
    zebra_gray = colors.HexColor('#F8F8F8')

    table.setStyle(TableStyle([
        # Header styling - mesmo teal + texto branco usado nos demais relatórios
        ('BACKGROUND', (0, 0), (-1, 0), header_bg),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
        ('TOPPADDING', (0, 0), (-1, 0), 5),
        
        # Body styling
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        
        # Alignments
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),   # ID Trans.
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),   # Tipo
        ('ALIGN', (9, 1), (9, -1), 'CENTER'),   # Status
        ('ALIGN', (10, 1), (10, -1), 'CENTER'), # Tamanho
        ('ALIGN', (11, 1), (11, -1), 'CENTER'), # Tara
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),     # Data/Hora
        ('ALIGN', (3, 1), (3, -1), 'LEFT'),     # Nº Container
        ('ALIGN', (4, 1), (4, -1), 'LEFT'),     # Motorista
        ('ALIGN', (5, 1), (5, -1), 'LEFT'),     # CPF
        ('ALIGN', (6, 1), (6, -1), 'LEFT'),     # Placa Cavalo
        ('ALIGN', (7, 1), (7, -1), 'LEFT'),     # Placa Carreta
        ('ALIGN', (8, 1), (8, -1), 'LEFT'),     # Transportadora
        ('ALIGN', (12, 1), (12, -1), 'LEFT'),   # Armador
        ('ALIGN', (13, 1), (13, -1), 'LEFT'),   # Booking
        
        # Bordas finas cinza, mesmo peso usado nos demais relatórios
        ('GRID', (0, 0), (-1, -1), 0.5, border_gray),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor(f'#{PRIMARY_COLOR}')),

        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        # Zebra striping
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, zebra_gray]),
    ]))

    elements.append(table)

    # ========== SAÍDAS POR BOOKING (resumo) ==========
    booking_counts = {}
    for m in movements:
        if m.get('operation_type') == 'SAIDA':
            booking = (m.get('booking') or '').strip()
            key = booking if booking else 'Sem Booking'
            booking_counts[key] = booking_counts.get(key, 0) + 1

    if booking_counts:
        elements.append(Spacer(1, 14))

        summary_title_style = ParagraphStyle(
            'BookingSummaryTitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor(f'#{PRIMARY_COLOR}'),
            alignment=TA_LEFT,
            fontName='Helvetica-Bold',
            spaceAfter=4
        )
        total_exits_summary = sum(booking_counts.values())
        elements.append(Paragraph(
            f"Saídas por Booking ({total_exits_summary} container{'s' if total_exits_summary != 1 else ''})",
            summary_title_style
        ))

        # Ordenar por quantidade decrescente, depois por nome do booking
        sorted_bookings = sorted(booking_counts.items(), key=lambda x: (-x[1], x[0]))
        summary_data = [['Booking', 'Qtd. Containers (Saída)']]
        for booking_name, qty in sorted_bookings:
            summary_data.append([booking_name, str(qty)])
        summary_data.append(['TOTAL', str(total_exits_summary)])

        summary_table = Table(summary_data, colWidths=[180, 110])
        summary_table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), header_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
            ('TOPPADDING', (0, 0), (-1, 0), 5),

            # Body
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),

            # Total row (last)
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor(f'#{HEADER_BG_COLOR}')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),

            ('GRID', (0, 0), (-1, -1), 0.25, border_gray),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(summary_table)

    footer = _make_pdf_footer(c['name'])
    doc.build(elements, onFirstPage=footer, onLaterPages=footer)
    pdf_bytes = buffer.getvalue()
    buffer.close()

    return pdf_bytes


def _bsoft_style_excel(ws, title, info_text, headers, data_rows, col_widths, center_cols=None, right_align_cols=None, number_fmt_cols=None, total_col=None, stats_text=None, company_name=None, logo_buffer=None, total_number_format='R$ #,##0.00'):
    """
    Shared Bsoft-style Excel formatting matching the J.A LOGÍSTICA template.
    - ws: worksheet
    - title: report subtitle text (e.g., "Relatório de Estoque Atual - Cliente: X")
    - info_text: stats line (e.g., "Containers em Estoque: 182")
    - headers: list of header strings
    - data_rows: list of lists (each inner list = 1 row of data)
    - col_widths: dict {col_letter: width}
    - center_cols: set of 0-based col indices for center alignment
    - right_align_cols: set of 0-based col indices for right alignment
    - number_fmt_cols: dict {0-based col index: format_string}
    - total_col: 0-based col index for SUM total row (or None)
    - stats_text: optional separate stats text for row 6
    """
    from openpyxl.utils import get_column_letter

    if center_cols is None:
        center_cols = set()
    if right_align_cols is None:
        right_align_cols = set()
    if number_fmt_cols is None:
        number_fmt_cols = {}

    num_cols = len(headers)
    first_col = 2  # column B (1-indexed)
    last_col = first_col + num_cols - 1
    first_letter = get_column_letter(first_col)
    last_letter = get_column_letter(last_col)

    # Colors from template
    TEAL_COLOR = '008B7B'
    STATS_BG = 'E8F4F5'
    ZEBRA_GRAY = 'F8F8F8'

    # Column A spacer
    ws.column_dimensions['A'].width = 3

    # Logo da empresa (cadastrado em "Dados da Empresa"), alinhada ao bloco
    # do nome da empresa (linhas 2-3), mantendo a proporção original da imagem
    if logo_buffer is not None:
        try:
            target_height_px = 92  # aprox. altura das linhas 2-3 (30pt + 40.5pt)
            try:
                logo_buffer.seek(0)
                with PILImage.open(logo_buffer) as pil_img:
                    orig_w, orig_h = pil_img.size
                target_width_px = int(target_height_px * orig_w / orig_h) if orig_h else target_height_px
            except Exception:
                target_width_px = target_height_px

            logo_buffer.seek(0)
            logo_img = XLImage(logo_buffer)
            logo_img.height = target_height_px
            logo_img.width = target_width_px
            ws.add_image(logo_img, 'A2')
        except Exception as e:
            logger.error(f"Error adding logo to Excel: {e}")

    # Apply col widths (keys are B, C, D...)
    for letter, w in col_widths.items():
        ws.column_dimensions[letter].width = w

    # Border style - all thin
    thin_side = Side(style='thin', color='000000')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # ======== HEADER ========
    # Row 2-3: Company name "J.A LOGÍSTICA" (merged B2:O3)
    ws.merge_cells(f'{first_letter}2:{last_letter}3')
    c = ws[f'{first_letter}2']
    c.value = company_name or DEFAULT_COMPANY['name']
    c.font = Font(name='Calibri', size=38, bold=True, color=TEAL_COLOR)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 40.5
    # Apply border to all cells in merged range
    for r in [2, 3]:
        for ci in range(first_col, last_col + 1):
            ws.cell(row=r, column=ci).border = thin_border

    # Row 4: Report subtitle (e.g., "Relatório de Estoque Atual - Cliente: X")
    ws.merge_cells(f'{first_letter}4:{last_letter}4')
    c = ws[f'{first_letter}4']
    c.value = title
    c.font = Font(name='Calibri', size=16, color=TEAL_COLOR)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = Border(left=thin_side, right=thin_side, top=Side(), bottom=thin_side)
    ws.row_dimensions[4].height = 21
    for ci in range(first_col, last_col + 1):
        cell = ws.cell(row=4, column=ci)
        cell.border = Border(left=thin_side, right=thin_side, top=Side(), bottom=thin_side)

    # Row 5: Empty spacer row
    ws.row_dimensions[5].height = 13

    # Row 6: Stats bar (e.g., "Containers em Estoque: 182")
    ws.merge_cells(f'{first_letter}6:{last_letter}6')
    c = ws[f'{first_letter}6']
    c.value = stats_text if stats_text else info_text
    c.font = Font(name='Calibri', size=12, bold=True, color=TEAL_COLOR)
    c.fill = PatternFill(start_color=STATS_BG, end_color=STATS_BG, fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border
    ws.row_dimensions[6].height = 28
    for ci in range(first_col, last_col + 1):
        cell = ws.cell(row=6, column=ci)
        cell.fill = PatternFill(start_color=STATS_BG, end_color=STATS_BG, fill_type='solid')
        cell.border = thin_border

    # Row 7: Generation date/time
    ws.merge_cells(f'{first_letter}7:{last_letter}7')
    c = ws[f'{first_letter}7']
    c.value = f"Gerado em: {now_brt().strftime('%d/%m/%Y %H:%M')} | Fuso: UTC-3 (Brasília)"
    c.font = Font(name='Calibri', size=9, color='808080')
    c.alignment = Alignment(horizontal='center')
    c.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=Side())
    ws.row_dimensions[7].height = 15
    for ci in range(first_col, last_col + 1):
        ws.cell(row=7, column=ci).border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=Side())

    # ======== COLUMN HEADERS (Row 8) ========
    header_row = 8
    header_fill = PatternFill(start_color=TEAL_COLOR, end_color=TEAL_COLOR, fill_type='solid')
    header_font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for i, h in enumerate(headers):
        ci = first_col + i
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 18

    # ======== DATA ROWS ========
    data_start = header_row + 1
    total_data = len(data_rows)
    left_align = Alignment(horizontal='left', vertical='center')
    center_align_data = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    data_font = Font(name='Calibri', size=9)

    for row_offset, row_data in enumerate(data_rows):
        row_num = data_start + row_offset
        # Zebra striping: odd rows get gray background
        is_gray_row = (row_offset % 2 == 1)
        row_fill = PatternFill(start_color=ZEBRA_GRAY, end_color=ZEBRA_GRAY, fill_type='solid') if is_gray_row else PatternFill(fill_type=None)
        
        for i, val in enumerate(row_data):
            ci = first_col + i
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.font = data_font
            cell.border = thin_border
            if is_gray_row:
                cell.fill = row_fill
            
            # Apply alignment based on column type
            if i in center_cols:
                cell.alignment = center_align_data
            elif i in right_align_cols or i in number_fmt_cols:
                cell.alignment = right_align
            else:
                cell.alignment = left_align

            if i in number_fmt_cols:
                cell.number_format = number_fmt_cols[i]

    # ======== TOTAL ROW ========
    if total_col is not None and total_data > 0:
        total_row_num = data_start + total_data
        # Label
        label_ci = first_col + total_col - 1
        label_cell = ws.cell(row=total_row_num, column=label_ci, value='TOTAL:')
        label_cell.font = Font(name='Calibri', size=9, bold=True)
        label_cell.alignment = Alignment(horizontal='right', vertical='center')

        # Sum formula
        val_ci = first_col + total_col
        val_letter = get_column_letter(val_ci)
        sum_formula = f'=SUM({val_letter}{data_start}:{val_letter}{data_start + total_data - 1})'
        sum_cell = ws.cell(row=total_row_num, column=val_ci, value=sum_formula)
        sum_cell.font = Font(name='Calibri', size=9, bold=True)
        sum_cell.number_format = total_number_format
        sum_cell.border = thin_border
        sum_cell.fill = PatternFill(start_color=STATS_BG, end_color=STATS_BG, fill_type='solid')
        last_row = total_row_num
    else:
        last_row = data_start + total_data - 1

    # ======== IMPRESSÃO ========
    # Sem isso, o Excel abre/imprime a planilha larga (13-15 colunas) no padrão
    # retrato dele, cortando colunas em várias páginas — configura para caber
    # numa página de largura, em paisagem, com o cabeçalho fixo ao rolar/imprimir.
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f'{first_letter}{header_row + 1}'
    ws.print_area = f'A1:{last_letter}{max(last_row, header_row)}'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True


def generate_delivery_status_excel(status: dict, company: dict = None) -> bytes:
    """Gera Excel de um Status de Entrega (um motorista/veículo por linha)."""
    try:
        c = merge_company(company)
        wb = Workbook()
        ws = wb.active
        ws.title = "Status de Entrega"

        status_date_value = status.get('status_date', '')
        if status_date_value:
            try:
                status_date_value = datetime.fromisoformat(status_date_value.replace('Z', '+00:00')).strftime('%d/%m/%Y')
            except Exception:
                pass

        title = f"Status de Entrega Nº {status.get('status_number', '-')} - Programação #{status.get('schedule_number', '-')}"
        stats_text = (
            f"Cliente Destino: {status.get('destination_client_name', '-')}"
            f"  |  Data: {status_date_value or '-'}"
        )

        items = status.get('items', []) or []
        has_bag_numbers = any((item.get('bag_number') or '').strip() for item in items)

        headers = ['#', 'Motorista', 'CPF', 'Cavalo', 'Carreta', 'Container', 'Local', 'Agend. Porto', 'Chegada', 'Início Carreg.', 'Término Carreg.', 'Saída', 'Entrega Finalizada']
        col_widths = {
            'B': 4.5, 'C': 26, 'D': 15, 'E': 12, 'F': 12, 'G': 16, 'H': 22,
            'I': 12, 'J': 12, 'K': 13, 'L': 14, 'M': 10, 'N': 16
        }
        center_cols = {0, 7, 8, 9, 10, 11, 12}

        if has_bag_numbers:
            headers.append('Nº da Bolsa')
            longest_bag_number = max((len(item.get('bag_number') or '') for item in items), default=0)
            col_widths['O'] = max(24, longest_bag_number + 8)
            center_cols.add(13)

        data_rows = []
        for idx, item in enumerate(items, 1):
            row = [
                idx,
                item.get('driver_name', '-') or '-',
                item.get('driver_cpf', '-') or '-',
                item.get('cavalo_plate', '-') or '-',
                item.get('carreta_plate', '-') or '-',
                item.get('container_number', '-') or '-',
                item.get('loading_location', '-') or '-',
                item.get('port_schedule_time', '-') or '-',
                item.get('arrival_time', '-') or '-',
                item.get('loading_start_time', '-') or '-',
                item.get('loading_end_time', '-') or '-',
                item.get('departure_time', '-') or '-',
                item.get('delivery_completed', '-') or '-',
            ]
            if has_bag_numbers:
                row.append(item.get('bag_number') or '-')
            data_rows.append(row)

        _bsoft_style_excel(
            ws, title, stats_text, headers, data_rows, col_widths,
            center_cols=center_cols,
            stats_text=stats_text,
            company_name=c['name'],
            logo_buffer=download_logo(company)
        )

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    except Exception as e:
        logger.error(f"Error generating delivery status Excel: {e}")
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Erro ao gerar relatório."
        ws['A1'].font = Font(size=14, bold=True, color="FF0000")
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


def generate_excel_report(movements: list, report_title: str = "Relatório de Movimentações", company: dict = None) -> bytes:
    """Generate Excel report following Bsoft template style."""
    try:
        c = merge_company(company)
        wb = Workbook()
        ws = wb.active
        ws.title = "Movimentações"

        total_records = len(movements)
        total_entries = sum(1 for m in movements if m.get('operation_type') == 'ENTRADA')
        total_exits = sum(1 for m in movements if m.get('operation_type') == 'SAIDA')

        if "Estoque" in report_title:
            stats_text = f"Containers em Estoque: {total_records}"
        elif "Entradas" in report_title:
            stats_text = f"Total de Entradas: {total_records}"
        elif "Saídas" in report_title:
            stats_text = f"Total de Saídas: {total_records}"
        else:
            stats_text = f"Total: {total_records}  |  Entradas: {total_entries}  |  Saídas: {total_exits}"

        # Headers matching template (14 columns)
        headers = [
            'ID Trans.', 'Data/Hora', 'Tipo', 'Nº Container', 'Motorista', 'CPF',
            'Placa Cavalo', 'Placa Carreta', 'Transportadora',
            'Status', 'Tamanho', 'Tara', 'Armador', 'Booking'
        ]

        data_rows = []
        for m in movements:
            dt_brt = to_brt(m.get('created_at'))
            created_at = dt_brt.strftime('%d/%m/%Y %H:%M') if dt_brt else str(m.get('created_at', ''))
            data_rows.append([
                m.get('transaction_id', '-'),
                created_at,
                "ENTRADA" if m.get('operation_type') == 'ENTRADA' else "SAÍDA",
                m.get('container_number', '-'),
                m.get('driver_name', '-'),
                m.get('driver_cpf', '-'),
                m.get('truck_plate', '-'),
                m.get('trailer_plate_1', '-') or '-',
                m.get('transport_company', '-'),
                m.get('status', 'VAZIO'),
                m.get('size_type', '-'),
                m.get('tare', '') or '-',
                m.get('shipping_line', '-'),
                m.get('booking', '') or '-'
            ])

        # Column widths matching template (14 columns: B to O)
        col_widths = {
            'B': 7.3, 'C': 13.7, 'D': 8, 'E': 14, 'F': 30, 'G': 14,
            'H': 12, 'I': 14, 'J': 27.3, 'K': 5.6, 'L': 7.6,
            'M': 4.5, 'N': 13.2, 'O': 12
        }

        # Center alignment for: ID Trans.(0), Tipo(2), Status(9), Tamanho(10), Tara(11)
        center_cols = {0, 2, 9, 10, 11}

        _bsoft_style_excel(
            ws, report_title, stats_text, headers, data_rows, col_widths,
            center_cols=center_cols,
            stats_text=stats_text,
            company_name=c['name'],
            logo_buffer=download_logo(company)
        )

        # ========== SAÍDAS POR BOOKING (resumo no final da planilha) ==========
        booking_counts = {}
        for m in movements:
            if m.get('operation_type') == 'SAIDA':
                booking_val = (m.get('booking') or '').strip()
                key = booking_val if booking_val else 'Sem Booking'
                booking_counts[key] = booking_counts.get(key, 0) + 1

        if booking_counts:
            from openpyxl.styles import Font as XLFont, Alignment as XLAlign, PatternFill as XLFill, Border as XLBorder, Side as XLSide

            thin_side = XLSide(style='thin', color='000000')
            thin_border = XLBorder(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            header_fill = XLFill(start_color='008B7B', end_color='008B7B', fill_type='solid')
            total_fill = XLFill(start_color='E8F4F5', end_color='E8F4F5', fill_type='solid')

            start_row = ws.max_row + 3  # deixa espaço da tabela principal

            # Título do bloco
            total_exits_summary = sum(booking_counts.values())
            ws.cell(row=start_row, column=2,
                    value=f"Saídas por Booking ({total_exits_summary} container{'s' if total_exits_summary != 1 else ''})")
            ws.cell(row=start_row, column=2).font = XLFont(name='Calibri', size=11, bold=True, color='008B7B')
            ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=3)

            # Cabeçalho da tabela
            header_row = start_row + 1
            ws.cell(row=header_row, column=2, value='Booking')
            ws.cell(row=header_row, column=3, value='Qtd. Containers (Saída)')
            for col_idx in (2, 3):
                cell = ws.cell(row=header_row, column=col_idx)
                cell.font = XLFont(name='Calibri', size=10, bold=True, color='FFFFFF')
                cell.alignment = XLAlign(horizontal='center', vertical='center')
                cell.fill = header_fill
                cell.border = thin_border

            # Linhas de dados (ordenado por quantidade desc, depois alfabético)
            sorted_bookings = sorted(booking_counts.items(), key=lambda x: (-x[1], x[0]))
            current_row = header_row + 1
            for booking_name, qty in sorted_bookings:
                ws.cell(row=current_row, column=2, value=booking_name)
                ws.cell(row=current_row, column=3, value=qty)
                ws.cell(row=current_row, column=2).alignment = XLAlign(horizontal='left', vertical='center')
                ws.cell(row=current_row, column=3).alignment = XLAlign(horizontal='center', vertical='center')
                ws.cell(row=current_row, column=2).border = thin_border
                ws.cell(row=current_row, column=3).border = thin_border
                ws.cell(row=current_row, column=2).font = XLFont(name='Calibri', size=10)
                ws.cell(row=current_row, column=3).font = XLFont(name='Calibri', size=10)
                current_row += 1

            # Linha de total
            ws.cell(row=current_row, column=2, value='TOTAL')
            ws.cell(row=current_row, column=3, value=total_exits_summary)
            for col_idx in (2, 3):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = XLFont(name='Calibri', size=10, bold=True)
                cell.alignment = XLAlign(
                    horizontal='left' if col_idx == 2 else 'center', vertical='center'
                )
                cell.fill = total_fill
                cell.border = thin_border

            # O resumo de bookings fica abaixo da tabela principal — estende a área
            # de impressão pra ele não ficar de fora ao imprimir/exportar em PDF.
            from openpyxl.utils import get_column_letter as _get_col_letter
            ws.print_area = f'A1:{_get_col_letter(2 + len(headers) - 1)}{current_row}'

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    except Exception as e:
        logger.error(f"Error generating Excel report: {e}")
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Erro ao gerar relatório."
        ws['A1'].font = Font(size=14, bold=True, color="FF0000")
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


def generate_billing_excel(movements: list, company: dict = None) -> bytes:
    """Generate billing Excel report with Bsoft template style."""
    try:
        c = merge_company(company)
        wb = Workbook()
        ws = wb.active
        ws.title = "Faturamento"

        total_value = sum(m.get('service_value', 0) or 0 for m in movements)
        val_str = f"R$ {total_value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        stats_text = f"Total: {len(movements)} movimentações  |  Valor Total: {val_str}"

        headers = [
            'ID', 'Data/Hora', 'Tipo', 'Nº Container', 'Cliente',
            'Placa', 'Transportadora', 'Armador', 'Status', 'Tamanho',
            'Tipo de Serviço', 'Nota Fiscal', 'Valor da Operação'
        ]

        data_rows = []
        for m in movements:
            dt_brt = to_brt(m.get('created_at'))
            created_at = dt_brt.strftime('%d/%m/%Y %H:%M') if dt_brt else str(m.get('created_at', ''))
            data_rows.append([
                m.get('transaction_id', '-'),
                created_at,
                "ENTRADA" if m.get('operation_type') == 'ENTRADA' else "SAÍDA",
                m.get('container_number') or '-',
                m.get('client_name') or '-',
                m.get('truck_plate') or '-',
                m.get('transport_company') or '-',
                m.get('shipping_line') or '-',
                m.get('status') or '-',
                m.get('size_type') or '-',
                m.get('service_type', '') or '-',
                m.get('invoice_number', '') or '-',
                m.get('service_value') if m.get('service_value') else 0,
            ])

        col_widths = {
            'B': 7.3, 'C': 13.7, 'D': 8, 'E': 14, 'F': 25,
            'G': 12, 'H': 20, 'I': 14, 'J': 6, 'K': 8,
            'L': 16, 'M': 12, 'N': 18
        }

        # Center alignment for: ID(0), Tipo(2), Status(8), Tamanho(9)
        center_cols = {0, 2, 8, 9}

        _bsoft_style_excel(
            ws, "Relatório de Faturamento", stats_text, headers, data_rows, col_widths,
            center_cols=center_cols,
            right_align_cols={12},
            number_fmt_cols={12: 'R$ #,##0.00'},
            total_col=12,
            stats_text=stats_text,
            company_name=c['name'],
            logo_buffer=download_logo(company)
        )

        # ========== DADOS BANCÁRIOS ==========
        # Encontrar a última linha com dados
        last_row = ws.max_row + 3  # Pular 3 linhas após a tabela

        # Estilo para o título dos dados bancários
        bank_title_font = Font(size=12, bold=True, color=PRIMARY_COLOR)
        bank_label_font = Font(size=10, bold=True)
        bank_value_font = Font(size=10)

        # Título "DADOS BANCÁRIOS PARA PAGAMENTO"
        ws.cell(row=last_row, column=2, value="DADOS BANCÁRIOS PARA PAGAMENTO")
        ws.cell(row=last_row, column=2).font = bank_title_font
        ws.merge_cells(start_row=last_row, start_column=2, end_row=last_row, end_column=6)

        # Linha separadora
        last_row += 1

        # Dados do banco
        bank_data = [
            ("Banco:", c['bank_name']),
            ("Agência:", c['bank_agency']),
            ("Conta Corrente:", c['bank_account']),
            ("CNPJ:", c['cnpj']),
            ("Beneficiário:", c['name']),
            ("", ""),
            ("Chave PIX:", c['pix_key']),
        ]
        
        for label, value in bank_data:
            last_row += 1
            if label:
                ws.cell(row=last_row, column=2, value=label)
                ws.cell(row=last_row, column=2).font = bank_label_font
                ws.cell(row=last_row, column=3, value=value)
                ws.cell(row=last_row, column=3).font = bank_value_font
            elif value:
                ws.cell(row=last_row, column=2, value=value)
                ws.cell(row=last_row, column=2).font = bank_value_font

        # Estende a área de impressão pra incluir o bloco de dados bancários,
        # que fica abaixo da tabela principal.
        ws.print_area = f'A1:N{last_row}'

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    except Exception as e:
        logger.error(f"Error generating billing Excel: {e}")
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Erro ao gerar relatório."
        ws['A1'].font = Font(size=14, bold=True, color="FF0000")
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


def generate_billing_pdf_report(movements: list, report_title: str = "Relatório de Faturamento", company: dict = None) -> bytes:
    """
    Generate PDF billing report following Bsoft layout style.
    Header: Logo left, company name + address center, generation info right
    Table with financial columns and total row
    """
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=10*mm,
        leftMargin=10*mm,
        topMargin=10*mm,
        bottomMargin=15*mm
    )

    c = merge_company(company)
    elements = []
    styles = getSampleStyleSheet()

    # ========== HEADER ==========
    logo_buffer = download_logo(company)
    header_elements = _build_pdf_header(styles, logo_buffer, report_title, company=company, content_width=doc.width)
    elements.extend(header_elements)

    # ========== STATISTICS BAR ==========
    total_records = len(movements)
    total_value = sum(m.get('service_value', 0) or 0 for m in movements)
    total_billed = sum(1 for m in movements if m.get('billed'))
    total_unbilled = total_records - total_billed
    value_str = f"R$ {total_value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

    stats_text = f"Total: {total_records}  |  Faturadas: {total_billed}  |  Não Faturadas: {total_unbilled}  |  Valor Total: {value_str}"

    stats_style = ParagraphStyle(
        'BillingStatsBar',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor(f'#{PRIMARY_COLOR}'),
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )

    stats_data = [[Paragraph(stats_text, stats_style)]]
    stats_table = Table(stats_data, colWidths=[760])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(f'#{HEADER_BG_COLOR}')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor(f'#{PRIMARY_COLOR}')),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 10))

    # ========== TABLE ==========
    data = [[
        'ID', 'Data/Hora', 'Tipo', 'Nº Container', 'Cliente', 'Placa',
        'Transportadora', 'Armador', 'Status', 'Tamanho',
        'Tipo Serviço', 'Nota Fiscal', 'Valor'
    ]]

    for m in movements:
        dt_brt = to_brt(m.get('created_at'))
        created_at = dt_brt.strftime('%d/%m/%Y %H:%M') if dt_brt else str(m.get('created_at', '-'))

        service_value = m.get('service_value')
        val_str = f"R$ {service_value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if service_value else '-'

        data.append([
            str(m.get('transaction_id', '-')),
            created_at,
            "ENTRADA" if m.get('operation_type') == 'ENTRADA' else "SAÍDA",
            m.get('container_number', '-'),
            m.get('client_name', '-') or '-',
            m.get('truck_plate', '-') or '-',
            m.get('transport_company', '-') or '-',
            m.get('shipping_line', '-') or '-',
            m.get('status', '-') or '-',
            m.get('size_type', '-') or '-',
            m.get('service_type', '-') or '-',
            m.get('invoice_number', '-') or '-',
            val_str
        ])

    # Total row
    data.append(['', '', '', '', '', '', '', '', '', '', '', 'TOTAL:', value_str])

    col_widths = [30, 58, 42, 62, 75, 45, 70, 55, 40, 42, 62, 52, 58]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{PRIMARY_COLOR}')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
        ('TOPPADDING', (0, 0), (-1, 0), 5),

        # Body styling
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        
        # Alignments
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),   # ID
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),   # Tipo
        ('ALIGN', (8, 1), (8, -1), 'CENTER'),   # Status
        ('ALIGN', (9, 1), (9, -1), 'CENTER'),   # Tamanho
        ('ALIGN', (12, 1), (12, -1), 'RIGHT'),  # Valor
        
        # Borders
        ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#CCCCCC')),
        ('BOX', (0, 0), (-1, -2), 1, colors.HexColor(f'#{PRIMARY_COLOR}')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8F8F8')]),

        # Total row styling
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor(f'#{HEADER_BG_COLOR}')),
        ('FONTNAME', (11, -1), (12, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (11, -1), (12, -1), 9),
        ('ALIGN', (11, -1), (11, -1), 'RIGHT'),
        ('ALIGN', (12, -1), (12, -1), 'RIGHT'),
        ('TOPPADDING', (0, -1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 6),
        ('BOX', (11, -1), (12, -1), 1, colors.HexColor(f'#{PRIMARY_COLOR}')),
    ]))

    elements.append(table)

    # Build with footer
    footer = _make_pdf_footer(c['name'])
    doc.build(elements, onFirstPage=footer, onLaterPages=footer)
    pdf_bytes = buffer.getvalue()
    buffer.close()

    return pdf_bytes


def generate_invoice_pdf(invoice: dict, movements: list, company: dict = None) -> bytes:
    """
    Generate PDF for a specific invoice following Bsoft layout style.
    Format: Landscape with header, client info, movements table, and total.
    """
    buffer = io.BytesIO()
    
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4), 
        rightMargin=10*mm, 
        leftMargin=10*mm, 
        topMargin=10*mm, 
        bottomMargin=15*mm
    )
    
    c = merge_company(company)
    elements = []
    styles = getSampleStyleSheet()

    # ========== HEADER SECTION ==========
    logo_buffer = download_logo(company)
    report_title = f"FATURA Nº {invoice.get('invoice_number', '-')}"
    header_elements = _build_pdf_header(styles, logo_buffer, report_title, company=company, content_width=doc.width)
    elements.extend(header_elements)
    
    # ========== CLIENT INFO BAR ==========
    total_value = sum(m.get('service_value', 0) or 0 for m in movements)
    total_str = f"R$ {total_value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    
    client_info_text = f"Cliente: {invoice.get('client_name', '-')}  |  CNPJ: {invoice.get('client_cnpj', '-') or '-'}  |  Movimentações: {len(movements)}  |  Valor Total: {total_str}"
    
    client_style = ParagraphStyle(
        'ClientInfo',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor(f'#{PRIMARY_COLOR}'),
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    client_data = [[Paragraph(client_info_text, client_style)]]
    client_table = Table(client_data, colWidths=[760])
    client_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(f'#{HEADER_BG_COLOR}')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor(f'#{PRIMARY_COLOR}')),
    ]))
    elements.append(client_table)
    elements.append(Spacer(1, 10))
    
    # ========== MOVEMENTS TABLE ==========
    table_data = [[
        'ID', 'Data/Hora', 'Tipo', 'Nº Container', 'Cliente', 'Placa', 
        'Transportadora', 'Armador', 'Status', 'Tamanho', 
        'Tipo de Serviço', 'Nota Fiscal', 'Valor da Operação'
    ]]
    
    for m in movements:
        _m_dt = to_brt(m.get('created_at'))
        mov_date = _m_dt.strftime('%d/%m/%Y %H:%M') if _m_dt else '-'
        
        service_value = m.get('service_value')
        value_str = f"R$ {service_value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if service_value else '-'
        
        table_data.append([
            str(m.get('transaction_id', '-')),
            mov_date,
            "ENTRADA" if m.get('operation_type') == 'ENTRADA' else "SAÍDA",
            m.get('container_number', '-'),
            m.get('client_name', '-') or '-',
            m.get('truck_plate', '-') or '-',
            m.get('transport_company', '-') or '-',
            m.get('shipping_line', '-') or '-',
            m.get('status', '-') or '-',
            m.get('size_type', '-') or '-',
            m.get('service_type', '-') or '-',
            m.get('invoice_number', '-') or '-',
            value_str
        ])
    
    # Total row
    table_data.append(['', '', '', '', '', '', '', '', '', '', '', 'TOTAL:', total_str])
    
    col_widths = [28, 55, 42, 62, 75, 42, 70, 52, 40, 42, 65, 50, 62]
    
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{PRIMARY_COLOR}')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
        ('TOPPADDING', (0, 0), (-1, 0), 5),
        
        # Body
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        
        # Alignments
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),   # ID
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),   # Tipo
        ('ALIGN', (8, 1), (8, -1), 'CENTER'),   # Status
        ('ALIGN', (9, 1), (9, -1), 'CENTER'),   # Tamanho
        ('ALIGN', (12, 1), (12, -1), 'RIGHT'),  # Valor
        
        # Borders
        ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#CCCCCC')),
        ('BOX', (0, 0), (-1, -2), 1, colors.HexColor(f'#{PRIMARY_COLOR}')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8F8F8')]),
        
        # Total row
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor(f'#{HEADER_BG_COLOR}')),
        ('FONTNAME', (11, -1), (12, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (11, -1), (12, -1), 9),
        ('ALIGN', (11, -1), (11, -1), 'RIGHT'),
        ('ALIGN', (12, -1), (12, -1), 'RIGHT'),
        ('TOPPADDING', (0, -1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 6),
        ('BOX', (11, -1), (12, -1), 1, colors.HexColor(f'#{PRIMARY_COLOR}')),
    ]))
    
    elements.append(table)
    
    # ========== NOTES SECTION ==========
    if invoice.get('notes'):
        elements.append(Spacer(1, 10))
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor(f'#{PRIMARY_COLOR}'),
            fontName='Helvetica-Bold',
            spaceAfter=5
        )
        elements.append(Paragraph("OBSERVAÇÕES:", section_style))
        notes_style = ParagraphStyle(
            'NotesStyle',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.black,
        )
        notes_data = [[Paragraph(invoice.get('notes', ''), notes_style)]]
        notes_table = Table(notes_data, colWidths=[760])
        notes_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FFF9E6')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#FFD700')),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(notes_table)

    # ========== DADOS BANCÁRIOS ==========
    # A versão Excel desta mesma fatura já traz os dados de pagamento — o PDF
    # (o formato que de fato vai pro cliente) precisa da mesma informação.
    elements.append(Spacer(1, 12))
    bank_title_style = ParagraphStyle(
        'BankTitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor(f'#{PRIMARY_COLOR}'),
        fontName='Helvetica-Bold',
        spaceAfter=4
    )
    elements.append(Paragraph("DADOS BANCÁRIOS PARA PAGAMENTO", bank_title_style))

    bank_rows = [
        ['Banco:', c['bank_name'], 'Agência:', c['bank_agency'], 'Conta Corrente:', c['bank_account']],
        ['CNPJ:', c['cnpj'], 'Beneficiário:', c['name'], 'Chave PIX:', c['pix_key']],
    ]
    bank_table = Table(bank_rows, colWidths=[55, 165, 65, 130, 75, 205])
    bank_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(f'#{HEADER_BG_COLOR}')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor(f'#{PRIMARY_COLOR}')),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
    ]))
    # Colunas de valor (1, 3, 5) ficam com peso normal, só os rótulos são negrito
    bank_table.setStyle(TableStyle([('FONTNAME', (c, 0), (c, -1), 'Helvetica') for c in (1, 3, 5)]))
    elements.append(bank_table)

    # Build with footer
    footer = _make_pdf_footer(c['name'])
    doc.build(elements, onFirstPage=footer, onLaterPages=footer)
    pdf_bytes = buffer.getvalue()
    buffer.close()

    return pdf_bytes


def generate_invoice_excel(invoice: dict, movements: list, company: dict = None) -> bytes:
    """Generate invoice Excel with Bsoft template style."""
    try:
        c = merge_company(company)
        wb = Workbook()
        ws = wb.active
        # Nomes de planilha no Excel: máx. 31 caracteres, sem \ / ? * [ ] :
        sheet_title = re.sub(r'[\\/?*\[\]:]', '-', f"Fatura {invoice.get('invoice_number', '')}").strip()
        ws.title = sheet_title[:31] or "Fatura"

        _inv_dt = to_brt(invoice.get('created_at'))
        created_at = _inv_dt.strftime('%d/%m/%Y %H:%M') if _inv_dt else str(invoice.get('created_at', '-'))

        # Calculate total value
        total_value = sum(m.get('service_value', 0) or 0 for m in movements)
        val_str = f"R$ {total_value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        
        title = f"FATURA Nº {invoice.get('invoice_number', '-')} - Cliente: {invoice.get('client_name', '-')}"
        stats_text = f"Total: {len(movements)} movimentações  |  Valor: {val_str}  |  Data: {created_at}"

        headers = [
            'ID', 'Data/Hora', 'Tipo', 'Nº Container', 'Cliente', 'Placa',
            'Transportadora', 'Armador', 'Status', 'Tamanho',
            'Tipo de Serviço', 'Nota Fiscal', 'Valor da Operação'
        ]

        data_rows = []
        for m in movements:
            _m_dt = to_brt(m.get('created_at'))
            mov_date = _m_dt.strftime('%d/%m/%Y %H:%M') if _m_dt else '-'
            data_rows.append([
                m.get('transaction_id', '-'),
                mov_date,
                "ENTRADA" if m.get('operation_type') == 'ENTRADA' else "SAÍDA",
                m.get('container_number', '-'),
                m.get('client_name', '-') or '-',
                m.get('truck_plate', '-') or '-',
                m.get('transport_company', '-') or '-',
                m.get('shipping_line', '-') or '-',
                m.get('status', '-') or '-',
                m.get('size_type', '-') or '-',
                m.get('service_type', '-') or '-',
                m.get('invoice_number', '-') or '-',
                m.get('service_value') if m.get('service_value') else 0,
            ])

        col_widths = {
            'B': 7.3, 'C': 13.7, 'D': 8, 'E': 14, 'F': 25, 'G': 12,
            'H': 20, 'I': 14, 'J': 6, 'K': 8, 'L': 16, 'M': 12, 'N': 18
        }

        # Center alignment for: ID(0), Tipo(2), Status(8), Tamanho(9)
        center_cols = {0, 2, 8, 9}

        _bsoft_style_excel(
            ws, title, stats_text,
            headers, data_rows, col_widths,
            center_cols=center_cols,
            right_align_cols={12},
            number_fmt_cols={12: 'R$ #,##0.00'},
            total_col=12,
            stats_text=stats_text,
            company_name=c['name'],
            logo_buffer=download_logo(company)
        )

        # Notes section
        data_start = 9  # Updated to match new header rows (8 = header row)
        notes_row = data_start + len(data_rows) + 2
        
        if invoice.get('notes'):
            ws.cell(row=notes_row, column=2, value="OBSERVAÇÕES:")
            ws.cell(row=notes_row, column=2).font = Font(name='Calibri', size=10, bold=True)
            ws.merge_cells(f'B{notes_row + 1}:N{notes_row + 1}')
            notes_cell = ws.cell(row=notes_row + 1, column=2, value=invoice.get('notes', ''))
            notes_cell.font = Font(name='Calibri', size=10)
            notes_cell.alignment = Alignment(wrap_text=True)
            notes_cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            ws.row_dimensions[notes_row + 1].height = 30
            bank_start_row = notes_row + 3  # Pular 1 linha após observações
        else:
            bank_start_row = notes_row + 1
        
        # ========== DADOS BANCÁRIOS ==========
        # Estilo para o título dos dados bancários
        bank_title_font = Font(name='Calibri', size=12, bold=True, color=PRIMARY_COLOR)
        bank_label_font = Font(name='Calibri', size=10, bold=True)
        bank_value_font = Font(name='Calibri', size=10)

        # Título "DADOS BANCÁRIOS PARA PAGAMENTO"
        ws.cell(row=bank_start_row, column=2, value="DADOS BANCÁRIOS PARA PAGAMENTO")
        ws.cell(row=bank_start_row, column=2).font = bank_title_font
        ws.merge_cells(start_row=bank_start_row, start_column=2, end_row=bank_start_row, end_column=6)

        # Dados do banco - SEM linha em branco entre Beneficiário e Chave PIX
        bank_data = [
            ("Banco:", c['bank_name']),
            ("Agência:", c['bank_agency']),
            ("Conta Corrente:", c['bank_account']),
            ("CNPJ:", c['cnpj']),
            ("Beneficiário:", c['name']),
            ("Chave PIX:", c['pix_key']),
        ]

        current_row = bank_start_row
        for label, value in bank_data:
            current_row += 1
            ws.cell(row=current_row, column=2, value=label)
            ws.cell(row=current_row, column=2).font = bank_label_font
            ws.cell(row=current_row, column=3, value=value)
            ws.cell(row=current_row, column=3).font = bank_value_font

        # Estende a área de impressão pra incluir observações e dados bancários,
        # que ficam abaixo da tabela principal.
        ws.print_area = f'A1:N{current_row}'

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    except Exception as e:
        logger.error(f"Error generating invoice Excel: {e}")
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Erro ao gerar fatura."
        ws['A1'].font = Font(size=14, bold=True, color="FF0000")
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()



def generate_intl_invoice_pdf(invoice: dict, company: dict = None) -> bytes:
    """
    Generate International Invoice PDF.
    Same style as movement reports but for international invoices.
    """
    c = merge_company(company)
    buffer = io.BytesIO()
    
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4, 
        rightMargin=15*mm, 
        leftMargin=15*mm, 
        topMargin=15*mm, 
        bottomMargin=20*mm
    )
    
    elements = []
    styles = getSampleStyleSheet()

    currency = invoice.get('currency', 'USD')

    # ========== HEADER SECTION ==========
    logo_buffer = download_logo(company)
    report_title = f"INVOICE Nº {invoice.get('invoice_number', '-')}"
    header_elements = _build_pdf_header(styles, logo_buffer, report_title, company=company, content_width=doc.width)
    elements.extend(header_elements)

    # ========== INVOICE INFO BAR ==========
    total_str = format_currency(invoice.get('total', 0), currency)
    info_text = f"Moeda: {invoice.get('currency', '-')}  |  Emissão: {invoice.get('issue_date', '-')}  |  Vencimento: {invoice.get('due_date', '-')}  |  Total: {total_str}"
    
    info_bar_style = ParagraphStyle(
        'InfoBar',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor(f'#{PRIMARY_COLOR}'),
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    info_bar_data = [[Paragraph(info_text, info_bar_style)]]
    info_bar_table = Table(info_bar_data, colWidths=[510])
    info_bar_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(f'#{HEADER_BG_COLOR}')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor(f'#{PRIMARY_COLOR}')),
    ]))
    elements.append(info_bar_table)
    elements.append(Spacer(1, 15))
    
    # ========== PAYER INFO (Apenas Pagador) ==========
    section_title_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=11,
        textColor=colors.HexColor(f'#{PRIMARY_COLOR}'),
        fontName='Helvetica-Bold',
        spaceBefore=5,
        spaceAfter=5
    )
    
    info_style = ParagraphStyle(
        'InfoText',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
        textColor=colors.black
    )
    
    payer_text = f"""
    <b>{invoice.get('payer_company', '-')}</b><br/>
    CNPJ: {invoice.get('payer_cnpj', '-') or '-'}<br/>
    Contato: {invoice.get('payer_contact', '-') or '-'}<br/>
    E-mail: {invoice.get('payer_email', '-') or '-'}<br/>
    {invoice.get('payer_address', '-')}
    """
    
    parties_data = [
        [Paragraph("PAGADOR / PAYER", section_title_style)],
        [Paragraph(payer_text, info_style)]
    ]
    
    parties_table = Table(parties_data, colWidths=[510])
    parties_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor(f'#{PRIMARY_COLOR}')),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{HEADER_BG_COLOR}')),
    ]))
    elements.append(parties_table)
    elements.append(Spacer(1, 15))
    
    # ========== ITEMS TABLE ==========
    elements.append(Paragraph("SERVIÇOS / SERVICES", section_title_style))
    elements.append(Spacer(1, 5))
    
    items_header = ['Descrição / Description', 'Qtd', 'Valor Unitário', 'Total']
    items_data = [items_header]

    item_desc_style = ParagraphStyle('ItemDesc', parent=styles['Normal'], fontSize=9, leading=12)

    for item in invoice.get('items', []):
        items_data.append([
            Paragraph(item.get('description', '-') or '-', item_desc_style),
            str(item.get('quantity', 1)),
            format_currency(item.get('unit_price', 0), currency),
            format_currency(item.get('total', 0), currency)
        ])
    
    # Total row
    items_data.append(['', '', 'TOTAL:', total_str])
    
    col_widths = [260, 50, 100, 100]
    
    items_table = Table(items_data, colWidths=col_widths, repeatRows=1)
    items_table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{PRIMARY_COLOR}')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        
        # Body
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        
        # Alignments
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),    # Description
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Qty
        ('ALIGN', (2, 1), (2, -1), 'RIGHT'),   # Unit price
        ('ALIGN', (3, 1), (3, -1), 'RIGHT'),   # Total
        
        # Borders
        ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#CCCCCC')),
        ('BOX', (0, 0), (-1, -2), 1, colors.HexColor(f'#{PRIMARY_COLOR}')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8F8F8')]),
        
        # Total row
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor(f'#{HEADER_BG_COLOR}')),
        ('FONTNAME', (2, -1), (3, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (2, -1), (3, -1), 11),
        ('ALIGN', (2, -1), (2, -1), 'RIGHT'),
        ('ALIGN', (3, -1), (3, -1), 'RIGHT'),
        ('TOPPADDING', (0, -1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
        ('BOX', (2, -1), (3, -1), 1, colors.HexColor(f'#{PRIMARY_COLOR}')),
    ]))
    
    elements.append(items_table)
    
    # ========== NOTES SECTION ==========
    if invoice.get('notes'):
        elements.append(Spacer(1, 15))
        elements.append(Paragraph("OBSERVAÇÕES / NOTES", section_title_style))
        elements.append(Spacer(1, 5))
        
        notes_style = ParagraphStyle(
            'Notes',
            parent=styles['Normal'],
            fontSize=9,
            leading=12,
            textColor=colors.HexColor('#333333'),
            borderColor=colors.HexColor('#CCCCCC'),
            borderWidth=1,
            borderPadding=8
        )
        
        notes_data = [[Paragraph(invoice['notes'], notes_style)]]
        notes_table = Table(notes_data, colWidths=[510])
        notes_table.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FAFAFA')),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(notes_table)
    
    # ========== BARCODE SECTION ==========
    elements.append(Spacer(1, 25))
    
    # Generate barcode for invoice number
    import barcode
    from barcode.writer import ImageWriter
    
    barcode_buffer = io.BytesIO()
    try:
        invoice_num_str = str(invoice.get('invoice_number', '0')).zfill(3)
        code128 = barcode.get_barcode_class('code128')
        barcode_obj = code128(invoice_num_str, writer=ImageWriter())
        barcode_obj.write(barcode_buffer, options={
            'module_width': 0.3,
            'module_height': 12,
            'font_size': 10,
            'text_distance': 5,
            'quiet_zone': 2
        })
        barcode_buffer.seek(0)
        barcode_image = Image(barcode_buffer, width=120, height=50)
    except Exception as e:
        logger.error(f"Error generating barcode: {e}")
        barcode_image = Paragraph(f"[{invoice.get('invoice_number', '-')}]", styles['Normal'])
    
    # User and timestamp info
    created_by = invoice.get('created_by_name', 'Sistema')
    print_date = now_brt().strftime('%d/%m/%Y')
    
    user_info_style = ParagraphStyle(
        'UserInfo',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.black,
        fontName='Helvetica-Bold',
        leading=14
    )
    
    user_info = [
        Paragraph(f"Usuário: {created_by}", user_info_style),
        Spacer(1, 8),
        Paragraph(f"Data da impressão: {print_date}", user_info_style),
    ]
    
    # Barcode section table: Barcode | User Info
    barcode_section_data = [[barcode_image, user_info]]
    barcode_section_table = Table(barcode_section_data, colWidths=[150, 360])
    barcode_section_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'LEFT'),
        ('LEFTPADDING', (1, 0), (1, 0), 20),
    ]))
    elements.append(barcode_section_table)
    
    # ========== FOOTER INFO ==========
    elements.append(Spacer(1, 15))
    
    footer_style = ParagraphStyle(
        'FooterInfo',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#666666'),
        alignment=TA_CENTER
    )
    
    elements.append(Paragraph(f"Documento gerado em {now_brt().strftime('%d/%m/%Y')} | ContainerLogix - {c['name']}", footer_style))

    # Build PDF
    footer = _make_pdf_footer(c['name'])
    doc.build(elements, onFirstPage=footer, onLaterPages=footer)
    return buffer.getvalue()


def generate_expense_report_pdf(report: dict, company: dict = None) -> bytes:
    """
    Gera o PDF final da Prestação de Contas: período, depósitos recebidos,
    lançamentos de compra, saldo, código de barras de controle e os recibos
    anexados a cada compra.
    """
    c = merge_company(company)
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15*mm,
        leftMargin=15*mm,
        topMargin=15*mm,
        bottomMargin=20*mm
    )

    elements = []
    styles = getSampleStyleSheet()

    def money(value):
        return format_currency(value)

    # ========== HEADER SECTION ==========
    logo_buffer = download_logo(company)
    report_title = f"PRESTAÇÃO DE CONTAS Nº {report.get('report_number_formatted', '-')}"
    header_elements = _build_pdf_header(styles, logo_buffer, report_title, company=company, content_width=doc.width)
    elements.extend(header_elements)

    # ========== INFO BAR: período / responsável / status ==========
    status_label = 'Concluída' if report.get('status') == 'CONCLUIDA' else 'Em Andamento'
    info_text = (
        f"Período: {fmt_date(report.get('period_start'))} a {fmt_date(report.get('period_end'))}"
        f"  |  Responsável: {report.get('created_by_name', '-')}"
        f"  |  Status: {status_label}"
    )
    info_style = ParagraphStyle(
        'ExpenseInfoBar', parent=styles['Normal'], fontSize=10,
        textColor=colors.HexColor(f'#{PRIMARY_COLOR}'), alignment=TA_CENTER, fontName='Helvetica-Bold'
    )
    info_data = [[Paragraph(info_text, info_style)]]
    info_table = Table(info_data, colWidths=[510])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(f'#{HEADER_BG_COLOR}')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor(f'#{PRIMARY_COLOR}')),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 15))

    section_title_style = ParagraphStyle(
        'ExpenseSectionTitle', parent=styles['Normal'], fontSize=11,
        textColor=colors.HexColor(f'#{PRIMARY_COLOR}'), fontName='Helvetica-Bold', spaceAfter=5
    )
    # Texto livre (nomes, observações) precisa quebrar linha dentro da célula —
    # string simples não quebra e pode vazar da coluna.
    obs_style = ParagraphStyle('ExpenseObsCell', parent=styles['Normal'], fontSize=8, leading=10)

    # ========== DEPÓSITOS RECEBIDOS (antes dos lançamentos de compra) ==========
    elements.append(Paragraph("DEPÓSITOS RECEBIDOS", section_title_style))
    deposits = report.get('deposits', []) or []
    deposits_data = [['Data', 'Enviado Por', 'Valor']]
    for d in deposits:
        deposits_data.append([fmt_date(d.get('date')), Paragraph(d.get('sent_by', '-') or '-', obs_style), money(d.get('amount'))])
    deposits_data.append(['', 'TOTAL DEPÓSITOS:', money(report.get('total_deposits'))])

    deposits_table = Table(deposits_data, colWidths=[100, 300, 110], repeatRows=1)
    deposits_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{PRIMARY_COLOR}')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#CCCCCC')),
        ('BOX', (0, 0), (-1, -2), 1, colors.HexColor(f'#{PRIMARY_COLOR}')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8F8F8')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor(f'#{HEADER_BG_COLOR}')),
        ('FONTNAME', (1, -1), (2, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(deposits_table)
    elements.append(Spacer(1, 15))

    # ========== LANÇAMENTOS DE COMPRAS ==========
    elements.append(Paragraph("LANÇAMENTOS DE COMPRAS", section_title_style))
    purchases = report.get('purchases', []) or []
    purchases = sorted(purchases, key=lambda p: p.get('purchase_date') or '')
    purchases_data = [['Local de Compra', 'Data', 'Valor', 'Observação']]
    for p in purchases:
        purchases_data.append([
            Paragraph(p.get('supplier_name') or '-', obs_style),
            fmt_date(p.get('purchase_date')),
            money(p.get('amount')),
            Paragraph(p.get('observation') or '-', obs_style)
        ])
    purchases_data.append(['', 'TOTAL COMPRAS:', money(report.get('total_purchases')), ''])

    purchases_table = Table(purchases_data, colWidths=[160, 80, 90, 180], repeatRows=1)
    purchases_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{PRIMARY_COLOR}')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#CCCCCC')),
        ('BOX', (0, 0), (-1, -2), 1, colors.HexColor(f'#{PRIMARY_COLOR}')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8F8F8')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor(f'#{HEADER_BG_COLOR}')),
        ('FONTNAME', (1, -1), (2, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(purchases_table)
    elements.append(Spacer(1, 15))

    # ========== TOTAIS / SALDO ==========
    balance = report.get('balance', 0) or 0
    if balance > 0:
        balance_label = 'VALOR A RESSARCIR AO FUNCIONÁRIO'
        balance_color = '#B45309'
    elif balance < 0:
        balance_label = 'SALDO A DEVOLVER PELO FUNCIONÁRIO'
        balance_color = '#B91C1C'
    else:
        balance_label = 'QUITADO'
        balance_color = f'#{PRIMARY_COLOR}'

    totals_data = [
        ['Total de Compras:', money(report.get('total_purchases'))],
        ['Total de Depósitos:', money(report.get('total_deposits'))],
        [balance_label + ':', money(abs(balance))],
    ]
    totals_table = Table(totals_data, colWidths=[300, 210])
    totals_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, 1), 10),
        ('FONTSIZE', (0, 2), (-1, 2), 12),
        ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 2), (-1, 2), colors.HexColor(balance_color)),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('LINEABOVE', (0, 2), (-1, 2), 1, colors.HexColor(balance_color)),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor(f'#{PRIMARY_COLOR}')),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(f'#{HEADER_BG_COLOR}')),
    ]))
    elements.append(totals_table)

    # ========== BARCODE SECTION ==========
    elements.append(Spacer(1, 20))

    import barcode
    from barcode.writer import ImageWriter

    barcode_buffer = io.BytesIO()
    try:
        barcode_str = str(report.get('report_number_formatted', '0'))
        code128 = barcode.get_barcode_class('code128')
        barcode_obj = code128(barcode_str, writer=ImageWriter())
        barcode_obj.write(barcode_buffer, options={
            'module_width': 0.3,
            'module_height': 12,
            'font_size': 10,
            'text_distance': 5,
            'quiet_zone': 2
        })
        barcode_buffer.seek(0)
        barcode_image = Image(barcode_buffer, width=140, height=50)
    except Exception as e:
        logger.error(f"Error generating barcode: {e}")
        barcode_image = Paragraph(f"[{report.get('report_number_formatted', '-')}]", styles['Normal'])

    created_by = report.get('created_by_name', 'Sistema')
    print_date = now_brt().strftime('%d/%m/%Y %H:%M')

    user_info_style = ParagraphStyle(
        'ExpenseUserInfo', parent=styles['Normal'], fontSize=10,
        textColor=colors.black, fontName='Helvetica-Bold', leading=14
    )
    user_info = [
        Paragraph(f"Usuário: {created_by}", user_info_style),
        Spacer(1, 6),
        Paragraph(f"Data da impressão: {print_date}", user_info_style),
    ]

    barcode_section_data = [[barcode_image, user_info]]
    barcode_section_table = Table(barcode_section_data, colWidths=[160, 350])
    barcode_section_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (1, 0), (1, 0), 20),
    ]))
    elements.append(barcode_section_table)

    # ========== RECIBOS ANEXADOS ==========
    if any(p.get('receipts') for p in purchases):
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("RECIBOS ANEXADOS", section_title_style))
        elements.append(Spacer(1, 5))

        receipt_label_style = ParagraphStyle(
            'ExpenseReceiptLabel', parent=styles['Normal'], fontSize=9,
            fontName='Helvetica-Bold', textColor=colors.HexColor(f'#{PRIMARY_COLOR}'), spaceAfter=4
        )
        unavailable_style = ParagraphStyle('ExpenseReceiptUnavailable', parent=styles['Normal'], fontSize=8, textColor=colors.grey)

        images_per_row = 3
        for p in purchases:
            receipts = p.get('receipts', []) or []
            if not receipts:
                continue

            label = f"{p.get('supplier_name') or '-'} — {fmt_date(p.get('purchase_date'))} — {money(p.get('amount'))}"
            elements.append(Paragraph(label, receipt_label_style))

            rows = []
            current_row = []
            for r in receipts:
                img_flowable = None
                try:
                    url = r.get('url', '') or ''
                    marker = '/expense_reports/'
                    if marker in url:
                        relative = url.split(marker, 1)[1]
                        file_path = UPLOADS_DIR / 'expense_reports' / relative
                        if file_path.exists():
                            img_flowable = Image(str(file_path), width=150, height=200, kind='proportional')
                except Exception as e:
                    logger.error(f"Error loading receipt image: {e}")
                if img_flowable is None:
                    img_flowable = Paragraph("[Recibo indisponível]", unavailable_style)
                current_row.append(img_flowable)
                if len(current_row) == images_per_row:
                    rows.append(current_row)
                    current_row = []
            if current_row:
                while len(current_row) < images_per_row:
                    current_row.append('')
                rows.append(current_row)

            receipts_table = Table(rows, colWidths=[170] * images_per_row)
            receipts_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ]))
            elements.append(receipts_table)
            elements.append(Spacer(1, 8))

    # Build with page footer
    footer = _make_pdf_footer(c['name'])
    doc.build(elements, onFirstPage=footer, onLaterPages=footer)
    return buffer.getvalue()


VEHICLE_CHECKLIST_TEMPLATE_SECTIONS = list(VEHICLE_CHECKLIST_TEMPLATE.keys())


def generate_vehicle_checklist_pdf(checklist: dict, company: dict = None) -> bytes:
    """Gera o PDF do Checklist de Veículo (LVT), com os itens e respostas SIM/NÃO por seção."""
    c = merge_company(company)
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=20 * mm
    )

    elements = []
    styles = getSampleStyleSheet()
    # Largura útil real da página (A4 - margens), não um valor fixo — evita vazar
    # da margem, que era o que acontecia com o 540 fixo usado antes aqui.
    SECTION_WIDTH = doc.width

    label_style = ParagraphStyle('CLLabel', parent=styles['Normal'], fontSize=7.5, fontName='Helvetica', textColor=colors.HexColor('#555555'))
    value_style = ParagraphStyle('CLValue', parent=styles['Normal'], fontSize=9.5, fontName='Helvetica-Bold', textColor=colors.black)
    section_title_style = ParagraphStyle('CLSection', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', textColor=colors.white)
    item_style = ParagraphStyle('CLItem', parent=styles['Normal'], fontSize=8, fontName='Helvetica', leading=10)

    # ========== HEADER ==========
    logo_buffer = download_logo(company)
    header_elements = _build_pdf_header(styles, logo_buffer, f"CHECKLIST DE VEÍCULO Nº {checklist.get('checklist_number', '-')}", company=company, content_width=doc.width)
    elements.extend(header_elements)

    # ========== FALHA / ATENÇÃO ==========
    has_failure = any(
        item.get('answer') == 'NAO'
        for section in VEHICLE_CHECKLIST_TEMPLATE_SECTIONS
        for item in (checklist.get(f"{section}_items") or [])
    )
    if has_failure:
        warn_style = ParagraphStyle('CLWarn', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', textColor=colors.white, alignment=TA_CENTER)
        warn_table = Table([[Paragraph("ATENÇÃO: há item(ns) reprovado(s) neste checklist. O carregamento deve ser cancelado.", warn_style)]], colWidths=[SECTION_WIDTH])
        warn_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#DC2626')),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(warn_table)
        elements.append(Spacer(1, 8))

    # ========== INFO DO CABEÇALHO ==========
    def info_pair(label, value):
        return [Paragraph(label, label_style), Paragraph(str(value) if value not in (None, '') else '-', value_style)]

    info_rows = [
        [info_pair("Expedidor", checklist.get('expedidor')), info_pair("UN", checklist.get('un'))],
        [info_pair("Data/Hora da Vistoria", fmt_datetime(checklist.get('inspection_datetime'))), info_pair("Cód. Agendamento", checklist.get('scheduling_code'))],
        [info_pair("Cliente", checklist.get('client_name')), info_pair("Transportadora", checklist.get('transport_company_name'))],
        [info_pair("Número ORP/ODP", checklist.get('orp_odp_number')), info_pair("Número da NF", checklist.get('nf_number'))],
        [info_pair("Motorista", checklist.get('driver_name')), info_pair("CPF Motorista", checklist.get('driver_cpf'))],
        [info_pair("CNH Nº / Categoria", f"{checklist.get('cnh_number') or '-'} / {checklist.get('cnh_category') or '-'}"), info_pair("CNH Vencimento", fmt_date(checklist.get('cnh_expiry')))],
        [info_pair("Produto(s)", checklist.get('products_description')), info_pair("Código SAP", checklist.get('sap_code'))],
        [info_pair("Placa do Cavalo / Ano", f"{checklist.get('cavalo_plate') or '-'} / {checklist.get('cavalo_year') or '-'}"), info_pair("", "")],
        [info_pair("Placa Carreta 1 / Ano", f"{checklist.get('carreta1_plate') or '-'} / {checklist.get('carreta1_year') or '-'}"), info_pair("Capacidade Carreta 1", checklist.get('carreta1_capacity'))],
        [info_pair("Placa Carreta 2 / Ano", f"{checklist.get('carreta2_plate') or '-'} / {checklist.get('carreta2_year') or '-'}"), info_pair("Capacidade Carreta 2", checklist.get('carreta2_capacity'))],
    ]
    info_table = Table(info_rows, colWidths=[SECTION_WIDTH / 2, SECTION_WIDTH / 2])
    info_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 10))

    # ========== SEÇÕES DE ITENS ==========
    for section in VEHICLE_CHECKLIST_TEMPLATE_SECTIONS:
        items = checklist.get(f"{section}_items") or []
        if not items:
            continue
        has_expiry = section == 'documentos'

        sec_header = Table([[Paragraph(VEHICLE_CHECKLIST_SECTION_LABELS[section], section_title_style)]], colWidths=[SECTION_WIDTH])
        sec_header.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(f'#{PRIMARY_COLOR}')),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(sec_header)

        header_row = ["Descrição", "Sim", "Não"] + (["Vencimento"] if has_expiry else [])
        rows = [header_row]
        for item in items:
            answer = item.get('answer')
            row = [
                Paragraph(item.get('text', ''), item_style),
                "X" if answer == 'SIM' else "",
                "X" if answer == 'NAO' else "",
            ]
            if has_expiry:
                row.append(fmt_date(item.get('expiry')))
            rows.append(row)

        # Colunas fixas (Sim/Não/Vencimento) + Descrição absorvendo o resto da
        # largura útil da página — nunca ultrapassa a margem.
        col_widths = [SECTION_WIDTH - 140, 35, 35, 70] if has_expiry else [SECTION_WIDTH - 70, 35, 35]
        items_table = Table(rows, colWidths=col_widths, repeatRows=1)
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{HEADER_BG_COLOR}')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(items_table)
        elements.append(Spacer(1, 8))

    # ========== PRODUTOS / RÓTULOS DE RISCO ==========
    products = checklist.get('products') or []
    if products:
        prod_header = Table([[Paragraph("Produtos Transportados", section_title_style)]], colWidths=[SECTION_WIDTH])
        prod_header.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(f'#{PRIMARY_COLOR}')),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(prod_header)

        prod_rows = [["Produto", "ONU", "Nº de Risco", "Subclasse"]]
        for p in products:
            prod_rows.append([
                Paragraph(p.get('product') or '-', item_style),
                p.get('un_number') or '-', p.get('risk_number') or '-', p.get('subclass') or '-'
            ])
        # ONU/Risco/Subclasse são códigos curtos e fixos; Produto absorve o resto.
        prod_table = Table(prod_rows, colWidths=[SECTION_WIDTH - 300, 100, 100, 100])
        prod_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{HEADER_BG_COLOR}')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(prod_table)
        elements.append(Spacer(1, 10))

    # ========== KIT: VALIDADES / ÚLTIMAS 3 VIAGENS ==========
    extra_rows = [
        [info_pair("Validade Calço/Extintor 1", fmt_date(checklist.get('kit_validity_1'))), info_pair("Validade 2", fmt_date(checklist.get('kit_validity_2')))],
        [info_pair("Validade 3", fmt_date(checklist.get('kit_validity_3'))), info_pair("", "")],
        [info_pair("Últ. Viagem - Produto 1", checklist.get('last_trip_product_1')), info_pair("Produto 2", checklist.get('last_trip_product_2'))],
        [info_pair("Produto 3", checklist.get('last_trip_product_3')), info_pair("", "")],
    ]
    extra_table = Table(extra_rows, colWidths=[SECTION_WIDTH / 2, SECTION_WIDTH / 2])
    extra_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(extra_table)
    elements.append(Spacer(1, 10))

    # ========== OBSERVAÇÕES ==========
    if checklist.get('observations'):
        obs_header = Table([[Paragraph("Observações", section_title_style)]], colWidths=[SECTION_WIDTH])
        obs_header.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(f'#{PRIMARY_COLOR}')),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(obs_header)
        obs_content_style = ParagraphStyle('CLObs', parent=styles['Normal'], fontSize=9, fontName='Helvetica')
        obs_table = Table([[Paragraph(checklist['observations'], obs_content_style)]], colWidths=[SECTION_WIDTH])
        obs_table.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(obs_table)
        elements.append(Spacer(1, 10))

    # ========== RESPONSÁVEIS ==========
    resp_header = Table([[Paragraph("Responsáveis", section_title_style)]], colWidths=[SECTION_WIDTH])
    resp_header.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(f'#{PRIMARY_COLOR}')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(resp_header)
    resp_rows = [
        [info_pair("Resp. Transportadora", checklist.get('transport_responsible_name')), info_pair("RG", checklist.get('transport_responsible_rg'))],
        [info_pair("Recebedor da LVT", checklist.get('lvt_receiver_name')), info_pair("Matrícula", checklist.get('lvt_receiver_registration'))],
        [info_pair("Resp. pela Vistoria", checklist.get('inspection_responsible_name')), info_pair("Matrícula", checklist.get('inspection_responsible_registration'))],
        [info_pair("Registro de Mérito", checklist.get('merit_record')), info_pair("Registro de Ocorrências", checklist.get('occurrence_record'))],
        [info_pair("Documento do Condutor", checklist.get('driver_document')), info_pair("Liberação do Veículo", fmt_datetime(checklist.get('release_datetime')))],
    ]
    resp_table = Table(resp_rows, colWidths=[SECTION_WIDTH / 2, SECTION_WIDTH / 2])
    resp_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(resp_table)

    footer = _make_pdf_footer(c['name'])
    doc.build(elements, onFirstPage=footer, onLaterPages=footer)
    return buffer.getvalue()


# ==================== CHECKLIST DE VEÍCULO — MODELO PETROBRAS (LVT) ====================
# Réplica das informações da "Lista de Verificação de Transporte - LVT N° 14" (Petrobras),
# usada quando o cliente contratante do checklist é a Manuport. O logo oficial da Petrobras
# não está embutido (não temos o arquivo da marca) — usamos um cabeçalho em texto no lugar.

def generate_petrobras_lvt_pdf(checklist: dict, company: dict = None) -> bytes:
    """Gera o PDF do Checklist de Veículo no formato LVT da Petrobras."""
    c = merge_company(company)
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=15 * mm
    )

    elements = []
    styles = getSampleStyleSheet()
    # Largura útil real da página (A4 - margens), não o valor fixo 546 usado antes
    # (vazava ~19pt da margem em todas as tabelas do documento). As tabelas abaixo
    # que não usam SECTION_WIDTH diretamente têm suas larguras escaladas por
    # _scale, preservando as proporções originais do layout (réplica do formulário
    # oficial da Petrobras) dentro da área realmente disponível.
    SECTION_WIDTH = doc.width
    _scale = SECTION_WIDTH / 546
    BLACK = colors.black

    label_style = ParagraphStyle('LVTLabel', parent=styles['Normal'], fontSize=7.5, fontName='Helvetica')
    value_style = ParagraphStyle('LVTValue', parent=styles['Normal'], fontSize=8.5, fontName='Helvetica-Bold')
    section_label_style = ParagraphStyle('LVTSectionLabel', parent=styles['Normal'], fontSize=7.5, fontName='Helvetica-Bold', textColor=BLACK)
    item_style = ParagraphStyle('LVTItem', parent=styles['Normal'], fontSize=7.5, fontName='Helvetica', leading=9.5)
    static_style = ParagraphStyle('LVTStatic', parent=styles['Normal'], fontSize=7.5, fontName='Helvetica', leading=10)
    static_bold_style = ParagraphStyle('LVTStaticBold', parent=styles['Normal'], fontSize=8, fontName='Helvetica-Bold', leading=10)
    heading_style = ParagraphStyle('LVTHeading', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold', alignment=TA_CENTER)

    def field(label, value):
        return Paragraph(f"<b>{label}:</b> {value if value not in (None, '') else ''}", label_style)

    # ========== CABEÇALHO (logo em texto, já que não temos o arquivo oficial) ==========
    logo_text = Paragraph("<b>BR</b><br/><font size=6>PETROBRAS</font>", ParagraphStyle('LVTLogo', parent=styles['Normal'], fontSize=16, fontName='Helvetica-Bold', textColor=colors.HexColor('#006633'), alignment=TA_CENTER, leading=16))
    title_block = [
        [Paragraph("LISTA DE VERIFICAÇÃO DE TRANSPORTE - LVT N° 14", heading_style), Paragraph("Versão: 01/2023 Rev. 3", ParagraphStyle('LVTVer', parent=styles['Normal'], fontSize=7, alignment=TA_RIGHT))],
        [Paragraph("INSPEÇÃO DE VEÍCULOS PARA TRANSPORTE RODOVIÁRIO DE DERIVADOS DE PETRÓLEO", ParagraphStyle('LVTSub', parent=styles['Normal'], fontSize=8, fontName='Helvetica-BoldOblique', alignment=TA_CENTER)), ""],
    ]
    title_table = Table(title_block, colWidths=[440 * _scale, 106 * _scale])
    title_table.setStyle(TableStyle([
        ('SPAN', (0, 1), (1, 1)),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
    ]))
    header_table = Table([[logo_text, title_table]], colWidths=[70 * _scale, 476 * _scale])
    header_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1.2, BLACK),
        ('INNERGRID', (0, 0), (-1, -1), 1.2, BLACK),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 4))

    # ========== INFORMAÇÕES DO CABEÇALHO (mesma ordem/agrupamento do documento original) ==========
    cnh_line = f"{checklist.get('cnh_number') or '-'} / {checklist.get('cnh_category') or '-'} / {fmt_date(checklist.get('cnh_expiry'))}"
    header_rows = [
        [field("Expedidor", checklist.get('expedidor') or "Petróleo Brasileiro S.A."), field("UN", checklist.get('un')), field("Data e Hora da vistoria", fmt_datetime(checklist.get('inspection_datetime')))],
        [field("Endereço da UN", checklist.get('un_address')), field("Telefones", checklist.get('phone')), field("Fax", checklist.get('fax'))],
        [field("Cliente", checklist.get('client_name')), field("Número ORP / ODP", checklist.get('orp_odp_number')), field("Número da NF (Descarga)", checklist.get('nf_number'))],
        [field("Transportadora", checklist.get('transport_company_name')), field("Motorista", checklist.get('driver_name')), field("Produto(s)", checklist.get('products_description'))],
        [field("Placa do cavalo", checklist.get('cavalo_plate')), field("Placa da carreta 1 / Ano", f"{checklist.get('carreta1_plate') or '-'} / {checklist.get('carreta1_year') or '-'}"), field("Placa da carreta 2 / Ano", f"{checklist.get('carreta2_plate') or '-'} / {checklist.get('carreta2_year') or '-'}")],
        [field("Ano de Fabricação do Cavalo", checklist.get('cavalo_year')), field("Capacidade carreta 1", checklist.get('carreta1_capacity')), field("Capacidade carreta 2", checklist.get('carreta2_capacity'))],
        [field("Nº CNH / Categoria / Vencimento", cnh_line), field("Código SAP", checklist.get('sap_code')), ""],
    ]
    header_info_table = Table(header_rows, colWidths=[SECTION_WIDTH / 3] * 3)
    header_info_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, BLACK),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, BLACK),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(header_info_table)
    elements.append(Spacer(1, 6))

    # ========== ATENÇÃO / REPROVAÇÃO ==========
    has_failure = any(
        item.get('answer') == 'NAO'
        for section in VEHICLE_CHECKLIST_TEMPLATE_SECTIONS
        for item in (checklist.get(f"{section}_items") or [])
    )
    if has_failure:
        warn_style = ParagraphStyle('LVTWarn', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold', textColor=colors.white, alignment=TA_CENTER)
        warn_table = Table([[Paragraph("ATENÇÃO: há item que não atende aos requisitos — o carregamento deve ser cancelado.", warn_style)]], colWidths=[SECTION_WIDTH])
        warn_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#DC2626')), ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5)]))
        elements.append(warn_table)
        elements.append(Spacer(1, 6))

    # ========== DESCRIÇÃO — itens SIM/NÃO por seção ==========
    desc_header = Table([["DESCRIÇÃO", "SIM", "NÃO", "DOCUMENTOS / VENCIMENTO"]], colWidths=[380 * _scale, 30 * _scale, 30 * _scale, 106 * _scale])
    desc_header.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F0F0F0')),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('BOX', (0, 0), (-1, -1), 1, BLACK),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, BLACK),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(desc_header)

    for section in VEHICLE_CHECKLIST_TEMPLATE_SECTIONS:
        items = checklist.get(f"{section}_items") or []
        if not items:
            continue
        has_expiry = section == 'documentos'
        sec_rows = []
        for item in items:
            answer = item.get('answer')
            sec_rows.append([
                Paragraph(item.get('text', ''), item_style),
                "X" if answer == 'SIM' else "",
                "X" if answer == 'NAO' else "",
                fmt_date(item.get('expiry')) if has_expiry else "",
            ])
        sec_table = Table(
            [[Paragraph(VEHICLE_CHECKLIST_SECTION_LABELS[section], section_label_style), "", "", ""]] + sec_rows,
            colWidths=[380 * _scale, 30 * _scale, 30 * _scale, 106 * _scale]
        )
        sec_style = [
            ('SPAN', (0, 0), (-1, 0)),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9D9D9')),
            ('ALIGN', (1, 1), (2, -1), 'CENTER'),
            ('FONTNAME', (1, 1), (2, -1), 'Helvetica-Bold'),
            ('BOX', (0, 0), (-1, -1), 1, BLACK),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, BLACK),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 2.5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]
        sec_table.setStyle(TableStyle(sec_style))
        elements.append(sec_table)

    elements.append(Spacer(1, 8))

    # ========== DECLARAÇÃO DO MOTORISTA / TRANSPORTADORA (texto padrão do documento) ==========
    declaration_text = (
        "<b>A transportadora e o motorista declaram que:</b> "
        "1) Observarão as Normas do Decreto nº 96.044 de 18 de maio de 1988, e demais legislações pertinentes, "
        "para o carregamento e transporte. "
        "2) Manifestam a concordância e anuência com a participação do motorista nas operações de carregamento, "
        "descarregamento e transbordo de carga (quando aplicável). "
        "3) O motorista está com Curso Básico em Norma Regulamentadora nº 20 concluído e válido, com carga horária "
        "de 8 horas e com certificado emitido. "
        "4) O motorista está capacitado para trabalho em altura (quando aplicável), para atividades acima de 2,00m "
        "do nível inferior, onde haja risco de queda, conforme atendimento à Norma Regulamentadora nº 35. "
        "5) O motorista está portando a FISPQ do produto."
    )
    decl_table = Table([[Paragraph(declaration_text, static_style)]], colWidths=[SECTION_WIDTH])
    decl_table.setStyle(TableStyle([('BOX', (0, 0), (-1, -1), 1, BLACK), ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6), ('LEFTPADDING', (0, 0), (-1, -1), 6)]))
    elements.append(decl_table)
    elements.append(Spacer(1, 10))

    # ========== PRODUTOS / RÓTULOS DE RISCO ==========
    products = checklist.get('products') or []
    prod_title = Table([[Paragraph("PRODUTOS TRANSPORTADOS (válido para o(s) seguinte(s) produto(s))", static_bold_style)]], colWidths=[SECTION_WIDTH])
    prod_title.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F0F0F0')), ('BOX', (0, 0), (-1, -1), 1, BLACK), ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4), ('LEFTPADDING', (0, 0), (-1, -1), 6)]))
    elements.append(prod_title)
    prod_rows = [["Produto", "ONU", "Nº de Risco", "Subclasse"]]
    for p in products:
        prod_rows.append([
            Paragraph(p.get('product') or '-', item_style),
            p.get('un_number') or '-', p.get('risk_number') or '-', p.get('subclass') or '-'
        ])
    if len(prod_rows) == 1:
        prod_rows.append(['-', '-', '-', '-'])
    prod_table = Table(prod_rows, colWidths=[246 * _scale, 100 * _scale, 100 * _scale, 100 * _scale])
    prod_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F0F0F0')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('BOX', (0, 0), (-1, -1), 1, BLACK),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, BLACK),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(prod_table)
    elements.append(Spacer(1, 10))

    # ========== UNIFORME E EPI (texto de referência fixo do documento) ==========
    uniforme_epi_text = (
        "<b>UNIFORME:</b> Calça de brim ou algodão; Camisa de brim ou algodão (mangas compridas); Bota de couro fechada.<br/><br/>"
        "<b>EQUIPAMENTOS DE PROTEÇÃO INDIVIDUAL (EPI)</b> — todos devem possuir Certificado de Aprovação (CA): "
        "Luvas de PVC; Capacete com jugular; Protetor auricular; Proteção respiratória semi-facial com filtro químico "
        "para vapores orgânicos; Óculos ampla visão com proteção lateral; Cinto de segurança para trabalho em altura "
        "(uso obrigatório nas baias de carregamento, quando acessar a parte superior do caminhão)."
    )
    epi_table = Table([[Paragraph("UNIFORME E EQUIPAMENTOS DE PROTEÇÃO INDIVIDUAL (EPI)", static_bold_style)], [Paragraph(uniforme_epi_text, static_style)]], colWidths=[SECTION_WIDTH])
    epi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#F0F0F0')),
        ('BOX', (0, 0), (-1, -1), 1, BLACK),
        ('LINEBELOW', (0, 0), (-1, 0), 1, BLACK),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(epi_table)
    elements.append(Spacer(1, 8))

    # ========== KIT DE EMERGÊNCIA (texto de referência fixo + validades preenchidas) ==========
    kit_text = (
        "Calços, com dimensões mínimas de 150mm x 200mm x 150mm (2 unid. em caminhão ou caminhão-trator com "
        "semirreboque; 4 unid. em caminhão com reboque, bitrem, bitrenzão ou rodotrem); Jogo de ferramentas adequado "
        "para reparos em situações de emergência (alicate universal, chave de fenda ou Philips, chave para desconexão "
        "do cabo da bateria); Triângulo ou equipamento similar reflexivo; 1 extintor de incêndio de Pó Químico ou "
        "Gás Carbônico de 2 kg."
    )
    validity_line = f"Validades: {fmt_date(checklist.get('kit_validity_1'))}, {fmt_date(checklist.get('kit_validity_2'))} e {fmt_date(checklist.get('kit_validity_3'))}."
    kit_table = Table([
        [Paragraph("KIT DE EMERGÊNCIA", static_bold_style)],
        [Paragraph(kit_text, static_style)],
        [Paragraph(f"<b>{validity_line}</b>", static_style)],
    ], colWidths=[SECTION_WIDTH])
    kit_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#F0F0F0')),
        ('BOX', (0, 0), (-1, -1), 1, BLACK),
        ('LINEBELOW', (0, 0), (-1, 0), 1, BLACK),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(kit_table)
    elements.append(Spacer(1, 8))

    elements.append(Paragraph(
        "<b>Aparelhos eletrônicos (celular, bip, nextel) OBRIGATORIAMENTE DEVEM ESTAR DESLIGADOS na Base de Carregamento.</b>",
        ParagraphStyle('LVTPhoneWarn', parent=styles['Normal'], fontSize=8, fontName='Helvetica-Bold', alignment=TA_CENTER)
    ))
    elements.append(Spacer(1, 8))

    # ========== REGISTRO DE OBSERVAÇÕES ==========
    if checklist.get('observations'):
        obs_table = Table([[Paragraph("REGISTRO DE OBSERVAÇÕES", static_bold_style)], [Paragraph(checklist['observations'], static_style)]], colWidths=[SECTION_WIDTH])
        obs_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#F0F0F0')),
            ('BOX', (0, 0), (-1, -1), 1, BLACK),
            ('LINEBELOW', (0, 0), (-1, 0), 1, BLACK),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(obs_table)
        elements.append(Spacer(1, 8))

    # ========== DECLARAÇÕES (texto padrão do documento) ==========
    trips = ", ".join(filter(None, [checklist.get('last_trip_product_1'), checklist.get('last_trip_product_2'), checklist.get('last_trip_product_3')])) or '-'
    expedidor_decl = (
        "<b>EXPEDIDOR/TRANSPORTADOR:</b> 1 - Declaramos que o veículo acima caracterizado foi inspecionado e que "
        "neste momento encontra-se em perfeito estado de conservação, que a documentação exigida foi entregue, que "
        "foram informados os riscos e características do(s) produto(s) a ser(em) transportado(s), que as embalagens "
        "atendem às legislações, que o veículo está apto ao transporte e que estão colocados corretamente os rótulos "
        "de risco e painéis de segurança. "
        "2 - Declaramos que o veículo está com as válvulas de descarregamento na posição fechada e em pleno "
        "funcionamento. "
        f"3 - Declaramos que as últimas 3 viagens do veículo foram com os seguintes produtos: {trips}."
    )
    condutor_decl = (
        "<b>CONDUTOR:</b> 1 - Declaro que o veículo acima caracterizado foi inspecionado pelo transportador/expedidor "
        "e que neste momento encontra-se em perfeito estado de conservação, que a documentação exigida foi recebida, "
        "que foram recebidas as informações sobre os riscos e características do(s) produto(s) e que todos os "
        "documentos, identificação e equipamentos permanecerão no veículo até o destino final da carga. "
        "2 - Declaro que cumpri o descanso previsto na legislação."
    )
    decl2_table = Table([[Paragraph(expedidor_decl, static_style)], [Paragraph(condutor_decl, static_style)]], colWidths=[SECTION_WIDTH])
    decl2_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, BLACK),
        ('LINEBELOW', (0, 0), (-1, 0), 0.5, BLACK),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(decl2_table)
    elements.append(Spacer(1, 8))

    # ========== RESPONSÁVEIS / ASSINATURAS ==========
    def info_pair(label, value):
        return [Paragraph(label, label_style), Paragraph(str(value) if value not in (None, '') else '-', value_style)]

    resp_rows = [
        [info_pair("Responsável da Transportadora", checklist.get('transport_responsible_name')), info_pair("RG", checklist.get('transport_responsible_rg'))],
        [info_pair("Recebedor da LVT", checklist.get('lvt_receiver_name')), info_pair("Matrícula", checklist.get('lvt_receiver_registration'))],
        [info_pair("Responsável pela Vistoria", checklist.get('inspection_responsible_name')), info_pair("Matrícula", checklist.get('inspection_responsible_registration'))],
        [info_pair("Registro de Mérito", checklist.get('merit_record')), info_pair("Registro de Ocorrências", checklist.get('occurrence_record'))],
        [info_pair("Data e Horário da Liberação do Veículo", fmt_datetime(checklist.get('release_datetime'))), info_pair("Documento do Condutor", checklist.get('driver_document'))],
    ]
    resp_table = Table(resp_rows, colWidths=[SECTION_WIDTH / 2, SECTION_WIDTH / 2])
    resp_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, BLACK),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(resp_table)
    elements.append(Spacer(1, 10))

    # ========== RODAPÉ DE ADVERTÊNCIA ==========
    footer_warn = Table([[Paragraph("NÃO SERÁ ACEITO NENHUM TIPO DE RASURA NESTE DOCUMENTO", ParagraphStyle('LVTFooterWarn', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold', alignment=TA_CENTER))]], colWidths=[SECTION_WIDTH])
    footer_warn.setStyle(TableStyle([('BOX', (0, 0), (-1, -1), 1, BLACK), ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5)]))
    elements.append(footer_warn)

    footer = _make_pdf_footer(c['name'])
    doc.build(elements, onFirstPage=footer, onLaterPages=footer)
    return buffer.getvalue()
